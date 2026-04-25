from fastapi import FastAPI, Query
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import JSONResponse, StreamingResponse
import pandas as pd
import re, math, json, glob, os, io, urllib.request
from typing import Optional, List
import uvicorn
import requests
import webbrowser
import threading
import shutil
from fastapi.staticfiles import StaticFiles
from fastapi.responses import FileResponse

app = FastAPI(title="Housing Dashboard API")
app.add_middleware(CORSMiddleware, allow_origins=["*"], allow_methods=["*"], allow_headers=["*"])

# Path to your build
FRONTEND_DIR = os.path.join(os.path.dirname(__file__), "frontend")

# Serve static assets (JS, CSS, images)
app.mount(
    "/spain_new_frontend/assets",
    StaticFiles(directory=os.path.join(FRONTEND_DIR, "assets")),
    name="assets"
)

# Serve index
@app.get("/spain_new_frontend")
@app.get("/spain_new_frontend/")
def serve_frontend():
    return FileResponse(os.path.join(FRONTEND_DIR, "index.html"))

# React routing fallback
@app.get("/spain_new_frontend/{full_path:path}")
def serve_frontend_routes(full_path: str):
    return FileResponse(os.path.join(FRONTEND_DIR, "index.html"))


MONTH_ORDER = {"Jan":1,"Feb":2,"Mar":3,"Apr":4,"May":5,"Jun":6,
               "Jul":7,"Aug":8,"Sep":9,"Oct":10,"Nov":11,"Dec":12}

def _clean(obj):
    if isinstance(obj, list):  return [_clean(i) for i in obj]
    if isinstance(obj, dict):  return {k: _clean(v) for k, v in obj.items()}
    if isinstance(obj, float) and (math.isnan(obj) or math.isinf(obj)): return None
    # pandas nullable integers/booleans (Int64, boolean dtype) → unwrap via .item()
    if hasattr(obj, 'item'):
        try: return _clean(obj.item())
        except: pass
    # pandas NA / NaT / NaN scalar
    try:
        if pd.isna(obj): return None
    except (TypeError, ValueError): pass
    # numpy arrays
    if hasattr(obj, 'tolist'): return _clean(obj.tolist())
    return obj

def safe_json(data):
    return JSONResponse(content=_clean(data))

# ── load & enrich ──────────────────────────────────────────────────────────
# Auto-discover all .xlsx files in the data/ folder (drop any into data/ to add)
# Tries multiple base directories so it works regardless of where uvicorn is launched from
_SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
_CWD        = os.getcwd()
_CANDIDATES = [
    os.path.join(_SCRIPT_DIR, "data"),   # relative to main.py (most reliable)
    os.path.join(_CWD,        "data"),   # relative to working dir
    _SCRIPT_DIR,                          # legacy: xlsx files next to main.py
    _CWD,                                 # legacy: xlsx files in working dir
]
_EXCLUDED_XLSX = {"expired_listing.xlsx"}
_xlsx_files = []
_DATA_DIR   = None
for _candidate in _CANDIDATES:
    _found = sorted(f for f in glob.glob(os.path.join(_candidate, "*.xlsx"))
                    if os.path.basename(f) not in _EXCLUDED_XLSX)
    if _found:
        _xlsx_files = _found
        _DATA_DIR   = _candidate
        break

if not _xlsx_files:
    raise RuntimeError(
        f"No .xlsx files found. Tried: {_CANDIDATES}\n"
        f"Create a 'data/' folder next to main.py and drop your .xlsx files there."
    )

print(f"[data] Data dir : {_DATA_DIR}")
print(f"[data] Loading  : {[os.path.basename(f) for f in _xlsx_files]}")
_raw = pd.concat([pd.read_excel(f) for f in _xlsx_files], ignore_index=True)
print(f"[data] Rows: {len(_raw):,}  |  provinces: {sorted(_raw['province'].dropna().unique().tolist())}")

# Load expired_listing sheet: sub_listing → removed_date
_EXPIRED_FILE = os.path.join(_DATA_DIR, "expired_listing.xlsx")
_expired_df   = pd.DataFrame()
if os.path.exists(_EXPIRED_FILE):
    try:
        _expired_df = pd.read_excel(_EXPIRED_FILE)
        # Normalise column names
        _expired_df.columns = [c.strip().lower().replace(" ","_") for c in _expired_df.columns]
        # Parse removed_date → formatted string "D Mon YYYY"
        _expired_df["removed_date"] = pd.to_datetime(_expired_df["removed_date"], dayfirst=True, errors="coerce")
        _expired_df["removed_date_str"] = _expired_df["removed_date"].apply(
            lambda d: f"{d.day} {d.strftime('%b')} {d.year}" if pd.notna(d) else None)
        print(f"[data] Expired listings loaded: {len(_expired_df):,} rows")
    except Exception as e:
        print(f"[data] Could not load expired_listing.xlsx: {e}")

# Build dicts keyed by sub_listing_id
_sub_to_sold_date: dict = {}
if not _expired_df.empty and "sub_listing" in _expired_df.columns:
    for _, row in _expired_df.iterrows():
        sid = row.get("sub_listing")
        dt  = row.get("removed_date_str")
        if pd.notna(sid) and dt:
            _sub_to_sold_date[int(sid)] = dt

# Build listing_id → set of expired sub_listing_ids (populated after df is built below)
# Will be computed after df is ready; see _build_listing_expired_counts()
_listing_expired_counts: dict = {}   # listing_id → int (sold/expired sub-listing count)

# Load geocoded streets CSV if present
_STREETS_CSV = os.path.join(_DATA_DIR, "combined_geocoded.csv")
_GEOCODED_STREETS = pd.DataFrame()
if os.path.exists(_STREETS_CSV):
    try:
        _GEOCODED_STREETS = pd.read_csv(_STREETS_CSV)
        # Normalise street names: clean encoding issues
        _GEOCODED_STREETS["street"] = _GEOCODED_STREETS["street"].apply(
            lambda s: str(s).encode("latin1").decode("utf-8", errors="replace") if isinstance(s, str) else s
        )
        print(f"[data] Geocoded streets loaded: {len(_GEOCODED_STREETS):,}")
    except Exception as e:
        print(f"[data] Could not load geocoded streets: {e}")

# Canonical snapshot label: "Feb 2026"
_raw["period"] = _raw["Month"].astype(str) + " " + _raw["Year"].astype(str)
_raw["period_ord"] = _raw["Month"].map(MONTH_ORDER) + (_raw["Year"].astype(int)-2000)*100

# Normalise municipality: take last segment after comma  e.g. "La Conarda, Betera" -> "Betera"
def _clean_municipality(s):
    if pd.isna(s): return s
    s = str(s).strip()
    return s.split(",")[-1].strip() if "," in s else s

_raw["municipality"] = _raw["municipality"].apply(_clean_municipality)

# Normalise province to Title Case  e.g. "alicante" -> "Alicante", "Valencia" stays "Valencia"
_raw["province"] = _raw["province"].apply(lambda s: str(s).strip().title() if pd.notna(s) else s)

# Full df (with duplicates) used only for snapshot-aware queries; deduplicated per sub+period for unit data
_full = _raw.copy()
df = _raw.drop_duplicates(subset=["sub_listing_id","period"]).copy()

# ── Normalise property names/developers to the latest-period value ─────────
# The same listing_id can have slightly different property_name or developer
# strings across months (typos, rebranding, etc.).  Using the raw name as a
# groupby key creates phantom duplicate rows.  Map everything to the canonical
# latest-period value so all endpoints are consistent.
_latest_rows = df[df["period"] == df.groupby("listing_id")["period_ord"].transform("max")]
for _col in ["property_name", "developer"]:
    if _col in df.columns:
        _canon = (
            _latest_rows.drop_duplicates("listing_id")
            .set_index("listing_id")[_col]
            .to_dict()
        )
        df[_col] = df["listing_id"].map(_canon).fillna(df[_col])
        _raw[_col] = _raw["listing_id"].map(_canon).fillna(_raw[_col])
print("[data] Property names/developers normalised to latest-period values")

PERIODS_SORTED = sorted(df["period"].unique(), key=lambda p: df[df["period"]==p]["period_ord"].iloc[0])
LATEST_PERIOD  = "Apr 2026"   # Display label override — actual data period is PERIODS_SORTED[-1]
_LATEST_DATA_PERIOD = PERIODS_SORTED[-1]   # Used for data queries
PREV_PERIOD    = PERIODS_SORTED[-2] if len(PERIODS_SORTED) > 1 else None

# Listing IDs new in the latest period (not present in any prior period)
_latest_listing_ids = set(df[df["period"] == _LATEST_DATA_PERIOD]["listing_id"].unique()) if _LATEST_DATA_PERIOD else set()
_prev_listing_ids   = set(df[df["period"] != _LATEST_DATA_PERIOD]["listing_id"].unique()) if _LATEST_DATA_PERIOD else set()
_new_this_month_ids = list(int(x) for x in (_latest_listing_ids - _prev_listing_ids))

# Per-province latest period (provinces may have different update cadences)
_prov_latest = (
    df.groupby("province")["period_ord"]
    .max()
    .reset_index()
    .rename(columns={"period_ord": "max_ord"})
)
_prov_latest = dict(zip(_prov_latest["province"], _prov_latest["max_ord"]))

# Mark each row: is it the latest snapshot for its province?
df["_is_latest"] = df.apply(lambda r: r["period_ord"] == _prov_latest.get(r["province"], -1), axis=1)
_full["_is_latest"] = _full.apply(lambda r: r["period_ord"] == _prov_latest.get(r["province"], -1), axis=1)

def _latest_df(df_src=None):
    """Return rows that represent the latest period for each province."""
    d = df_src if df_src is not None else df
    return d[d["_is_latest"]]

# Global set of active sub_listing_ids (latest period, all municipalities).
# A sub-listing is "sold" only if its ID does not appear here.
_global_active_sub_ids: set = set(
    df[df["_is_latest"]]["sub_listing_id"].dropna().astype(int).unique()
)

# Build per-listing expired sub-listing count from expired_listing.xlsx
# A sub-listing is "sold" if it appears in _sub_to_sold_date AND is NOT in the latest period
def _build_listing_expired_counts():
    if not _sub_to_sold_date:
        return {}
    expired_sub_ids = set(_sub_to_sold_date.keys())
    # Active sub-listing IDs (latest period) — these are NOT sold even if in expired file
    active_sub_ids = set(df[df["_is_latest"]]["sub_listing_id"].dropna().astype(int).unique())
    truly_expired = expired_sub_ids - active_sub_ids
    if not truly_expired:
        return {}
    # Map sub_listing_id → listing_id from the full df
    sub_to_lid = (
        df[df["sub_listing_id"].isin(truly_expired)][["sub_listing_id","listing_id"]]
        .drop_duplicates("sub_listing_id")
        .set_index("sub_listing_id")["listing_id"]
        .astype(int)
        .to_dict()
    )
    counts = {}
    for sub_id in truly_expired:
        lid = sub_to_lid.get(sub_id)
        if lid is not None:
            counts[lid] = counts.get(lid, 0) + 1
    return counts

_listing_expired_counts = _build_listing_expired_counts()
print(f"[data] Listings with expired/sold sub-listings: {len(_listing_expired_counts)}")

_UNIT_COUNT_PATTERNS = [
    re.compile(r'development of\s+(\d+)\s+(?:homes?|apartments?|properties|residences?|units?|dwellings?|viviendas?)', re.I),
    re.compile(r'consisting of\s+(\d+)\s+(?:homes?|apartments?|properties|residences?|units?|dwellings?)', re.I),
    re.compile(r'composed of\s+(\d+)\s+(?:homes?|apartments?|properties|residences?|units?|dwellings?)', re.I),
    re.compile(r'made up of\s+(\d+)\s+(?:homes?|apartments?|properties|residences?|units?|dwellings?)', re.I),
    re.compile(r'consists of\s+(\d+)\s+(?:homes?|apartments?|properties|residences?|units?|dwellings?)', re.I),
    re.compile(r'has\s+(\d+)\s+(?:homes?|apartments?|residences?|bright homes?|multi-family)', re.I),
    re.compile(r'(\d+)\s+(?:multi-family homes?|single-family homes?|townhouses?)', re.I),
    re.compile(r'(\d+)\s+\w+\s+(?:multi-family homes?|single-family homes?|townhouses?)', re.I),
    re.compile(r'(\d+)\s+(?:\w+\s+){1,2}(?:homes?|apartments?|villas?|properties|residences?)\s+(?:with|and|in|distributed|for)\b', re.I),
    re.compile(r'(\d+)\s+(?:exclusive|luxury|modern|new|bright|independent|private|unique|detached|semi-detached|terraced)\s+(?:homes?|apartments?|villas?|properties|residences?)', re.I),
    re.compile(r'(\d+)\s+(?:homes?|apartments?)\s+(?:with|and|in|distributed|for)\b', re.I),
    re.compile(r'(?:^|[\s,])(\d+)\s+homes?(?=[\s,\.]|$)', re.I),
    re.compile(r'(?:^|[\s,])(\d+)\s+apartments?(?=[\s,\.]|$)', re.I),
    re.compile(r'(\d+)\s+(?:viviendas?|pisos?|apartamentos?)\s', re.I),
]

def _extract_stated_units(description: str):
    if not description or not isinstance(description, str):
        return None
    for pat in _UNIT_COUNT_PATTERNS:
        m = pat.search(description)
        if m:
            n = int(m.group(1))
            if 2 <= n <= 2000:
                return n
    return None

# ── Precompute stated total units from descriptions (done once at startup) ──
_desc_col = next((c for c in ["description","property_description","descripcion","desc","comments"]
                  if c in df.columns), None)
if _desc_col:
    _latest_listings = _latest_df().drop_duplicates(subset=["listing_id"])[["listing_id", _desc_col]].copy()
    _latest_listings["_stated"] = _latest_listings[_desc_col].apply(
        lambda v: _extract_stated_units(str(v)) if pd.notna(v) else None)
    _STATED_TOTAL_UNITS = int(_latest_listings["_stated"].dropna().sum())
    _STATED_UNIT_COUNT  = int(_latest_listings["_stated"].notna().sum())
    print(f"[data] Stated total units (from descriptions): {_STATED_TOTAL_UNITS:,} across {_STATED_UNIT_COUNT:,} listings")
else:
    _latest_listings = pd.DataFrame(columns=["listing_id", "_stated"])
    _STATED_TOTAL_UNITS = 0
    _STATED_UNIT_COUNT  = 0

_STRIP_LINES = [
    "This comment was automatically translated and may not be 100% accurate.",
    "See description in the original language",
]
_DESC_COLS = ["description","property_description","descripcion","desc","comments"]

def _clean_description(text):
    """Strip disclaimer lines and return cleaned description or None."""
    if not text or not isinstance(text, str) or text.strip().lower() in ("nan","none",""):
        return None
    lines = [l for l in text.splitlines() if not any(s.lower() in l.lower() for s in _STRIP_LINES)]
    result = " ".join(l.strip() for l in lines if l.strip())
    return result if result else None

def _first_desc(row, columns):
    """Get first non-null description from a row, cleaned."""
    for c in _DESC_COLS:
        if c in columns and pd.notna(row.get(c)):
            return _clean_description(str(row[c]))
    return None

def _year(s):
    if pd.isna(s): return None
    m = re.search(r"(\d{4})", str(s))
    return int(m.group(1)) if m else None

def _quarter(s):
    if pd.isna(s): return None
    s2, year = str(s).lower(), _year(s)
    if year is None: return None
    if "immediate" in s2: return f"{year} Q1"
    for q, kw in [("Q1",["january","february","march","first quarter","first semester"]),
                  ("Q2",["april","may","june","second quarter","second semester"]),
                  ("Q3",["july","august","september","third quarter"]),
                  ("Q4",["october","november","december","fourth quarter"])]:
        if any(k in s2 for k in kw): return f"{year} {q}"
    return f"{year} Q2"

def _esg(s):
    if pd.isna(s): return None
    m = re.search(r"Consumption: ([A-E])", str(s))
    return m.group(1) if m else None

def _floor_num(f):
    if pd.isna(f) or str(f).strip() in ["-","ext."]: return None
    if "ground" in str(f).lower(): return 0
    m = re.search(r"Floor (\d+)", str(f), re.I)
    return int(m.group(1)) if m else None

def _parse_amenities(s):
    _s_raw = s
    if pd.isna(s): s = ""
    s = str(s)
    m_bed  = re.search(r"(\d+)\s+bedroom", s)
    m_bath = re.search(r"(\d+)\s+bathroom", s)
    m_fa   = re.search(r"(\d+)\s+m.*?floor area", s)
    sl = s.lower()
    # If amenities is missing/empty, house_type is unknown (None) — don't default to Apartments
    if not sl.strip():
        ht = None
    elif "semi-detached" in sl or "semidetached" in sl or "semi detached" in sl:
        ht = "Semi-detached house"
    elif "detached house" in sl:
        ht = "Detached house"
    elif "terraced house" in sl:
        ht = "Terraced house"
    elif "flat" in sl or "apartment" in sl:
        ht = "Apartments"
    else:
        ht = "Apartments"
    return {"bedrooms": int(m_bed.group(1)) if m_bed else (0 if "No bedroom" in s else None),
            "bathrooms": int(m_bath.group(1)) if m_bath else None,
            "floor_area_m2": int(m_fa.group(1)) if m_fa else None,
            "has_terrace":  "Terrace" in s, "has_parking": "Parking" in s,
            "has_pool": "Swimming pool" in s, "has_garden": "Garden" in s,
            "has_lift": "lift" in s.lower(), "has_ac": "Air conditioning" in s,
            "has_storage": "Storage room" in s, "has_wardrobes": "wardrobe" in s.lower(),
            "house_type": ht}

for _d in [_raw, df, _full]:
    _d["delivery_year"]    = _d["delivery_date"].apply(_year)
    _d["delivery_quarter"] = _d["delivery_date"].apply(_quarter)
    _d["esg_grade"]        = _d["esg_certificate"].apply(_esg)
    _d["floor_num"]        = _d["floor"].apply(_floor_num)
    _am = _d["amenities"].apply(_parse_amenities).apply(pd.Series)
    for col in _am.columns:
        _d[col] = _am[col]

# Property classification — from a 'property_type' column if present, else derived from unit_type
_PROP_TYPE_COL = next((c for c in ["property_type","tipo_propiedad","property_class","type"] if c in df.columns), None)
_UT_TO_CLASS = {t: "Apartment" for t in ["Studio","1BR","2BR","3BR","4BR","5BR","Penthouse"]}

def _get_property_class(row):
    if _PROP_TYPE_COL and pd.notna(row.get(_PROP_TYPE_COL)):
        raw = str(row[_PROP_TYPE_COL]).strip().lower()
        if any(x in raw for x in ["semi","pareado"]): return "Semi-detached"
        if any(x in raw for x in ["detached","chalet","villa","casa","house"]): return "Detached"
        if any(x in raw for x in ["terrace","terraced","adosado"]): return "Terraced"
        if any(x in raw for x in ["country","rural","finca","cortijo"]): return "Country House"
        if any(x in raw for x in ["apartment","flat","piso","duplex"]): return "Apartment"
    return _UT_TO_CLASS.get(str(row.get("unit_type","")), "Apartment")

for _d in [df, _full]:
    _d["property_class"] = _d.apply(_get_property_class, axis=1)

def _filter(municipality=None, unit_type=None, year=None, esg=None, period=None, province=None, df_src=None, house_type=None, all_periods=False):
    base = df_src if df_src is not None else df
    # When no period specified, use per-province latest (so all provinces show even if not in sync)
    if period:
        d = base[base["period"].isin(period)].copy()
    elif all_periods:
        d = base.copy()
    else:
        d = _latest_df(base).copy()
    if province:     d = d[d["province"].isin(province)]
    if municipality: d = d[d["municipality"].isin(municipality)]
    if unit_type:    d = d[d["unit_type"].isin(unit_type)]
    if year:         d = d[d["delivery_year"].isin([int(y) for y in year])]
    if esg:          d = d[d["esg_grade"].isin(esg)]
    if house_type and "house_type" in d.columns:
        d = d[d["house_type"].isin(house_type)]
    return d

# ══════════════════════════════════════════════════════════════════════════
#  META
# ══════════════════════════════════════════════════════════════════════════
@app.get("/data-sources")
def get_data_sources():
    """Returns list of loaded data files and row counts per province."""
    sources = []
    for f in _xlsx_files:
        name = os.path.basename(f)
        province_name = name.replace("_all_units.xlsx","").replace("_"," ").title()
        count = int(df[df["province"] == province_name.lower()]["sub_listing_id"].nunique()) \
                if province_name.lower() in df["province"].str.lower().values else None
        sources.append({"file": name, "province": province_name})
    return safe_json({"sources": sources, "total_rows": len(df), "provinces": sorted(df["province"].dropna().unique().tolist())})


@app.get("/filters")
def get_filters():
    # Build province -> municipalities mapping
    province_munis = {}
    for _, row in df[["province","municipality"]].drop_duplicates().iterrows():
        p = str(row["province"]) if pd.notna(row["province"]) else "Other"
        m = str(row["municipality"]) if pd.notna(row["municipality"]) else None
        if m:
            province_munis.setdefault(p, [])
            if m not in province_munis[p]:
                province_munis[p].append(m)
    for p in province_munis:
        province_munis[p] = sorted(province_munis[p])

    house_types = sorted([h for h in df["house_type"].dropna().unique().tolist() if h]) if "house_type" in df.columns else []
    return {"municipalities": sorted(df["municipality"].dropna().unique().tolist()),
            "provinces":      sorted(df["province"].dropna().unique().tolist()),
            "province_munis": province_munis,
            "unit_types":     sorted(df["unit_type"].dropna().unique().tolist()),
            "house_types":    house_types,
            "delivery_years": sorted([int(y) for y in df["delivery_year"].dropna().unique()]),
            "esg_grades":     sorted(df["esg_grade"].dropna().unique().tolist()),
            "periods":              PERIODS_SORTED,
            "latest_period":        _LATEST_DATA_PERIOD,
            "prev_period":          PREV_PERIOD,
            "new_this_month_ids":   _new_this_month_ids}

# ══════════════════════════════════════════════════════════════════════════
#  SUMMARY / SNAPSHOT  (latest period by default)
# ══════════════════════════════════════════════════════════════════════════
@app.get("/stats")
def get_stats(municipality: Optional[List[str]] = Query(None),
              province:     Optional[List[str]] = Query(None),
              unit_type:    Optional[List[str]] = Query(None),
              year:         Optional[List[str]] = Query(None),
              esg:          Optional[List[str]] = Query(None),
              period:       Optional[List[str]] = Query(None),
              house_type:   Optional[List[str]] = Query(None)):
    d = _filter(municipality, unit_type, year, esg, period, province, house_type=house_type)
    p = _filter(municipality, unit_type, year, esg, [PREV_PERIOD], province, house_type=house_type) if PREV_PERIOD else None
    def _s(d): return {"total_units": len(d),
                       "avg_price":    round(float(d["price"].mean()))   if len(d) else 0,
                       "avg_price_m2": round(float(d["price_per_m2"].mean()),1) if len(d) else 0,
                       "avg_size":     round(float(d["size"].mean()),1)  if len(d) else 0,
                       "total_developments": int(d["listing_id"].nunique())}
    cur = _s(d)
    _lids_in_d = set(d["listing_id"].unique())
    _sl = _latest_listings[_latest_listings["listing_id"].isin(_lids_in_d)]
    cur["total_stated_units"] = int(_sl["_stated"].dropna().sum())
    cur["stated_unit_count"]  = int(_sl["_stated"].notna().sum())
    cur["prev"] = _s(p) if p is not None else None
    cur["prev_period"] = PREV_PERIOD
    cur["new_this_month"] = int(len(set(d["listing_id"].unique()) & set(_new_this_month_ids)))
    return cur

@app.get("/charts/price-by-unit-type")
def price_by_unit_type(municipality: Optional[List[str]] = Query(None),
                       province:     Optional[List[str]] = Query(None),
                       unit_type:    Optional[List[str]] = Query(None),
                       year:         Optional[List[str]] = Query(None),
                       esg:          Optional[List[str]] = Query(None),
                       period:       Optional[List[str]] = Query(None),
                       house_type:   Optional[List[str]] = Query(None)):
    d = _filter(municipality, unit_type, year, esg, period, province, house_type=house_type)
    r = d.groupby("unit_type").agg(
        avg_price=("price","mean"), min_price=("price","min"), max_price=("price","max"),
        count=("sub_listing_id","nunique"), avg_size=("size","mean"), avg_price_m2=("price_per_m2","mean")
    ).reset_index()
    r["avg_price"] = r["avg_price"].round(0); r["min_price"] = r["min_price"].round(0); r["max_price"] = r["max_price"].round(0)
    r["avg_size"] = r["avg_size"].round(1); r["avg_price_m2"] = r["avg_price_m2"].round(0)
    order = ["Studio","1BR","2BR","3BR","4BR","5BR","Penthouse"]
    r["_s"] = r["unit_type"].apply(lambda x: order.index(x) if x in order else 99)
    return safe_json(r.sort_values("_s").drop("_s",axis=1).to_dict(orient="records"))

@app.get("/charts/delivery-timeline")
def delivery_timeline(municipality: Optional[List[str]] = Query(None),
                      province:     Optional[List[str]] = Query(None),
                      unit_type:    Optional[List[str]] = Query(None),
                      year:         Optional[List[str]] = Query(None),
                      esg:          Optional[List[str]] = Query(None),
                      period:       Optional[List[str]] = Query(None),
                      house_type:   Optional[List[str]] = Query(None)):
    d = _filter(municipality, unit_type, year, esg, period, province, house_type=house_type).dropna(subset=["delivery_quarter"])
    r = d.groupby("delivery_quarter").agg(count=("price","count"), avg_price=("price","mean")).reset_index()
    r["avg_price"] = r["avg_price"].round(0)
    return safe_json(r.sort_values("delivery_quarter").to_dict(orient="records"))

@app.get("/charts/price-distribution")
def price_distribution(municipality: Optional[List[str]] = Query(None),
                       province:     Optional[List[str]] = Query(None),
                       unit_type:    Optional[List[str]] = Query(None),
                       year:         Optional[List[str]] = Query(None),
                       esg:          Optional[List[str]] = Query(None),
                       period:       Optional[List[str]] = Query(None),
                       house_type:   Optional[List[str]] = Query(None)):
    d = _filter(municipality, unit_type, year, esg, period, province, house_type=house_type)
    bins   = [0,150000,200000,250000,300000,400000,500000,700000,10000000]
    labels = ["<150k","150-200k","200-250k","250-300k","300-400k","400-500k","500-700k",">700k"]
    d2 = d.copy(); d2["bin"] = pd.cut(d2["price"], bins=bins, labels=labels)
    result = d2.groupby("bin", observed=True).agg(count=("price","size")).reset_index()
    m2_bins   = [0,1000,1500,2000,2500,3000,3500,4000,5000,6000,100000]
    m2_labels = ["<1k","1-1.5k","1.5-2k","2-2.5k","2.5-3k","3-3.5k","3.5-4k","4-5k","5-6k",">6k"]
    d3 = d[d["price_per_m2"].notna() & (d["price_per_m2"] > 0)].copy()
    d3["bin"] = pd.cut(d3["price_per_m2"], bins=m2_bins, labels=m2_labels)
    m2_result = d3.groupby("bin", observed=True).agg(count=("price_per_m2","size")).reset_index()
    return safe_json({"price_dist": result.to_dict(orient="records"), "m2_dist": m2_result.to_dict(orient="records")})

@app.get("/charts/municipality-overview")
def municipality_overview(municipality: Optional[List[str]] = Query(None),
                          province:     Optional[List[str]] = Query(None),
                          unit_type:    Optional[List[str]] = Query(None),
                          year:         Optional[List[str]] = Query(None),
                          esg:          Optional[List[str]] = Query(None),
                          period:       Optional[List[str]] = Query(None),
                          house_type:   Optional[List[str]] = Query(None)):
    # Active listings — use _filter() exactly as the original endpoint did
    d_active = _filter(municipality, unit_type, year, esg, period, province, house_type=house_type)

    # Unit count per municipality (active sub-listings)
    unit_counts = (
        d_active.groupby("municipality")
        .agg(units=("price", "count"), listings=("listing_id", "nunique"))
        .reset_index()
    )

    # Per-listing averages first, then average across listings per municipality
    # This matches the detail page which uses mean-of-listing-averages (development-weighted)
    listing_avg = (
        d_active.groupby(["municipality", "listing_id"])
        .agg(avg_price=("price", "mean"), avg_price_m2=("price_per_m2", "mean"))
        .reset_index()
    )
    muni_avg = (
        listing_avg.groupby("municipality")
        .agg(avg_price=("avg_price", "mean"), avg_price_m2=("avg_price_m2", "mean"))
        .reset_index()
    )

    r = unit_counts.merge(muni_avg, on="municipality", how="left")
    r["avg_price"]    = r["avg_price"].round(0)
    r["avg_price_m2"] = r["avg_price_m2"].round(1)
    return safe_json(r.sort_values("units", ascending=False).to_dict(orient="records"))

@app.get("/debug/municipality-overview")
def debug_municipality_overview(muni: str = "Valencia"):
    """Debug: show active vs sold breakdown for a single municipality."""
    d_active = _filter()
    d_muni_active = d_active[d_active["municipality"] == muni]
    active_count = int(d_muni_active["price"].count())

    _active_lid_set = set(d_active["listing_id"].astype(int).tolist())
    d_all  = _filter(all_periods=True)
    d_hist = d_all[
        ~d_all["_is_latest"] &
        d_all["sub_listing_id"].notna() &
        d_all["listing_id"].astype(int).isin(_active_lid_set)
    ].copy()
    d_hist["_sid_i"] = d_hist["sub_listing_id"].astype(int)
    d_sold = d_hist[~d_hist["_sid_i"].isin(_global_active_sub_ids)]
    sold_count = int(d_sold[d_sold["municipality"] == muni]["_sid_i"].nunique())

    periods_in_data = sorted(d_all[d_all["municipality"]==muni]["period"].unique().tolist())
    return safe_json({
        "municipality": muni,
        "active_units_latest_period": active_count,
        "sold_units_truly_removed": sold_count,
        "total_units": active_count + sold_count,
        "periods_in_data": periods_in_data,
    })

@app.get("/charts/municipality-activity")
def municipality_activity(province: Optional[List[str]] = Query(None),
                          municipality: Optional[List[str]] = Query(None)):
    """Top municipalities by new listings and by sold-out listings (latest period)."""
    d = df.copy()
    if province:     d = d[d["province"].isin(province)]
    if municipality: d = d[d["municipality"].isin(municipality)]

    # New listings: listing_ids present in latest period but not in any prior period
    new_ids = set(_new_this_month_ids)
    d_new = d[d["listing_id"].isin(new_ids) & d["_is_latest"]]
    new_by_muni = (
        d_new.groupby("municipality")
        .agg(listings=("listing_id","nunique"), units=("sub_listing_id","nunique"))
        .reset_index()
        .sort_values("units", ascending=False)
        .head(15)
    )

    # Sold-out THIS period: present in prev period but absent from latest period
    has_latest = set(d[d["_is_latest"]]["listing_id"].unique())
    if PREV_PERIOD:
        has_prev = set(d[d["period"] == PREV_PERIOD]["listing_id"].unique())
        sold_this_period = has_prev - has_latest
        d_sold = d[d["listing_id"].isin(sold_this_period) & (d["period"] == PREV_PERIOD)]
    else:
        d_sold = d.iloc[:0]  # empty
    sold_by_muni = (
        d_sold.groupby("municipality")["listing_id"]
        .nunique().reset_index(name="listings")
        .sort_values("listings", ascending=False)
        .head(15)
    )

    return safe_json({
        "new_listings":  new_by_muni.to_dict(orient="records"),
        "sold_out":      sold_by_muni.to_dict(orient="records"),
    })

@app.get("/charts/municipality-soldout-trend")
def municipality_soldout_trend(province: Optional[List[str]] = Query(None),
                               municipality: Optional[List[str]] = Query(None)):
    """Sold-out listings per municipality per period (consecutive period diffs)."""
    d = df.copy()
    if province:     d = d[d["province"].isin(province)]
    if municipality: d = d[d["municipality"].isin(municipality)]

    rows = []
    for i in range(1, len(PERIODS_SORTED)):
        prev_p = PERIODS_SORTED[i - 1]
        curr_p = PERIODS_SORTED[i]
        has_prev = set(d[d["period"] == prev_p]["listing_id"].unique())
        has_curr = set(d[d["period"] == curr_p]["listing_id"].unique())
        sold = has_prev - has_curr
        if not sold:
            continue
        d_sold = d[d["listing_id"].isin(sold) & (d["period"] == prev_p)]
        by_muni = d_sold.groupby("municipality")["listing_id"].nunique().reset_index(name="listings")
        for _, r in by_muni.iterrows():
            rows.append({"period": curr_p, "municipality": str(r["municipality"]), "listings": int(r["listings"])})

    return safe_json(rows)

@app.get("/charts/esg-breakdown")
def esg_breakdown(municipality: Optional[List[str]] = Query(None),
                  province:     Optional[List[str]] = Query(None),
                  unit_type:    Optional[List[str]] = Query(None),
                  year:         Optional[List[str]] = Query(None),
                  esg:          Optional[List[str]] = Query(None),
                  period:       Optional[List[str]] = Query(None),
                  house_type:   Optional[List[str]] = Query(None)):
    d = _filter(municipality, unit_type, year, esg, period, province, house_type=house_type)
    r = d.groupby("esg_grade", dropna=False).agg(count=("price","count"), avg_price=("price","mean")).reset_index()
    r["esg_grade"] = r["esg_grade"].fillna("Unknown")
    r["avg_price"] = r["avg_price"].round(0)
    return safe_json(r.to_dict(orient="records"))

@app.get("/charts/unit-by-house-type")
def unit_by_house_type(municipality: Optional[List[str]] = Query(None),
                       province:     Optional[List[str]] = Query(None),
                       unit_type:    Optional[List[str]] = Query(None),
                       year:         Optional[List[str]] = Query(None),
                       esg:          Optional[List[str]] = Query(None),
                       period:       Optional[List[str]] = Query(None),
                       house_type:   Optional[List[str]] = Query(None)):
    d = _filter(municipality, unit_type, year, esg, period, province, house_type=house_type)
    if "house_type" not in d.columns or "unit_type" not in d.columns:
        return safe_json([])
    d2 = d[d["house_type"].notna() & (d["house_type"] != "") &
           d["unit_type"].notna()  & (d["unit_type"]  != "")]
    counts = d2.groupby(["house_type","unit_type"])["sub_listing_id"].nunique().reset_index()
    counts.columns = ["house_type","unit_type","count"]
    return safe_json(counts.to_dict(orient="records"))

@app.get("/charts/size-vs-price")
def size_vs_price(municipality: Optional[List[str]] = Query(None),
                  province:     Optional[List[str]] = Query(None),
                  unit_type:    Optional[List[str]] = Query(None),
                  year:         Optional[List[str]] = Query(None),
                  esg:          Optional[List[str]] = Query(None),
                  period:       Optional[List[str]] = Query(None),
                  house_type:   Optional[List[str]] = Query(None)):
    d = _filter(municipality, unit_type, year, esg, period, province, house_type=house_type)
    cols = ["sub_listing_id","listing_id","size","price","price_per_m2",
            "unit_type","house_type","municipality","city_area","property_name","floor",
            "bedrooms","unit_url"]
    d2 = d[[c for c in cols if c in d.columns]].dropna(subset=["size","price"])
    if len(d2) > 800: d2 = d2.sample(800, random_state=42)
    rows = []
    for _, r in d2.iterrows():
        lat, lng, map_url = _listing_coords(r["listing_id"], str(r["municipality"]) if pd.notna(r.get("municipality")) else "")
        row = {c: r[c] for c in d2.columns}
        row["lat"] = lat or 39.47
        row["lng"] = lng or -0.38
        row["map_url"] = map_url
        for k in list(row.keys()):
            v = row[k]
            if hasattr(v, 'item'): row[k] = v.item()
            if isinstance(v, float) and v != v: row[k] = None
        rows.append(row)
    return safe_json(rows)

# ══════════════════════════════════════════════════════════════════════════
#  TEMPORAL  — market-wide month-over-month
# ══════════════════════════════════════════════════════════════════════════
@app.get("/temporal/market-trend")
def market_trend(municipality: Optional[List[str]] = Query(None),
                 province:     Optional[List[str]] = Query(None),
                 unit_type:    Optional[List[str]] = Query(None)):
    d = _filter(municipality, unit_type, province=province, df_src=_full, all_periods=True)
    r = d.groupby(["period","period_ord"]).agg(
        avg_price    =("price","mean"),
        avg_price_m2 =("price_per_m2","mean"),
        total_units  =("sub_listing_id","nunique"),
        avg_size     =("size","mean"),
    ).reset_index().sort_values("period_ord")
    r["avg_price"]    = r["avg_price"].round(0)
    r["avg_price_m2"] = r["avg_price_m2"].round(1)
    r["avg_size"]     = r["avg_size"].round(1)
    return safe_json(r.drop("period_ord",axis=1).to_dict(orient="records"))

@app.get("/temporal/unit-type-trend")
def unit_type_trend(municipality: Optional[List[str]] = Query(None),
                    province:     Optional[List[str]] = Query(None)):
    d = _filter(municipality, province=province, df_src=_full, all_periods=True)
    r = d.groupby(["period","period_ord","unit_type"]).agg(
        avg_price=("price","mean"), count=("sub_listing_id","nunique")
    ).reset_index().sort_values(["unit_type","period_ord"])
    r["avg_price"] = r["avg_price"].round(0)
    return safe_json(r.drop("period_ord",axis=1).to_dict(orient="records"))

@app.get("/temporal/municipality-trend")
def municipality_trend(municipality: Optional[List[str]] = Query(None),
                       province:     Optional[List[str]] = Query(None)):
    d = _filter(municipality, province=province, df_src=_full, all_periods=True) if municipality else _full
    r = d.groupby(["municipality","period","period_ord"]).agg(
        avg_price    =("price","mean"),
        avg_price_m2 =("price_per_m2","mean"),
        units        =("sub_listing_id","nunique"),
    ).reset_index().sort_values(["municipality","period_ord"])
    r["avg_price"]    = r["avg_price"].round(0)
    r["avg_price_m2"] = r["avg_price_m2"].round(1)
    return safe_json(r.drop("period_ord",axis=1).to_dict(orient="records"))

@app.get("/temporal/inventory-trend")
def inventory_trend(municipality: Optional[List[str]] = Query(None),
                    province:     Optional[List[str]] = Query(None),
                    unit_type:    Optional[List[str]] = Query(None)):
    """Units available per period, plus new/removed vs prior period."""
    d = _filter(municipality, unit_type, province=province, df_src=_full, all_periods=True)
    result = []
    prev_ids = None
    for period in PERIODS_SORTED:
        ids = set(d[d["period"]==period]["sub_listing_id"].unique())
        new     = len(ids - prev_ids) if prev_ids is not None else 0
        removed = len(prev_ids - ids) if prev_ids is not None else 0
        result.append({"period":period, "total":len(ids), "new":new, "removed":removed})
        prev_ids = ids
    return result

# ══════════════════════════════════════════════════════════════════════════
#  DRILL-DOWN — municipality
# ══════════════════════════════════════════════════════════════════════════
@app.get("/drilldown/municipality/{municipality:path}")
def drilldown_municipality(municipality: str):
    d = df[df["municipality"] == municipality]
    if d.empty:
        return safe_json({"listings":[],"stats":{},"unit_type_mix":[],"price_dist":[],"trend":[]})

    # latest period snapshot for listings (use per-province latest)
    dl = d[d["_is_latest"]]
    # Detect "Tourist apartment" in any description column
    _desc_col = next((c for c in ["description","property_description","descripcion","desc","comments"] if c in dl.columns), None)
    if _desc_col:
        dl = dl.copy()
        dl["_is_tourist"] = dl[_desc_col].fillna("").str.contains("tourist apartment", case=False, na=False)
    else:
        dl = dl.copy()
        dl["_is_tourist"] = False

    _dd_agg = dict(
        units        =("sub_listing_id","nunique"),
        min_price    =("price","min"), max_price=("price","max"),
        avg_price    =("price","mean"), avg_price_m2=("price_per_m2","mean"),
        avg_size     =("size","mean"),
        unit_types   =("unit_type",  lambda x: ", ".join(sorted(x.dropna().unique().tolist()))),
        house_types  =("house_type", lambda x: ", ".join(sorted(t for t in x.dropna().unique() if t))),
        has_pool=("has_pool","max"), has_parking=("has_parking","max"),
        has_terrace=("has_terrace","max"), has_lift=("has_lift","max"),
        is_tourist   =("_is_tourist","max"),
    )
    _dd_desc_col = next((c for c in _DESC_COLS if c in dl.columns), None)
    if _dd_desc_col:
        _dd_agg["_desc_raw"] = (_dd_desc_col, "first")
    listings_grp = dl.groupby(["listing_id","property_name","developer","delivery_date","esg_grade"], dropna=False).agg(**_dd_agg).reset_index()
    listings_grp = listings_grp[listings_grp["listing_id"].notna()].copy()
    listings_grp["listing_id"] = listings_grp["listing_id"].astype(int)
    # Drop listings whose only latest-period rows have null sub_listing_id (fully sold out)
    listings_grp = listings_grp[listings_grp["units"] > 0].copy()
    # Total unique sub-listings ever seen for each listing (across all periods)
    _all_units = d.groupby("listing_id")["sub_listing_id"].nunique().reset_index(name="total_units_ever")
    _all_units["listing_id"] = _all_units["listing_id"].astype(int)
    listings_grp = listings_grp.merge(_all_units, on="listing_id", how="left")
    _dd_cols = dl.columns.tolist()
    if "_desc_raw" in listings_grp.columns:
        listings_grp["stated_total_units"] = listings_grp["_desc_raw"].apply(
            lambda v: _extract_stated_units(str(v)) if pd.notna(v) else None)
        listings_grp = listings_grp.drop(columns=["_desc_raw"])
    else:
        listings_grp["stated_total_units"] = None
    for c in ["avg_price","min_price","max_price"]:
        listings_grp[c] = listings_grp[c].round(0)
    listings_grp["avg_price_m2"] = listings_grp["avg_price_m2"].round(1)
    listings_grp["avg_size"]     = listings_grp["avg_size"].round(1)
    listings_grp["esg_grade"]    = listings_grp["esg_grade"].where(pd.notna(listings_grp["esg_grade"]), None)

    _new_in_muni = int(len(set(dl["listing_id"].unique()) & set(_new_this_month_ids)))
    stats = {"total_units": int(dl["sub_listing_id"].nunique()),
             "total_listings": int(dl["listing_id"].nunique()),
             "avg_price":    round(float(dl["price"].mean())),
             "avg_price_m2": round(float(dl["price_per_m2"].mean()),1),
             "price_range":  [int(dl["price"].min()), int(dl["price"].max())],
             "new_this_month": _new_in_muni}

    mix = dl.groupby("unit_type").size().reset_index(name="count")
    order = ["Studio","1BR","2BR","3BR","4BR","5BR","Penthouse"]
    mix["_s"] = mix["unit_type"].apply(lambda x: order.index(x) if x in order else 99)
    mix = mix.sort_values("_s").drop("_s",axis=1)

    # Full per-type stats for summary table
    ut_stats = dl.groupby("unit_type").agg(
        count     =("sub_listing_id","nunique"),
        min_price =("price","min"),
        avg_price =("price","mean"),
        max_price =("price","max"),
        avg_size  =("size","mean"),
        avg_pm2   =("price_per_m2","mean"),
    ).reset_index()
    ut_stats["min_price"] = ut_stats["min_price"].round(0)
    ut_stats["avg_price"] = ut_stats["avg_price"].round(0)
    ut_stats["max_price"] = ut_stats["max_price"].round(0)
    ut_stats["avg_size"]  = ut_stats["avg_size"].round(1)
    ut_stats["avg_pm2"]   = ut_stats["avg_pm2"].round(0)
    ut_stats["_s"] = ut_stats["unit_type"].apply(lambda x: order.index(x) if x in order else 99)
    ut_stats = ut_stats.sort_values("_s").drop("_s",axis=1)

    bins   = [0,150000,200000,250000,300000,400000,500000,700000,10000000]
    labels = ["<150k","150-200k","200-250k","250-300k","300-400k","400-500k","500-700k",">700k"]
    d2 = dl.copy(); d2["bin"] = pd.cut(d2["price"], bins=bins, labels=labels)
    price_dist_grp = d2.groupby("bin", observed=True).agg(count=("price","size")).reset_index()
    price_dist = price_dist_grp

    # Separate €/m² distribution with proper m² bins
    m2_bins   = [0,1000,1500,2000,2500,3000,3500,4000,5000,6000,100000]
    m2_labels = ["<1k","1-1.5k","1.5-2k","2-2.5k","2.5-3k","3-3.5k","3.5-4k","4-5k","5-6k",">6k"]
    d3 = dl[dl["price_per_m2"].notna() & (dl["price_per_m2"] > 0)].copy()
    d3["bin"] = pd.cut(d3["price_per_m2"], bins=m2_bins, labels=m2_labels)
    m2_dist_grp = d3.groupby("bin", observed=True).agg(count=("price_per_m2","size")).reset_index()
    m2_dist = m2_dist_grp

    # month-over-month trend for this municipality
    trend = d.groupby(["period","period_ord"]).agg(
        avg_price    =("price","mean"),
        avg_price_m2 =("price_per_m2","mean"),
        total_units  =("sub_listing_id","nunique"),
    ).reset_index().sort_values("period_ord")
    trend["avg_price"]    = trend["avg_price"].round(0)
    trend["avg_price_m2"] = trend["avg_price_m2"].round(1)

    # House type stats (unit-level, same structure as unit_type_stats)
    if "house_type" in dl.columns:
        ht_stats = dl[dl["house_type"].notna() & (dl["house_type"] != "")].groupby("house_type").agg(
            count     =("sub_listing_id","nunique"),
            min_price =("price","min"),
            avg_price =("price","mean"),
            max_price =("price","max"),
            avg_size  =("size","mean"),
            avg_pm2   =("price_per_m2","mean"),
        ).reset_index()
        for c in ["min_price","avg_price","max_price"]:
            ht_stats[c] = ht_stats[c].round(0)
        ht_stats["avg_size"] = ht_stats["avg_size"].round(1)
        ht_stats["avg_pm2"]  = ht_stats["avg_pm2"].round(0)
        ht_stats = ht_stats.sort_values("count", ascending=False)
    else:
        ht_stats = pd.DataFrame()

    # Per-listing, per-unit-type and per-house-type counts for accurate frontend breakdown
    _dd_lid_ut = (
        dl.groupby(["listing_id","unit_type"])["sub_listing_id"]
        .nunique().reset_index(name="cnt")
    )
    _dd_lid_ut_map = {}
    for _, _r in _dd_lid_ut.iterrows():
        _dd_lid_ut_map.setdefault(int(_r["listing_id"]), {})[_r["unit_type"]] = int(_r["cnt"])

    _dd_lid_ht = (
        dl[dl["house_type"].notna() & (dl["house_type"] != "")]
        .groupby(["listing_id","house_type"])["sub_listing_id"]
        .nunique().reset_index(name="cnt")
    )
    _dd_lid_ht_map = {}
    for _, _r in _dd_lid_ht.iterrows():
        _dd_lid_ht_map.setdefault(int(_r["listing_id"]), {})[_r["house_type"]] = int(_r["cnt"])

    # Per-listing, per-house-type × per-unit-type cross-reference (active period)
    # Allows frontend to accurately count unit types within a house type and vice versa
    _dd_lid_ht_ut_map = {}
    if "unit_type" in dl.columns and "house_type" in dl.columns:
        _dd_ht_ut_grp = (
            dl[dl["unit_type"].notna() & dl["house_type"].notna() & (dl["house_type"] != "")]
            .groupby(["listing_id","house_type","unit_type"])["sub_listing_id"]
            .nunique().reset_index(name="cnt")
        )
        for _, _r in _dd_ht_ut_grp.iterrows():
            _dd_lid_ht_ut_map.setdefault(int(_r["listing_id"]), {}).setdefault(_r["house_type"], {})[_r["unit_type"]] = int(_r["cnt"])

    # Previous period data for ALL listings in this municipality
    # sold = unique sub_listing_ids in non-latest periods that are NOT in latest period (truly removed)
    _all_lids = set(int(lid) for lid in listings_grp["listing_id"].unique())
    _dd_prev_ut_map        = {}
    _dd_prev_ut_stats      = {}
    _dd_prev_ht_map        = {}
    _dd_sold_per_lid: dict = {}  # listing_id → count of truly-removed (sold) sub-listings
    _dd_d_prev_removed_all = pd.DataFrame()  # initialized here; populated below if historical data exists
    # Active sub-listing count per listing (always computed — used in sold calculation below)
    _dd_active_sub_cnt = dl.groupby("listing_id")["sub_listing_id"].nunique().to_dict()
    # Per-listing per-unit-type price stats for active (latest) period
    _dd_ut_stats: dict = {}
    if "unit_type" in dl.columns:
        _dd_active_price_agg = dl[dl["unit_type"].notna()].groupby(["listing_id","unit_type"]).agg(
            avg_price=("price","mean"), min_price=("price","min"),
            max_price=("price","max"), avg_pm2=("price_per_m2","mean"), avg_size=("size","mean"),
        ).reset_index()
        for _, _r in _dd_active_price_agg.iterrows():
            _dd_ut_stats.setdefault(int(_r["listing_id"]), {})[_r["unit_type"]] = {
                "avg_price": round(float(_r["avg_price"])) if pd.notna(_r["avg_price"]) else None,
                "min_price": round(float(_r["min_price"])) if pd.notna(_r["min_price"]) else None,
                "max_price": round(float(_r["max_price"])) if pd.notna(_r["max_price"]) else None,
                "avg_pm2":   round(float(_r["avg_pm2"]))   if pd.notna(_r["avg_pm2"])   else None,
                "avg_size":  round(float(_r["avg_size"]),1) if pd.notna(_r["avg_size"]) else None,
            }
    # Truly removed = in non-latest periods AND sub_listing_id is not active anywhere globally.
    # Use the global active set so the sold count matches the overview grid.
    d_non_latest_all = d[~d["_is_latest"] & d["listing_id"].isin(_all_lids) & d["sub_listing_id"].notna()].copy()
    if not d_non_latest_all.empty:
        d_non_latest_all["_sid_i"] = d_non_latest_all["sub_listing_id"].astype(int)
        _dd_d_prev_removed_all = d_non_latest_all[~d_non_latest_all["_sid_i"].isin(_global_active_sub_ids)]
        # Build per-listing sold count from truly-removed unique sub-listings
        _dd_sold_counts = (
            _dd_d_prev_removed_all.groupby("listing_id")["sub_listing_id"].nunique()
        )
        _dd_sold_per_lid = {int(k): int(v) for k, v in _dd_sold_counts.items()}
        # For unit_type breakdown: only listings that actually have sold units
        _dd_lids_with_sold = set(_dd_sold_per_lid.keys())
        if _dd_lids_with_sold:
            _dd_d_prev_removed = _dd_d_prev_removed_all[_dd_d_prev_removed_all["listing_id"].isin(_dd_lids_with_sold)]
            d_non_latest = d_non_latest_all[d_non_latest_all["listing_id"].isin(_dd_lids_with_sold)]
            for _, _r in (_dd_d_prev_removed[_dd_d_prev_removed["unit_type"].notna()]
                          .groupby(["listing_id","unit_type"])["sub_listing_id"]
                          .nunique().reset_index(name="cnt")).iterrows():
                _dd_prev_ut_map.setdefault(int(_r["listing_id"]), {})[_r["unit_type"]] = int(_r["cnt"])
            # Price stats: most recent previous period per listing
            _dd_prev_max_ord = d_non_latest.groupby("listing_id")["period_ord"].max().reset_index().rename(columns={"period_ord":"_pm"})
            _dd_d_prev_latest = d_non_latest.merge(_dd_prev_max_ord, on="listing_id")
            _dd_d_prev_latest = _dd_d_prev_latest[_dd_d_prev_latest["period_ord"] == _dd_d_prev_latest["_pm"]]
            _dd_price_agg = _dd_d_prev_latest.groupby(["listing_id","unit_type"]).agg(
                avg_price=("price","mean"), min_price=("price","min"),
                max_price=("price","max"), avg_pm2=("price_per_m2","mean"), avg_size=("size","mean"),
            ).reset_index()
            for _, _r in _dd_price_agg.iterrows():
                _dd_prev_ut_stats.setdefault(int(_r["listing_id"]), {})[_r["unit_type"]] = {
                    "avg_price": round(float(_r["avg_price"])) if pd.notna(_r["avg_price"]) else None,
                    "min_price": round(float(_r["min_price"])) if pd.notna(_r["min_price"]) else None,
                    "max_price": round(float(_r["max_price"])) if pd.notna(_r["max_price"]) else None,
                    "avg_pm2":   round(float(_r["avg_pm2"]))   if pd.notna(_r["avg_pm2"])   else None,
                    "avg_size":  round(float(_r["avg_size"]),1) if pd.notna(_r["avg_size"]) else None,
                }
            # Historical house_type counts — only truly removed sub-listings (not still active)
            _dd_d_prev_ht = _dd_d_prev_removed[_dd_d_prev_removed["house_type"].notna() & (_dd_d_prev_removed["house_type"] != "")]
            for _, _r in _dd_d_prev_ht.groupby(["listing_id","house_type"])["sub_listing_id"].nunique().reset_index(name="cnt").iterrows():
                _dd_prev_ht_map.setdefault(int(_r["listing_id"]), {})[_r["house_type"]] = int(_r["cnt"])
            # Fallback: listings with sold unit-type counts but no sold house-type data
            # (sub-listings had unit_type but null house_type). Distribute sold total
            # proportionally across the listing's active house types.
            for _lid_s in _dd_lids_with_sold:
                if _lid_s in _dd_prev_ut_map and _lid_s not in _dd_prev_ht_map:
                    _active_ht = _dd_lid_ht_map.get(_lid_s, {})
                    if not _active_ht:
                        continue
                    _sold_total = sum(_dd_prev_ut_map[_lid_s].values())
                    _ht_total   = sum(_active_ht.values()) or 1
                    for _ht, _ht_cnt in _active_ht.items():
                        _dd_prev_ht_map.setdefault(_lid_s, {})[_ht] = round(_sold_total * _ht_cnt / _ht_total)

    listings_records = listings_grp.to_dict(orient="records")
    for rec in listings_records:
        lid = int(rec["listing_id"])
        _lat, _lng, _ = _listing_coords(lid, municipality)
        rec["lat"] = _lat
        rec["lng"] = _lng
        rec["nearest_beach_km"], rec["nearest_beach_name"] = _nearest_beach(lid, _lat, _lng)
        rec["unit_type_counts"]            = _dd_lid_ut_map.get(lid, {})
        rec["unit_type_stats"]             = _dd_ut_stats.get(lid, {})
        rec["house_type_counts"]           = _dd_lid_ht_map.get(lid, {})
        rec["house_type_unit_counts"]      = _dd_lid_ht_ut_map.get(lid, {})
        rec["prev_unit_type_counts"]       = _dd_prev_ut_map.get(lid, {})
        rec["prev_unit_type_stats"]        = _dd_prev_ut_stats.get(lid, {})
        rec["prev_house_type_counts"]      = _dd_prev_ht_map.get(lid, {})
        rec["prev_house_type_unit_counts"] = {}  # populated below after historical data is computed
        # sold = truly removed sub-listings (in non-latest periods, not present in latest at all)
        _active_cnt = _dd_active_sub_cnt.get(lid, int(rec.get("units", 0)))
        _sold = _dd_sold_per_lid.get(lid, 0)
        rec["prev_total_units"] = _sold  # sold count only
        # Partial delisted: recent sub-listing loss (PARTIAL_DELISTED_IDS) OR
        # sold units exist but active units have no type data ("phantom active" case)
        _active_has_types = bool(_dd_lid_ut_map.get(lid))
        rec["is_partial_delisted"] = (lid in PARTIAL_DELISTED_IDS) or (
            _active_cnt > 0 and _sold > 0 and not _active_has_types
        )
        # Exclude phantom active sub-listings (no unit_type data) from the display count
        _active_display = _active_cnt if _active_has_types else 0
        rec["units"] = _active_display + _sold

    # Historical (sold) unit-type and house-type aggregates for fallback price stats
    _dd_prev_ut_agg_df = pd.DataFrame()
    _dd_prev_ht_agg_df = pd.DataFrame()
    if _dd_d_prev_removed_all is not None and not _dd_d_prev_removed_all.empty:
        _tmp_ut = _dd_d_prev_removed_all[_dd_d_prev_removed_all["unit_type"].notna()]
        if not _tmp_ut.empty:
            _dd_prev_ut_agg_df = _tmp_ut.groupby("unit_type").agg(
                count=("sub_listing_id","nunique"), min_price=("price","min"),
                avg_price=("price","mean"), max_price=("price","max"),
                avg_size=("size","mean"), avg_pm2=("price_per_m2","mean"),
            ).reset_index()
            for _c in ["min_price","avg_price","max_price"]: _dd_prev_ut_agg_df[_c] = _dd_prev_ut_agg_df[_c].round(0)
            _dd_prev_ut_agg_df["avg_size"] = _dd_prev_ut_agg_df["avg_size"].round(1)
            _dd_prev_ut_agg_df["avg_pm2"]  = _dd_prev_ut_agg_df["avg_pm2"].round(0)
        _tmp_ht = _dd_d_prev_removed_all[_dd_d_prev_removed_all["house_type"].notna() & (_dd_d_prev_removed_all["house_type"] != "")]
        if not _tmp_ht.empty:
            _dd_prev_ht_agg_df = _tmp_ht.groupby("house_type").agg(
                count=("sub_listing_id","nunique"), min_price=("price","min"),
                avg_price=("price","mean"), max_price=("price","max"),
                avg_size=("size","mean"), avg_pm2=("price_per_m2","mean"),
            ).reset_index()
            for _c in ["min_price","avg_price","max_price"]: _dd_prev_ht_agg_df[_c] = _dd_prev_ht_agg_df[_c].round(0)
            _dd_prev_ht_agg_df["avg_size"] = _dd_prev_ht_agg_df["avg_size"].round(1)
            _dd_prev_ht_agg_df["avg_pm2"]  = _dd_prev_ht_agg_df["avg_pm2"].round(0)

    # Historical cross-reference: per-listing, per-house-type × per-unit-type (sold period)
    _dd_lid_prev_ht_ut_map = {}
    if not _dd_d_prev_removed_all.empty and "unit_type" in _dd_d_prev_removed_all.columns and "house_type" in _dd_d_prev_removed_all.columns:
        _dd_prev_ht_ut_grp = (
            _dd_d_prev_removed_all[
                _dd_d_prev_removed_all["unit_type"].notna() &
                _dd_d_prev_removed_all["house_type"].notna() &
                (_dd_d_prev_removed_all["house_type"] != "")
            ].groupby(["listing_id","house_type","unit_type"])["sub_listing_id"]
            .nunique().reset_index(name="cnt")
        )
        for _, _r in _dd_prev_ht_ut_grp.iterrows():
            _dd_lid_prev_ht_ut_map.setdefault(int(_r["listing_id"]), {}).setdefault(_r["house_type"], {})[_r["unit_type"]] = int(_r["cnt"])
    # Fallback: sold sub-listings often lack house_type; distribute prev_unit_type_counts
    # proportionally using the active cross-reference to fill prev_house_type_unit_counts
    for _lid in _all_lids:
        if _lid in _dd_prev_ut_map and _lid not in _dd_lid_prev_ht_ut_map:
            _active_xref = _dd_lid_ht_ut_map.get(_lid, {})
            if not _active_xref:
                continue
            _estimated: dict = {}
            for _ut, _sold_cnt in _dd_prev_ut_map[_lid].items():
                _ht_shares = {_ht: _active_xref[_ht].get(_ut, 0) for _ht in _active_xref if _ut in _active_xref[_ht]}
                _total = sum(_ht_shares.values()) or 1
                for _ht, _cnt in _ht_shares.items():
                    _estimated.setdefault(_ht, {})[_ut] = round(_sold_cnt * _cnt / _total)
            if _estimated:
                _dd_lid_prev_ht_ut_map[_lid] = _estimated
    for rec in listings_records:
        rec["prev_house_type_unit_counts"] = _dd_lid_prev_ht_ut_map.get(int(rec["listing_id"]), {})

    # Add sold units to stats.total_units
    _dd_total_sold = int(_dd_d_prev_removed_all["sub_listing_id"].nunique()) if not _dd_d_prev_removed_all.empty else 0
    stats["total_units"] += _dd_total_sold

    return safe_json({"listings": listings_records,
                      "stats": stats,
                      "unit_type_mix": mix.to_dict(orient="records"),
                      "unit_type_stats":      ut_stats.to_dict(orient="records"),
                      "prev_unit_type_stats": _clean(_dd_prev_ut_agg_df.to_dict(orient="records")) if not _dd_prev_ut_agg_df.empty else [],
                      "house_type_stats":      ht_stats.to_dict(orient="records"),
                      "prev_house_type_stats": _clean(_dd_prev_ht_agg_df.to_dict(orient="records")) if not _dd_prev_ht_agg_df.empty else [],
                      "price_dist": price_dist.to_dict(orient="records"),
                      "m2_dist": m2_dist.to_dict(orient="records"),
                      "trend": trend.drop("period_ord",axis=1).to_dict(orient="records")})

# ══════════════════════════════════════════════════════════════════════════
#  DRILL-DOWN — listing (development)
# ══════════════════════════════════════════════════════════════════════════
@app.get("/drilldown/listing/{listing_id}")
def drilldown_listing(listing_id: int):
    d = df[df["listing_id"] == listing_id]
    if d.empty: return safe_json({})

    meta = d[d["_is_latest"]].iloc[0] if not d[d["_is_latest"]].empty else d.iloc[0]
    dl   = d[d["_is_latest"]]

    apt_cols = ["sub_listing_id","unit_type","price","size","price_per_m2",
                "floor","floor_num","bedrooms","bathrooms","floor_area_m2",
                "has_terrace","has_parking","has_pool","has_garden",
                "has_lift","has_ac","has_storage","has_wardrobes","unit_url"]
    if "last_updated" in dl.columns: apt_cols = apt_cols + ["last_updated"]
    if "house_type"   in dl.columns: apt_cols = apt_cols + ["house_type"]
    apts = dl[apt_cols].copy()
    for col in ["floor_num","bedrooms","bathrooms","floor_area_m2"]:
        apts[col] = pd.to_numeric(apts[col], errors="coerce").astype("Int64")
    for col in ["has_terrace","has_parking","has_pool","has_garden","has_lift","has_ac","has_storage","has_wardrobes"]:
        apts[col] = apts[col].fillna(False).astype(bool)
    apts = apts.sort_values(["unit_type","price"])
    apt_records = _clean(apts.to_dict(orient="records"))
    apt_records = [{k:(None if str(v)=="<NA>" else v) for k,v in r.items()} for r in apt_records]
    # attach coords (same for all apts in same listing)
    lat, lng, map_url = _listing_coords(listing_id, str(meta["municipality"]) if pd.notna(meta.get("municipality")) else "")
    for r in apt_records:
        r["lat"] = lat or 39.47
        r["lng"] = lng or -0.38

    # ── All sub-listings: use last-seen record per sub_listing (active + sold) ──
    _d_all_sorted = d.sort_values("period_ord")
    _d_last = _d_all_sorted.drop_duplicates("sub_listing_id", keep="last").copy()
    _latest_sub_ids = set(dl["sub_listing_id"].tolist())
    _d_last["_status"] = _d_last["sub_listing_id"].apply(lambda x: "active" if x in _latest_sub_ids else "sold")

    floor_price = _d_last.dropna(subset=["floor_num"])[["floor_num","price","unit_type","size","sub_listing_id"]].copy()
    floor_price["floor_num"] = pd.to_numeric(floor_price["floor_num"], errors="coerce")
    floor_price = floor_price.dropna(subset=["floor_num"])
    floor_price["floor_num"] = floor_price["floor_num"].astype(int)

    # unit_comparison: include both active and sold sub-listings using last-known data
    _uc_grp = _d_last[_d_last["unit_type"].notna()].groupby("unit_type")
    _uc_rows = []
    _ut_order = ["Studio","1BR","2BR","3BR","4BR","5BR","Penthouse"]
    for ut, grp in _uc_grp:
        _uc_rows.append({
            "unit_type":    str(ut),
            "count":        int(len(grp)),
            "active_count": int((grp["_status"] == "active").sum()),
            "sold_count":   int((grp["_status"] == "sold").sum()),
            "avg_price":    round(float(grp["price"].mean()))    if grp["price"].notna().any()       else None,
            "min_price":    round(float(grp["price"].min()))     if grp["price"].notna().any()       else None,
            "max_price":    round(float(grp["price"].max()))     if grp["price"].notna().any()       else None,
            "avg_size":     round(float(grp["size"].mean()), 1)  if grp["size"].notna().any()        else None,
            "avg_price_m2": round(float(grp["price_per_m2"].mean()), 1) if grp["price_per_m2"].notna().any() else None,
        })
    _uc_rows.sort(key=lambda r: _ut_order.index(r["unit_type"]) if r["unit_type"] in _ut_order else 99)
    unit_comp = _uc_rows  # list of dicts, not DataFrame

    # ── LISTING TIME SERIES: avg price, inventory, unit mix per period ──────
    listing_trend = d.groupby(["period","period_ord"]).agg(
        avg_price    =("price","mean"),
        avg_price_m2 =("price_per_m2","mean"),
        total_units  =("sub_listing_id","nunique"),
        min_price    =("price","min"),
        max_price    =("price","max"),
    ).reset_index().sort_values("period_ord")
    listing_trend["avg_price"]    = listing_trend["avg_price"].round(0)
    listing_trend["avg_price_m2"] = listing_trend["avg_price_m2"].round(1)
    listing_trend["min_price"]    = listing_trend["min_price"].round(0)
    listing_trend["max_price"]    = listing_trend["max_price"].round(0)

    # per-unit-type trend
    ut_trend = d.groupby(["unit_type","period","period_ord"]).agg(
        avg_price=("price","mean"), count=("sub_listing_id","nunique")
    ).reset_index().sort_values(["unit_type","period_ord"])
    ut_trend["avg_price"] = ut_trend["avg_price"].round(0)

    # ── PER-APARTMENT TIME SERIES ─────────────────────────────────────────
    apt_trend = d.groupby(["sub_listing_id","period","period_ord"]).agg(
        price        =("price","first"),
        price_per_m2 =("price_per_m2","first"),
        unit_type    =("unit_type","first"),
    ).reset_index().sort_values(["sub_listing_id","period_ord"])
    # attach static metadata from latest record for each sub_listing
    d_sorted = d.sort_values("period_ord") if "period_ord" in d.columns else d
    apt_meta = d_sorted.drop_duplicates("sub_listing_id", keep="last")[["sub_listing_id","floor","size","unit_url","bedrooms"]]
    apt_trend = apt_trend.merge(apt_meta, on="sub_listing_id", how="left")
    apt_trend["price_per_m2"] = apt_trend["price_per_m2"].round(1)
    # flag rows where unit_type differs from latest period for that sub_listing
    latest_ut = apt_trend.groupby("sub_listing_id")["unit_type"].last().rename("current_unit_type")
    apt_trend = apt_trend.merge(latest_ut, on="sub_listing_id", how="left")
    apt_trend["unit_type_changed"] = apt_trend["unit_type"] != apt_trend["current_unit_type"]
    apt_trend.drop(columns=["current_unit_type"], inplace=True)
    apt_trend_records = _clean(apt_trend.drop("period_ord",axis=1).to_dict(orient="records"))

    return safe_json({
        "listing_id":    int(listing_id),
        "property_name": str(meta["property_name"]),
        "developer":     str(meta["developer"]),
        "municipality":  str(meta["municipality"]),
        "delivery_date": str(meta["delivery_date"]),
        "esg_grade":     str(meta["esg_grade"]) if pd.notna(meta["esg_grade"]) else None,
        "description":   next((str(meta[c]) for c in ["description","property_description","descripcion","desc","comments"] if c in d.columns and pd.notna(meta.get(c))), None),
        "is_tourist":    any(bool(str(meta.get(c,"")).lower().find("tourist apartment") >= 0) for c in ["description","property_description","descripcion","desc","comments"] if c in d.columns and pd.notna(meta.get(c))),
        "stated_total_units": _extract_stated_units(next((str(meta[c]) for c in ["description","property_description","descripcion","desc","comments"] if c in d.columns and pd.notna(meta.get(c))), None)),
        "nearest_beach_km":   (_nb := _nearest_beach(listing_id, lat, lng))[0],
        "nearest_beach_name": _nb[1],
        "active_units":  int(dl["sub_listing_id"].nunique()),
        "sold_units":    max(
            _listing_expired_counts.get(listing_id, 0),
            max(0, int(d["sub_listing_id"].nunique()) - int(dl["sub_listing_id"].nunique()))
        ),
        "total_units":   int(dl["sub_listing_id"].nunique()) + max(
            _listing_expired_counts.get(listing_id, 0),
            max(0, int(d["sub_listing_id"].nunique()) - int(dl["sub_listing_id"].nunique()))
        ),
        "periods":       PERIODS_SORTED,
        "apartments":    apt_records,
        "floor_price":   floor_price.to_dict(orient="records"),
        "unit_comparison": unit_comp,
        "house_type_comparison": (lambda rows: rows)(
            [{"house_type": ht,
              "count":        int(len(g)),
              "active_count": int((g["_status"] == "active").sum()),
              "sold_count":   int((g["_status"] == "sold").sum()),
              "avg_price":    round(float(g["price"].mean()),0)    if g["price"].notna().any() else None,
              "min_price":    round(float(g["price"].min()),0)     if g["price"].notna().any() else None,
              "max_price":    round(float(g["price"].max()),0)     if g["price"].notna().any() else None,
              "avg_size":     round(float(g["size"].mean()),1)     if g["size"].notna().any()  else None,
              "avg_price_m2": round(float(g["price_per_m2"].mean()),1) if g["price_per_m2"].notna().any() else None,
             } for ht, g in _d_last[_d_last["house_type"].notna() & (_d_last["house_type"] != "")].groupby("house_type")]
            if "house_type" in _d_last.columns else []
        ),
        "listing_trend": listing_trend.drop("period_ord",axis=1).to_dict(orient="records"),
        "unit_type_trend": ut_trend.drop("period_ord",axis=1).to_dict(orient="records"),
        "apt_trend":     apt_trend_records,
        "unit_url_sample": str(dl["unit_url"].dropna().iloc[0]) if "unit_url" in dl.columns and dl["unit_url"].notna().any() else None,
    })

# ══════════════════════════════════════════════════════════════════════════
#  APARTMENT PRICE MATRIX  — wide format, one row per apartment, col per period
# ══════════════════════════════════════════════════════════════════════════
@app.get("/drilldown/listing/{listing_id}/price-matrix")
def price_matrix(listing_id: int):
    d = df[df["listing_id"] == listing_id]
    if d.empty: return safe_json({"periods":[], "rows":[]})

    periods = PERIODS_SORTED  # use global sorted periods

    rows = []
    for sub_id, grp in d.groupby("sub_listing_id"):
        # Use latest period's record for metadata (unit_type may have changed)
        grp_sorted = grp.sort_values("period_ord") if "period_ord" in grp.columns else grp
        meta = grp_sorted.iloc[-1]
        row = {
            "sub_listing_id": int(sub_id),
            "unit_type":  str(meta["unit_type"]),
            "floor":      str(meta["floor"]) if pd.notna(meta["floor"]) else "—",
            "size":       float(meta["size"])       if pd.notna(meta["size"])    else None,
            "bedrooms":   int(meta["bedrooms"])     if pd.notna(meta.get("bedrooms")) else None,
            "bathrooms":  int(meta["bathrooms"])    if pd.notna(meta.get("bathrooms")) else None,
            "has_terrace":bool(meta.get("has_terrace", False)),
            "has_parking":bool(meta.get("has_parking", False)),
            "has_pool":   bool(meta.get("has_pool",   False)),
            "has_lift":   bool(meta.get("has_lift",   False)),
            "has_ac":     bool(meta.get("has_ac",     False)),
            "has_storage":bool(meta.get("has_storage",False)),
            "unit_url":   str(meta["unit_url"])     if pd.notna(meta.get("unit_url")) else None,
            "last_updated": str(meta["last_updated"]) if pd.notna(meta.get("last_updated")) else None,
            "house_type": str(meta["house_type"]) if pd.notna(meta.get("house_type")) else None,
        }

        price_vals = []
        for period in periods:
            pr = grp[grp["period"] == period]
            if len(pr):
                p  = int(pr["price"].iloc[0])
                m2 = round(float(pr["price_per_m2"].iloc[0]), 1)
                price_vals.append(p)
            else:
                p, m2 = None, None
            row[f"price_{period}"] = p
            row[f"ppm2_{period}"]  = m2

        # change vs first available period
        valid = [v for v in price_vals if v is not None]
        if len(valid) >= 2:
            row["price_change"]     = valid[-1] - valid[0]
            row["price_change_pct"] = round((valid[-1] - valid[0]) / valid[0] * 100, 2)
        else:
            row["price_change"]     = 0
            row["price_change_pct"] = 0.0

        # latest price for default sort
        row["latest_price"]    = valid[-1] if valid else None
        row["latest_ppm2"]     = row.get(f"ppm2_{periods[-1]}")

        # detect unit_type changes across periods
        ut_by_period = {}
        for period in periods:
            pr = grp[grp["period"] == period]
            if len(pr):
                ut_by_period[period] = str(pr["unit_type"].iloc[0])
        seen_uts = list(dict.fromkeys(ut_by_period.values()))  # ordered unique
        row["unit_type_changed"] = len(seen_uts) > 1
        row["unit_type_history"] = seen_uts if len(seen_uts) > 1 else []
        # per-period unit types for tooltip
        row["unit_type_by_period"] = ut_by_period

        rows.append(row)

    return safe_json({"periods": periods, "rows": rows})

# ══════════════════════════════════════════════════════════════════════════
#  GEOCOORDINATES  — municipality centroid lookup
# ══════════════════════════════════════════════════════════════════════════
MUNI_COORDS = {
    "Aiora, Valencia":(38.93,-0.97),"Alaquás":(39.46,-0.46),"Albal":(39.38,-0.43),
    "Albalat Dels Sorells":(39.55,-0.37),"Alberic":(39.12,-0.52),"Alcàsser":(39.40,-0.44),
    "Aldaia":(39.46,-0.48),"Alfafar":(39.41,-0.39),"Alfara de Baronia":(39.74,-0.28),
    "Alfara del Patriarca":(39.56,-0.39),"Algar de Palancia":(39.75,-0.34),
    "Algemesi":(39.18,-0.44),"Almardá, Sagunto/Sagunt":(39.68,-0.25),
    "Almàssera":(39.53,-0.36),"Antigua Moreria, Sagunto/Sagunt":(39.68,-0.27),
    "Arrancapins, Valencia":(39.47,-0.39),"Barranquet - El Salvador, Godella":(39.53,-0.42),
    "Barrio de Favara, Valencia":(39.44,-0.38),"Benaguasil":(39.59,-0.56),
    "Beneixida":(39.12,-0.58),"Benetusser":(39.42,-0.40),"Beniarjo":(38.93,-0.19),
    "Benicalap, Valencia":(39.50,-0.40),"Benimàmet":(39.49,-0.43),
    "Beniopa - San Pere, Gandia":(38.97,-0.18),"Beniparrell":(39.39,-0.42),
    "Benipeixcar - El Raval, Gandia":(38.98,-0.18),"Benisano":(39.60,-0.59),
    "Bonrepos i Mirambell":(39.55,-0.38),"Burjassot":(39.51,-0.41),
    "Catarroja":(39.40,-0.40),"Chiva":(39.47,-0.71),"Cullera":(39.16,-0.25),
    "Gandia":(38.97,-0.18),"Godella":(39.53,-0.42),"La Eliana":(39.57,-0.53),
    "L'Eliana":(39.57,-0.53),"Llíria":(39.62,-0.60),"Manises":(39.49,-0.46),
    "Massamagrell":(39.57,-0.34),"Mislata":(39.48,-0.42),"Moncada":(39.55,-0.40),
    "Montserrat":(39.35,-0.47),"Museros":(39.58,-0.36),"Náquera":(39.60,-0.41),
    "Paiporta":(39.42,-0.41),"Paterna":(39.50,-0.44),"Picanya":(39.43,-0.43),
    "Picassent":(39.36,-0.46),"Puçol":(39.62,-0.32),"Quart de Poblet":(39.47,-0.45),
    "Riba-roja de Túria":(39.56,-0.61),"Rocafort":(39.54,-0.41),"Sagunto":(39.68,-0.27),
    "Sedaví":(39.42,-0.39),"Silla":(39.36,-0.41),"Tavernes de la Valldigna":(39.07,-0.27),
    "Torrent":(39.43,-0.46),"Valencia":(39.47,-0.38),"Vilamarxant":(39.58,-0.65),
    "Xàtiva":(38.99,-0.52),"Xirivella":(39.46,-0.42),
    # Extra city-district entries
    "Campanar, Valencia":(39.49,-0.40),"Camins al Grau, Valencia":(39.47,-0.36),
    "El Pla del Real, Valencia":(39.48,-0.38),"Extramurs, Valencia":(39.47,-0.39),
    "La Saïdia, Valencia":(39.49,-0.37),"L'Eixample, Valencia":(39.47,-0.38),
    "Olivereta, Valencia":(39.46,-0.40),"Poblats Maritims, Valencia":(39.46,-0.33),
    "Quatre Carreres, Valencia":(39.44,-0.37),"Rascanya, Valencia":(39.50,-0.38),
    "Sant Marcel·lí, Valencia":(39.44,-0.39),"Tavernes Blanques":(39.53,-0.37),
    "Torrefiel, Valencia":(39.50,-0.38),"Zona Parc Central - Hort de Trenor, Torrent":(39.43,-0.47),
    "Natzaret, Valencia":(39.45,-0.34),"El Saler, Valencia":(39.35,-0.32),
    "Borbotó, Valencia":(39.54,-0.40),"Massarrojos, Valencia":(39.55,-0.39),
    "Castellar-l'Oliveral, Valencia":(39.41,-0.37),"Pobles del Nord, Valencia":(39.55,-0.38),
    "Benimaclet, Valencia":(39.49,-0.36),"Nou Moles, Valencia":(39.47,-0.40),
    "Fonteta de Sant Lluis, Valencia":(39.44,-0.39),"Jesús, Valencia":(39.45,-0.39),
    "Patraix, Valencia":(39.46,-0.40),"Sant Pau-Bon Pastor, Valencia":(39.48,-0.40),
    "Tormos, Valencia":(39.50,-0.38),"La Roqueta, Valencia":(39.47,-0.38),
}

def _get_coords(municipality):
    if not municipality: return None, None
    # exact match
    if municipality in MUNI_COORDS:
        return MUNI_COORDS[municipality]
    # try substring
    for k, v in MUNI_COORDS.items():
        if municipality.lower() in k.lower() or k.lower() in municipality.lower():
            return v
    return None, None

# Per-listing precise coordinates — built from latitude/longitude columns in Excel
# Falls back to municipality geocoding for listings without coords
def _build_listing_coords():
    coords = {}
    for lid, grp in _raw.groupby("listing_id"):
        row = grp.iloc[0]
        lat = row.get("latitude") if "latitude" in grp.columns else None
        lng = row.get("longitude") if "longitude" in grp.columns else None
        map_url = row.get("map") if "map" in grp.columns else None
        if pd.notna(lat) and pd.notna(lng) and float(lat) != 0 and float(lng) != 0:
            coords[int(lid)] = {
                "lat": float(lat),
                "lng": float(lng),
                "map_url": str(map_url) if pd.notna(map_url) else None,
            }
    return coords

LISTING_COORDS = _build_listing_coords()

# Extended coord lookup: all active listings, falling back to municipality coords when no explicit coords.
# Used by nearby endpoints so radius filter covers listings without GPS data in the Excel.
def _build_all_listing_coords():
    coords = dict(LISTING_COORDS)   # start with exact coords
    _lid_muni = (
        df[df["_is_latest"]]
        .drop_duplicates("listing_id")[["listing_id","municipality"]]
        .set_index("listing_id")["municipality"]
        .to_dict()
    )
    for lid, muni in _lid_muni.items():
        lid = int(lid)
        if lid not in coords:
            lat, lng = _get_coords(str(muni))
            if lat and lng:
                coords[lid] = {"lat": float(lat), "lng": float(lng)}
    return coords

ALL_LISTING_COORDS = _build_all_listing_coords()

# Pre-compute partial-delisted listing IDs (still active but lost some sub-listings vs prev period)
def _build_partial_delisted_ids():
    if not PREV_PERIOD:
        return set()
    d_latest = df[df["_is_latest"]]
    d_prev   = df[~df["_is_latest"]]
    if d_prev.empty:
        return set()
    prev_period_ord = d_prev["period_ord"].max()
    d_prev_last = d_prev[d_prev["period_ord"] == prev_period_ord]
    has_latest = set(d_latest["listing_id"].unique())
    partial = set()
    prev_subs_by_lid = d_prev_last.groupby("listing_id")["sub_listing_id"].apply(set).to_dict()
    latest_subs_by_lid = d_latest.groupby("listing_id")["sub_listing_id"].apply(set).to_dict()
    for lid in has_latest:
        prev_subs = prev_subs_by_lid.get(lid, set())
        if prev_subs and (prev_subs - latest_subs_by_lid.get(lid, set())):
            partial.add(lid)
    return partial

PARTIAL_DELISTED_IDS = _build_partial_delisted_ids()

def _listing_coords(listing_id, municipality=""):
    """Return (lat, lng, map_url) — from Excel coords first, then municipality fallback."""
    c = LISTING_COORDS.get(int(listing_id) if listing_id else -1)
    if c:
        return c["lat"], c["lng"], c.get("map_url")
    lat, lng = _get_coords(str(municipality))
    return (lat or 39.47), (lng or -0.38), None

def _parse_address(city_area):
    """Return {street, municipality, comarca, province} from city_area string."""
    if pd.isna(city_area): return {}
    parts = [p.strip() for p in str(city_area).split(",")]
    if len(parts) < 2: return {"raw": str(city_area)}
    province = parts[-1]
    comarca  = parts[-2] if len(parts) >= 3 else None
    rest     = parts[:-2]
    import re as _re
    if rest and _re.match(r'^(Calle|Carrer|Avinguda|Avda|Avenida|Plaza|Plaça|Urb)', rest[0], _re.I):
        street = _re.sub(r' NN$', '', rest[0])
        muni   = rest[1] if len(rest) > 1 else None
    else:
        street = None
        muni   = rest[0] if rest else None
    return {"street": street, "municipality": muni, "comarca": comarca, "province": province}

def _comarca(city_area):
    d = _parse_address(city_area)
    return d.get("comarca")

# Add derived columns to df
for _d in [df]:
    _d["comarca"]      = _d["city_area"].apply(_comarca)
    _d["addr_street"]  = _d["city_area"].apply(lambda x: _parse_address(x).get("street"))
    _d["addr_comarca"] = _d["city_area"].apply(lambda x: _parse_address(x).get("comarca"))

# ══════════════════════════════════════════════════════════════════════════
#  DESCRIPTION SEARCH INDEX  — pre-built at startup for fast keyword search
# ══════════════════════════════════════════════════════════════════════════

def _esg_grade_val(val):
    if not val or str(val).lower() in ("nan", "unknown", ""):
        return None
    m = re.findall(r':\s*([A-G])', str(val), re.IGNORECASE)
    if m:
        return sorted(m, key=lambda g: "ABCDEFG".index(g.upper()))[0].upper()
    m2 = re.match(r'^([A-G])$', str(val).strip(), re.IGNORECASE)
    return m2.group(1).upper() if m2 else None

import bisect as _bisect

_DSI: dict = {}
_DSI_words: list = []
_DSI_word_idx: dict = {}

_DSI_DEL: dict = {}
_DSI_DEL_words: list = []
_DSI_DEL_word_idx: dict = {}


def _build_description_index():
    _desc_col = next((c for c in _DESC_COLS if c in df.columns), None)
    if not _desc_col:
        return

    def _make_entry(lid, sub, sold_date=None):
        desc = ""
        if _desc_col in sub.columns:
            dv_rows = sub.dropna(subset=[_desc_col])
            if not dv_rows.empty:
                v = dv_rows.iloc[0][_desc_col]
                if pd.notna(v):
                    desc = str(v).strip()
                    for sl in ["This comment was automatically translated and may not be 100% accurate.",
                               "See description in the original language"]:
                        desc = desc.replace(sl, "").strip()
        if not desc:
            return None
        words = [w.lower() for w in re.split(r'\W+', desc) if len(w) >= 2]
        if not words:
            return None
        meta = sub.iloc[0]
        lat_v, lng_v, _ = _listing_coords(lid, str(meta.get("municipality", "")))
        ht_counts = dict(sub[sub["house_type"].notna()].groupby("house_type")["sub_listing_id"].nunique().astype(int)) \
                    if "house_type" in sub.columns else {}
        _ut_stats = {}
        if "unit_type" in sub.columns and "price" in sub.columns:
            _sub_ut = sub[sub["unit_type"].notna()].copy()
            if not _sub_ut.empty:
                _ag = {"avg_price": ("price", "mean"), "min_price": ("price", "min"), "max_price": ("price", "max")}
                if "price_per_m2" in sub.columns: _ag["avg_pm2"] = ("price_per_m2", "mean")
                if "size" in sub.columns:         _ag["avg_size"] = ("size", "mean")
                for _, _ur in _sub_ut.groupby("unit_type").agg(**_ag).reset_index().iterrows():
                    _ut_stats[_ur["unit_type"]] = {
                        "avg_price": round(float(_ur["avg_price"])) if pd.notna(_ur.get("avg_price")) else None,
                        "min_price": round(float(_ur["min_price"])) if pd.notna(_ur.get("min_price")) else None,
                        "max_price": round(float(_ur["max_price"])) if pd.notna(_ur.get("max_price")) else None,
                        "avg_pm2":   round(float(_ur["avg_pm2"]))   if "avg_pm2"  in _ur.index and pd.notna(_ur.get("avg_pm2"))  else None,
                        "avg_size":  round(float(_ur["avg_size"]), 1) if "avg_size" in _ur.index and pd.notna(_ur.get("avg_size")) else None,
                    }
        esg    = _esg_grade_val(meta.get("esg_certificate", ""))
        avg_p  = round(float(sub["price"].mean())) if not sub["price"].isna().all() else None
        min_p  = round(float(sub["price"].min()))  if not sub["price"].isna().all() else None
        max_p  = round(float(sub["price"].max()))  if not sub["price"].isna().all() else None
        avg_pm2 = round(float(sub["price_per_m2"].mean()), 1) \
                  if "price_per_m2" in sub.columns and not sub["price_per_m2"].isna().all() else None
        min_pm2 = round(float(sub["price_per_m2"].min()), 1) \
                  if "price_per_m2" in sub.columns and not sub["price_per_m2"].isna().all() else None
        max_pm2 = round(float(sub["price_per_m2"].max()), 1) \
                  if "price_per_m2" in sub.columns and not sub["price_per_m2"].isna().all() else None
        unit_types  = ", ".join(sorted(sub["unit_type"].dropna().unique())) if "unit_type" in sub.columns else ""
        house_types = ", ".join(sorted(t for t in (sub["house_type"].dropna().unique()
                                                    if "house_type" in sub.columns else []) if t))
        record = {
            "listing_id": lid, "property_name": _clean(meta.get("property_name", "")),
            "municipality": _clean(meta.get("municipality", "")), "province": _clean(meta.get("province", "")),
            "developer": _clean(meta.get("developer", "")), "city_area": _clean(meta.get("city_area", "")),
            "lat": lat_v, "lng": lng_v, "lat_exact": lid in LISTING_COORDS, "esg_grade": esg,
            "unit_types": unit_types, "house_types": house_types,
            "avg_price": avg_p, "avg_price_m2": avg_pm2, "min_price": min_p, "max_price": max_p,
            "units": int(sub["sub_listing_id"].nunique()) if "sub_listing_id" in sub.columns else 0,
            "unit_type_counts": dict(sub.groupby("unit_type")["sub_listing_id"].nunique().astype(int)) if "unit_type" in sub.columns else {},
            "house_type_counts": ht_counts,
            "prev_unit_type_counts": {}, "prev_house_type_counts": {},
            "unit_type_stats": _ut_stats, "house_type_unit_counts": {}, "prev_house_type_unit_counts": {},
            "is_partial_delisted": lid in PARTIAL_DELISTED_IDS,
            "stated_total_units": _extract_stated_units(desc),
            "nearest_beach_km": None, "nearest_beach_name": None,
            "description_snippet": "", "full_description_html": "",
        }
        if sold_date is not None:
            record["sold_date"] = sold_date
            record["delisted_type"] = "full"
        muni_str = str(meta.get("municipality", ""))
        prov_str = str(meta.get("province", "")) if "province" in meta.index else ""
        return {
            "words": words, "word_set": frozenset(words), "desc": desc, "record": record,
            "municipality": muni_str, "province": prov_str,
            "unit_type_set": frozenset(sub["unit_type"].dropna().unique()) if "unit_type" in sub.columns else frozenset(),
            "house_type_set": frozenset(t for t in (sub["house_type"].dropna().unique()
                                                     if "house_type" in sub.columns else []) if t),
            "esg_grade": esg, "min_price": min_p, "max_price": max_p, "min_pm2": min_pm2, "max_pm2": max_pm2,
        }

    _word_tmp: dict = {}
    for lid_val, sub in df[df["_is_latest"]].groupby("listing_id"):
        lid   = int(lid_val)
        entry = _make_entry(lid, sub)
        if not entry: continue
        _DSI[lid] = entry
        for w in entry["word_set"]:
            if w not in _word_tmp: _word_tmp[w] = set()
            _word_tmp[w].add(lid)
    _DSI_word_idx.update({w: frozenset(lids) for w, lids in _word_tmp.items()})
    _DSI_words.extend(sorted(_word_tmp.keys()))

    _latest_lids    = set(df[df["_is_latest"]]["listing_id"].unique())
    _nonlatest_lids = set(df[~df["_is_latest"]]["listing_id"].unique())
    _fully_del_lids = _nonlatest_lids - _latest_lids
    if _fully_del_lids:
        _word_del_tmp: dict = {}
        for lid_val, sub in df[df["listing_id"].isin(_fully_del_lids)].groupby("listing_id"):
            lid  = int(lid_val)
            _dp = sub[~sub["_is_latest"]].copy() if "_is_latest" in sub.columns else sub.copy()
            if _dp.empty: _dp = sub.copy()
            if "period_ord" in _dp.columns:
                _dp = _dp[_dp["period_ord"] == _dp["period_ord"].max()]
            sold_dt = None
            if "period" in sub.columns:
                _sp = sub["period"].dropna()
                if not _sp.empty: sold_dt = str(_sp.max())
            entry = _make_entry(lid, _dp, sold_date=sold_dt)
            if not entry: continue
            _DSI_DEL[lid] = entry
            for w in entry["word_set"]:
                if w not in _word_del_tmp: _word_del_tmp[w] = set()
                _word_del_tmp[w].add(lid)
        _DSI_DEL_word_idx.update({w: frozenset(lids) for w, lids in _word_del_tmp.items()})
        _DSI_DEL_words.extend(sorted(_word_del_tmp.keys()))


_build_description_index()

# ══════════════════════════════════════════════════════════════════════════
#  MAP DATA  — all listings with coords
# ══════════════════════════════════════════════════════════════════════════
@app.get("/map/listings")
def map_listings(municipality: Optional[List[str]] = Query(None)):
    d = _latest_df()
    if municipality:
        d = d[d["municipality"].isin(municipality)]
    grp = d.groupby(["listing_id","property_name","municipality","city_area","comarca"], dropna=False).agg(
        units      = ("sub_listing_id","nunique"),
        avg_price  = ("price","mean"),
        min_price  = ("price","min"),
    ).reset_index()
    # All-time unique sub-listing count per listing (for sold = all_time - active)
    _all_lids_map = set(grp["listing_id"].astype(int).unique())
    _alltime_map  = {
        int(k): int(v)
        for k, v in df[df["listing_id"].isin(_all_lids_map)].groupby("listing_id")["sub_listing_id"].nunique().items()
    }
    rows = []
    for _, r in grp.iterrows():
        lat, lng, map_url = _listing_coords(r["listing_id"], str(r["municipality"]) if pd.notna(r["municipality"]) else "")
        addr = _parse_address(r["city_area"])
        # Skip listings without real coords (fallback would cluster at 39.47,-0.38)
        if not lat or not lng: continue
        lid        = int(r["listing_id"])
        active_cnt = int(r["units"])
        expired_sold = _listing_expired_counts.get(lid, 0)
        period_sold  = max(0, _alltime_map.get(lid, active_cnt) - active_cnt)
        sold_cnt     = max(expired_sold, period_sold)
        rows.append({
            "listing_id":  lid,
            "property_name": str(r["property_name"]),
            "municipality": str(r["municipality"]) if pd.notna(r["municipality"]) else "",
            "comarca":     str(r["comarca"])     if pd.notna(r["comarca"])     else "",
            "street":      addr.get("street","") or "",
            "lat": lat, "lng": lng, "map_url": map_url,
            "units":     active_cnt + sold_cnt,
            "avg_price": round(float(r["avg_price"])),
            "min_price": int(r["min_price"]),
        })
    return safe_json(rows)


def _haversine_km(lat1, lng1, lat2, lng2):
    """Distance in km between two lat/lng points."""
    import math as _math
    R = 6371
    dlat = _math.radians(lat2 - lat1)
    dlng = _math.radians(lng2 - lng1)
    a = _math.sin(dlat/2)**2 + _math.cos(_math.radians(lat1)) * _math.cos(_math.radians(lat2)) * _math.sin(dlng/2)**2
    return R * 2 * _math.atan2(_math.sqrt(a), _math.sqrt(1-a))

# ══════════════════════════════════════════════════════════════════════════
#  BEACH DISTANCES  — pre-computed by precompute_beaches.py, loaded at startup
# ══════════════════════════════════════════════════════════════════════════
_BEACH_DIST_FILE = os.path.join(os.path.dirname(__file__), "beach_distances.json")
_BEACH_DISTANCES: dict = {}
if os.path.exists(_BEACH_DIST_FILE):
    try:
        with open(_BEACH_DIST_FILE) as _f:
            _BEACH_DISTANCES = {int(k): v for k, v in json.load(_f).items()}
        print(f"Loaded beach distances for {len(_BEACH_DISTANCES)} listings")
    except Exception as _e:
        print(f"Could not load beach_distances.json: {_e}")

def _nearest_beach(listing_id, *args, **kwargs):
    """Return (distance_km, beach_name) from pre-computed beach_distances.json."""
    rec = _BEACH_DISTANCES.get(int(listing_id) if listing_id else -1)
    if rec:
        return rec.get("nearest_beach_km"), rec.get("nearest_beach_name")
    return None, None

# ══════════════════════════════════════════════════════════════════════════
#  NEARBY COMPARISON  — same comarca, different listings
# ══════════════════════════════════════════════════════════════════════════
@app.get("/nearby/listings/{listing_id}")
def nearby_listings(listing_id: int, radius_km: Optional[float] = None):
    base = df[df["listing_id"]==listing_id]
    if base.empty: return safe_json({"comarca":"","listings":[]})
    comarca = base["comarca"].iloc[0]
    base_lat, base_lng, _ = _listing_coords(listing_id, str(base.iloc[0]["municipality"]))

    # Filter by radius if provided, else fall back to comarca
    if radius_km and base_lat and base_lng:
        # Use only listings that have real coords (not fallback)
        nearby_ids = [
            lid for lid, c in ALL_LISTING_COORDS.items()
            if _haversine_km(base_lat, base_lng, c["lat"], c["lng"]) <= radius_km
        ]
        d = df[df["listing_id"].isin(nearby_ids) & df["_is_latest"]]
    elif pd.isna(comarca):
        return safe_json({"comarca":"","listings":[]})
    else:
        d = df[(df["comarca"]==comarca) & df["_is_latest"]]
    grp = d.groupby(["listing_id","property_name","municipality","city_area","developer","delivery_date","esg_grade"], dropna=False).agg(
        units        = ("sub_listing_id","nunique"),
        avg_price    = ("price","mean"),
        min_price    = ("price","min"),
        max_price    = ("price","max"),
        avg_price_m2 = ("price_per_m2","mean"),
        avg_size     = ("size","mean"),
        unit_types   = ("unit_type", lambda x: ", ".join(sorted(set(x.dropna())))),
    ).reset_index()
    grp["avg_price"]    = grp["avg_price"].round(0)
    grp["avg_price_m2"] = grp["avg_price_m2"].round(1)
    grp["avg_size"]     = grp["avg_size"].round(1)
    grp["esg_grade"]    = grp["esg_grade"].where(pd.notna(grp["esg_grade"]), None)

    # Add coords
    rows = []
    for _, r in grp.iterrows():
        lat, lng, map_url = _listing_coords(r["listing_id"], str(r["municipality"]))
        addr = _parse_address(r["city_area"])
        rows.append({
            **{k: (_clean(r[k]) if not isinstance(r[k], (str,bool,type(None))) else r[k]) for k in r.index},
            "is_current": int(r["listing_id"]) == listing_id,
            "lat": lat or 39.47, "lng": lng or -0.38,
            "map_url": map_url,
            "street": addr.get("street","") or "",
            "addr_comarca": str(comarca),
        })
    return safe_json({"comarca": str(comarca), "listings": rows})

@app.get("/nearby/apartments/{listing_id}")
def nearby_apartments(listing_id: int, unit_type: Optional[str] = None, radius_km: Optional[float] = None):
    """Compare apartments across nearby listings (same comarca or radius)."""
    base = df[df["listing_id"]==listing_id]
    if base.empty: return safe_json({"comarca":"","apartments":[]})
    comarca = base["comarca"].iloc[0]
    base_lat, base_lng, _ = _listing_coords(listing_id, str(base.iloc[0]["municipality"]))

    if radius_km and base_lat and base_lng:
        nearby_ids = [
            lid for lid, c in ALL_LISTING_COORDS.items()
            if _haversine_km(base_lat, base_lng, c["lat"], c["lng"]) <= radius_km
        ]
        d = df[df["listing_id"].isin(nearby_ids) & df["_is_latest"]].copy()
    elif pd.isna(comarca):
        return safe_json({"comarca":"","apartments":[]})
    else:
        d = df[(df["comarca"]==comarca) & df["_is_latest"]].copy()
    if unit_type:
        d = d[d["unit_type"]==unit_type]

    apt_cols = ["sub_listing_id","listing_id","property_name","municipality","unit_type",
                "price","size","price_per_m2","floor","bedrooms","bathrooms",
                "has_terrace","has_parking","has_pool","has_lift","has_ac","unit_url","city_area"]
    if "house_type"    in d.columns: apt_cols = apt_cols + ["house_type"]
    if "last_updated"  in d.columns: apt_cols = apt_cols + ["last_updated"]
    apts = d[apt_cols].drop_duplicates("sub_listing_id").copy()
    for col in ["bedrooms","bathrooms"]:
        apts[col] = pd.to_numeric(apts[col], errors="coerce").astype("Int64")

    rows = []
    for _, r in apts.iterrows():
        lat, lng, map_url = _listing_coords(r["listing_id"], str(r["municipality"]))
        addr = _parse_address(r["city_area"])
        row = {c: r[c] for c in apt_cols}
        row["lat"] = lat or 39.47
        row["lng"] = lng or -0.38
        row["map_url"] = map_url
        row["street"] = addr.get("street","") or ""
        row["is_current_listing"] = int(r["listing_id"]) == listing_id
        for k in row:
            if hasattr(row[k], 'item'):
                row[k] = row[k].item()
            if isinstance(row[k], float) and (row[k] != row[k]):
                row[k] = None
        rows.append(row)

    rows.sort(key=lambda r: r.get("price") or 999999)
    return safe_json({"comarca": str(comarca), "apartments": rows})


@app.get("/nearby/apartments/{listing_id}/trend")
def nearby_apartments_trend(listing_id: int, unit_type: Optional[str] = None, radius_km: Optional[float] = None):
    """Avg price & price_per_m2 trend over all periods for similar nearby apartments."""
    base = df[df["listing_id"]==listing_id]
    if base.empty: return safe_json({"trend": []})
    comarca = base["comarca"].iloc[0]
    base_lat, base_lng, _ = _listing_coords(listing_id, str(base.iloc[0]["municipality"]))

    if radius_km and base_lat and base_lng:
        nearby_ids = [
            lid for lid, c in ALL_LISTING_COORDS.items()
            if _haversine_km(base_lat, base_lng, c["lat"], c["lng"]) <= radius_km
        ]
        d = df[df["listing_id"].isin(nearby_ids)].copy()
    elif pd.isna(comarca):
        return safe_json({"trend": []})
    else:
        d = df[df["comarca"] == comarca].copy()

    if unit_type:
        d = d[d["unit_type"] == unit_type]

    # Exclude the current listing itself
    d = d[d["listing_id"] != listing_id]
    if d.empty: return safe_json({"trend": []})

    trend = d.groupby(["period","period_ord"]).agg(
        avg_price    =("price","mean"),
        avg_price_m2 =("price_per_m2","mean"),
        count        =("sub_listing_id","nunique"),
    ).reset_index().sort_values("period_ord")
    trend["avg_price"]    = trend["avg_price"].round(0)
    trend["avg_price_m2"] = trend["avg_price_m2"].round(1)
    return safe_json({"trend": trend.drop("period_ord",axis=1).to_dict(orient="records")})


# ══════════════════════════════════════════════════════════════════════════
#  LISTING DETAIL — enhanced with address + coords
# ══════════════════════════════════════════════════════════════════════════
@app.get("/listing/meta/{listing_id}")
def listing_meta(listing_id: int):
    d = df[df["listing_id"]==listing_id]
    if d.empty: return safe_json({})
    r = d.iloc[0]
    lat, lng, map_url = _listing_coords(listing_id, str(r["municipality"]))
    addr = _parse_address(r["city_area"])
    return safe_json({
        "listing_id":   int(listing_id),
        "property_name":str(r["property_name"]),
        "municipality": str(r["municipality"]) if pd.notna(r["municipality"]) else "",
        "city_area":    str(r["city_area"])    if pd.notna(r["city_area"])    else "",
        "comarca":      str(r["comarca"])      if pd.notna(r["comarca"])      else "",
        "street":       addr.get("street","") or "",
        "lat": lat or 39.47, "lng": lng or -0.38, "map_url": map_url,
    })

# ══════════════════════════════════════════════════════════════════════════
#  DELISTED — developments + apartments present in prev period but not latest
# ══════════════════════════════════════════════════════════════════════════
@app.get("/delisted/listings")
def delisted_listings(province: Optional[List[str]] = Query(None),
                      municipality: Optional[List[str]] = Query(None)):
    d = df.copy()
    if province:     d = d[d["province"].isin(province)]
    if municipality: d = d[d["municipality"].isin(municipality)]

    # Per-province aware: a listing is fully delisted if it has rows in a non-latest period
    # but no rows with a valid sub_listing_id in its province's latest period
    has_latest     = set(d[d["_is_latest"] & d["sub_listing_id"].notna()]["listing_id"].unique())
    has_non_latest = set(d[~d["_is_latest"] & d["sub_listing_id"].notna()]["listing_id"].unique())
    fully_delisted = has_non_latest - has_latest

    # Partially delisted: still in latest period but lost some sub_listings vs previous period
    d_latest     = d[d["_is_latest"]]
    d_prev       = d[~d["_is_latest"]]
    partial_delisted = set()
    if PREV_PERIOD and not d_prev.empty:
        prev_period_ord = d_prev["period_ord"].max()
        d_prev_last = d_prev[d_prev["period_ord"] == prev_period_ord]
        for lid in has_latest:
            prev_subs   = set(d_prev_last[d_prev_last["listing_id"] == lid]["sub_listing_id"].dropna().unique())
            latest_subs = set(d_latest[d_latest["listing_id"] == lid]["sub_listing_id"].dropna().unique())
            if prev_subs and (prev_subs - latest_subs):
                partial_delisted.add(lid)

    all_delisted = fully_delisted | partial_delisted
    if not all_delisted:
        return safe_json({"listings": [], "summary": {"count":0,"units":0}, "periods": {"prev":PREV_PERIOD,"latest":LATEST_PERIOD}})

    def _build_records(listing_ids, use_latest=False):
        if use_latest:
            dp = d_latest[d_latest["listing_id"].isin(listing_ids)]
        else:
            d_nl = d[(d["listing_id"].isin(listing_ids)) & (~d["_is_latest"])]
            max_ords = d_nl.groupby("listing_id")["period_ord"].max().reset_index().rename(columns={"period_ord":"_max_ord"})
            d_nl = d_nl.merge(max_ords, on="listing_id")
            dp = d_nl[d_nl["period_ord"] == d_nl["_max_ord"]].drop("_max_ord", axis=1)
        grp = dp.groupby(["listing_id","property_name","developer","municipality","city_area","esg_grade","delivery_date"], dropna=False).agg(
            units        =("sub_listing_id","nunique"),
            avg_price    =("price","mean"),
            min_price    =("price","min"),
            max_price    =("price","max"),
            avg_price_m2 =("price_per_m2","mean"),
            avg_size     =("size","mean"),
            unit_types   =("unit_type",  lambda x: ", ".join(sorted(x.dropna().unique().tolist()))),
            house_types  =("house_type", lambda x: ", ".join(sorted(t for t in x.dropna().unique() if t))),
            has_pool     =("has_pool","max"),
            has_parking  =("has_parking","max"),
            has_terrace  =("has_terrace","max"),
            has_lift     =("has_lift","max"),
            last_period  =("period","max"),
        ).reset_index()
        for c in ["avg_price","min_price","max_price"]:
            grp[c] = grp[c].round(0)
        grp["avg_price_m2"] = grp["avg_price_m2"].round(1)
        grp["avg_size"]     = grp["avg_size"].round(1)
        grp["esg_grade"]    = grp["esg_grade"].where(pd.notna(grp["esg_grade"]), None)
        return grp, dp

    grp_full, dp_full = _build_records(fully_delisted, use_latest=False)
    grp_part, dp_part = _build_records(partial_delisted, use_latest=True)

    grp_full["delisted_type"] = "full"
    grp_part["delisted_type"] = "partial"

    grp = pd.concat([grp_full, grp_part], ignore_index=True)
    dp  = pd.concat([dp_full,  dp_part],  ignore_index=True)

    # Per-listing, per-unit-type and per-house-type counts
    _dl_lid_ut = (
        dp.groupby(["listing_id","unit_type"])["sub_listing_id"]
        .nunique().reset_index(name="cnt")
    )
    _dl_lid_ut_map = {}
    for _, _r in _dl_lid_ut.iterrows():
        _dl_lid_ut_map.setdefault(int(_r["listing_id"]), {})[_r["unit_type"]] = int(_r["cnt"])

    _dl_lid_ht = (
        dp[dp["house_type"].notna() & (dp["house_type"] != "")]
        .groupby(["listing_id","house_type"])["sub_listing_id"]
        .nunique().reset_index(name="cnt")
    )
    _dl_lid_ht_map = {}
    for _, _r in _dl_lid_ht.iterrows():
        _dl_lid_ht_map.setdefault(int(_r["listing_id"]), {})[_r["house_type"]] = int(_r["cnt"])

    records = grp.to_dict(orient="records")
    for r in records:
        lat, lng, map_url = _listing_coords(int(r["listing_id"]), str(r["municipality"]))
        r["lat"] = lat or 39.47
        r["lng"] = lng or -0.38
        r["lat_exact"] = int(r["listing_id"]) in LISTING_COORDS
        r["sold_date"] = _listing_sold_date(int(r["listing_id"]))
        r["unit_type_counts"]  = _dl_lid_ut_map.get(int(r["listing_id"]), {})
        r["house_type_counts"] = _dl_lid_ht_map.get(int(r["listing_id"]), {})

    summary = {
        "count": len(records),
        "units": int(dp["sub_listing_id"].nunique()),
        "avg_price": round(float(dp["price"].mean())) if len(dp) else 0,
        "avg_price_m2": round(float(dp["price_per_m2"].mean()), 1) if len(dp) else 0,
    }
    return safe_json({"listings": records, "summary": summary,
                      "periods": {"prev": PREV_PERIOD, "latest": LATEST_PERIOD}})


@app.get("/delisted/apartments/{listing_id}")
def delisted_apartments(listing_id: int):
    """All apartments from a delisted listing (from its last non-latest period)."""
    d_lid = df[df["listing_id"]==listing_id]
    d_non_latest = d_lid[~d_lid["_is_latest"]]
    if d_non_latest.empty:
        return safe_json({"apartments": []})
    # Use the most recent non-latest period for this listing
    max_ord = d_non_latest["period_ord"].max()
    dp = d_non_latest[d_non_latest["period_ord"] == max_ord]
    if dp.empty:
        return safe_json({"apartments": []})

    apt_cols = ["sub_listing_id","unit_type","price","size","price_per_m2",
                "floor","floor_num","bedrooms","bathrooms",
                "has_terrace","has_parking","has_pool","has_lift","has_ac","unit_url","last_updated"]
    apts = dp[[c for c in apt_cols if c in dp.columns]].copy()
    for col in ["floor_num","bedrooms","bathrooms"]:
        if col in apts.columns:
            apts[col] = pd.to_numeric(apts[col], errors="coerce").astype("Int64")
    for col in ["has_terrace","has_parking","has_pool","has_lift","has_ac"]:
        if col in apts.columns:
            apts[col] = apts[col].fillna(False).astype(bool)
    apts = apts.sort_values(["unit_type","price"])
    records = _clean(apts.to_dict(orient="records"))
    records = [{k:(None if str(v)=="<NA>" else v) for k,v in r.items()} for r in records]

    meta = dp.iloc[0]
    lat, lng, _ = _listing_coords(listing_id, str(meta["municipality"]))
    return safe_json({
        "property_name": str(meta["property_name"]),
        "municipality":  str(meta["municipality"]),
        "developer":     str(meta["developer"]) if pd.notna(meta.get("developer")) else None,
        "last_period":   str(dp.iloc[0]["period"]) if not dp.empty else PREV_PERIOD,
        "sold_date":     _listing_sold_date(listing_id),
        "city_area":     str(meta["city_area"]) if pd.notna(meta.get("city_area")) else None,
        "esg_grade":     str(meta["esg_grade"]) if pd.notna(meta.get("esg_grade")) else None,
        "description":   str(meta["description"]) if pd.notna(meta.get("description")) else None,
        "lat":           lat,
        "lng":           lng,
        "apartments":    records,
    })

# ══════════════════════════════════════════════════════════════════════════
#  IMAGE CLASSIFICATION — detect floor plans via Pillow image analysis
# ══════════════════════════════════════════════════════════════════════════
import json as _json, os as _os
_IMG_CACHE_FILE = _os.path.join(_os.path.dirname(__file__), "image_cache.json")
_IMG_CACHE: dict = {}

# Load persisted cache on startup
try:
    with open(_IMG_CACHE_FILE, "r") as _f:
        _IMG_CACHE = _json.load(_f)
except Exception:
    _IMG_CACHE = {}

def _save_img_cache():
    try:
        with open(_IMG_CACHE_FILE, "w") as _f:
            _json.dump(_IMG_CACHE, _f)
    except Exception:
        pass

def _is_floorplan(url: str) -> bool:
    """
    Classify an image as a floor plan using Pillow pixel analysis.
    Floor plans are nearly monochrome documents: very high white ratio + very low colour saturation.
    Thresholds calibrated on real Idealista data:
      - property photos:  white_ratio 0.0–0.31, saturation 35–79
      - floor plans:      white_ratio 0.7–0.85, saturation 2–8
    Results are cached to disk so each URL is only downloaded once.
    """
    if url in _IMG_CACHE:
        return _IMG_CACHE[url]

    result = False
    try:
        import requests as _req
        from PIL import Image as _PILImg
        import numpy as _np
        import io as _io

        resp = _req.get(url, timeout=6, headers={"User-Agent": "Mozilla/5.0"})
        img = _PILImg.open(_io.BytesIO(resp.content)).convert("RGB")

        # Downscale for speed (200×200 is plenty for colour stats)
        img.thumbnail((200, 200))
        arr = _np.array(img, dtype=float)

        # White pixel ratio: all channels > 210
        white_ratio = (_np.all(arr > 210, axis=2)).mean()

        # Colour saturation proxy: mean channel range per pixel
        ch_max = arr.max(axis=2)
        ch_min = arr.min(axis=2)
        avg_sat = (ch_max - ch_min).mean()

        # Floor plan: mostly white/light AND low saturation
        # Threshold loosened to catch coloured-room floor plans (light blue/green fills)
        result = bool(white_ratio > 0.45 and avg_sat < 35)

    except Exception:
        result = False

    _IMG_CACHE[url] = result
    # Persist every 20 new entries
    if len(_IMG_CACHE) % 20 == 0:
        _save_img_cache()
    return result

def _classify_images_parallel(urls: list) -> dict:
    """Download + classify a list of URLs concurrently. Returns {url: is_floorplan}."""
    from concurrent.futures import ThreadPoolExecutor, as_completed
    results = {}
    uncached = [u for u in urls if u not in _IMG_CACHE]
    if uncached:
        with ThreadPoolExecutor(max_workers=8) as pool:
            fut = {pool.submit(_is_floorplan, u): u for u in uncached}
            for f in as_completed(fut):
                u = fut[f]
                results[u] = f.result()
        _save_img_cache()
    for u in urls:
        results[u] = _IMG_CACHE.get(u, results.get(u, False))
    return results

# ══════════════════════════════════════════════════════════════════════════
#  LISTING PHOTOS — read from images column in data
# ══════════════════════════════════════════════════════════════════════════
@app.get("/listing/photos/{listing_id}")
def listing_photos(listing_id: int):
    import re as _re, ast as _ast
    rows = df[df["listing_id"] == listing_id]
    if rows.empty:
        return safe_json({"photos": []})

    _FLOORPLAN_RE    = _re.compile(r'plano|planta|floor[_\-]plan|blueprint|fp_', _re.IGNORECASE)
    _FLOORPLAN_TITLE = _re.compile(r'floor.?plan|plano|planta|blueprint', _re.IGNORECASE)

    seen = set()
    # url -> True (floor plan) / False (photo) / None (unknown — needs pixel analysis)
    url_classification = {}

    # ── Try images_dic first (has url + title) ────────────────────────────
    has_images_dic = "images_dic" in rows.columns
    for raw in (rows["images_dic"].dropna() if has_images_dic else []):
        s = str(raw).strip()
        if not s or s in ("0", "nan", "[]"):
            continue
        try:
            parsed = _ast.literal_eval(s)
        except Exception:
            parsed = []
        if not isinstance(parsed, list):
            continue
        for item in parsed:
            if not isinstance(item, dict):
                continue
            url   = str(item.get("url", "")).strip()
            title = str(item.get("title", "")).strip()
            if not url.startswith("http"):
                continue
            m   = _re.search(r'/([a-f0-9]+)\.(webp|jpg)$', url)
            key = m.group(1) if m else url
            if key in seen:
                continue
            seen.add(key)
            if title:
                # Title present — use it definitively
                url_classification[url] = bool(_FLOORPLAN_TITLE.search(title))
            else:
                # Title absent — mark for pixel fallback
                url_classification[url] = None

    # ── Fallback: plain images column for URLs not already seen ──────────
    if "images" in rows.columns:
        for raw in rows["images"].dropna():
            s = str(raw).strip()
            if not s or s == "0":
                continue
            urls = []
            try:
                parsed = _ast.literal_eval(s)
                if isinstance(parsed, list):
                    urls = [str(u).strip() for u in parsed]
            except Exception:
                urls = [u.strip().strip("'").strip('"') for u in s.strip("[]").split(",")]
            for url in urls:
                if not url.startswith("http"):
                    continue
                m   = _re.search(r'/([a-f0-9]+)\.(webp|jpg)$', url)
                key = m.group(1) if m else url
                if key in seen:
                    continue
                seen.add(key)
                url_classification[url] = None  # needs classification

    # ── Classify URLs with no title: URL keyword then pixel analysis ──────
    needs_classify = [u for u, v in url_classification.items() if v is None]
    fp_by_url_kw   = [u for u in needs_classify if     _FLOORPLAN_RE.search(u)]
    needs_pixel    = [u for u in needs_classify if not _FLOORPLAN_RE.search(u)]

    for u in fp_by_url_kw:
        url_classification[u] = True
    pixel_results = _classify_images_parallel(needs_pixel)
    for u in needs_pixel:
        url_classification[u] = pixel_results.get(u, False)

    photos      = [u for u, is_fp in url_classification.items() if not is_fp]
    floor_plans = [u for u, is_fp in url_classification.items() if is_fp]
    return safe_json({"photos": photos, "floor_plans": floor_plans})


@app.get("/listing/photos/{listing_id}/{sub_listing_id}")
def listing_photos_for_apt(listing_id: int, sub_listing_id: int):
    """Return floor plans specific to one sub_listing (apartment unit).
    Uses images_dic titles if available for that row; falls back to full
    listing floor plans if no sub_listing-specific data exists."""
    import re as _re, ast as _ast

    rows = df[df["listing_id"] == listing_id]
    if rows.empty:
        return safe_json({"photos": [], "floor_plans": [], "apt_specific": False})

    _FLOORPLAN_TITLE = _re.compile(r'floor.?plan|plano|planta|blueprint', _re.IGNORECASE)

    # Check if this sub_listing has its own images_dic entry
    if "images_dic" in rows.columns:
        apt_rows = rows[rows["sub_listing_id"] == sub_listing_id]
        apt_images_dic = apt_rows["images_dic"].dropna()
        if not apt_images_dic.empty:
            for raw in apt_images_dic:
                s = str(raw).strip()
                if not s or s in ("0", "nan", "[]"):
                    continue
                try:
                    parsed = _ast.literal_eval(s)
                except Exception:
                    continue
                if not isinstance(parsed, list) or not parsed:
                    continue
                # Has real data — extract floor plans from this apt's images_dic
                apt_floor_plans = []
                apt_photos = []
                for item in parsed:
                    if not isinstance(item, dict):
                        continue
                    url   = str(item.get("url", "")).strip()
                    title = str(item.get("title", "")).strip()
                    if not url.startswith("http"):
                        continue
                    if title and _FLOORPLAN_TITLE.search(title):
                        apt_floor_plans.append(url)
                    elif title:
                        apt_photos.append(url)
                    # no title — skip (don't pixel-classify per-apt images, too slow)
                if apt_floor_plans or apt_photos:
                    return safe_json({"photos": apt_photos, "floor_plans": apt_floor_plans, "apt_specific": True})

    # No sub_listing-specific data — fall back to full listing floor plans
    import json as _json
    full_resp = listing_photos(listing_id)
    full_data = _json.loads(full_resp.body)
    return safe_json({**full_data, "apt_specific": False})

# ── Search page endpoints ─────────────────────────────────────────────────

@app.get("/search/options")
def search_options(municipality: Optional[List[str]] = Query(None)):
    """Return all municipalities and all area/locality/street parts from city_area + geocoded CSV."""
    munis = sorted(_raw["municipality"].dropna().unique().tolist())

    base = _raw.copy()
    if municipality:
        base = base[base["municipality"].isin(municipality)]

    # Always exclude province-level names and municipality names
    exclude = {"valencia", "alicante", "spain", "españa", "comunitat valenciana"}
    all_munis = {m.lower() for m in _raw["municipality"].dropna().unique()}

    locations = set()

    # Extract from city_area column
    for val in base["city_area"].dropna():
        parts = [p.strip() for p in str(val).split(",")]
        for part in parts[:-1]:
            clean = re.sub(r'\s+NN\s*$', '', part, flags=re.IGNORECASE).strip()
            clean = re.sub(r'\s+\d+[a-zA-Z]*\s*$', '', clean).strip()
            if len(clean) < 3: continue
            if clean.lower() in exclude: continue
            if clean.lower() in all_munis: continue
            if re.match(r'^\d+$', clean): continue
            locations.add(clean)

    # Add geocoded streets from CSV
    if not _GEOCODED_STREETS.empty:
        geo = _GEOCODED_STREETS.copy()
        if municipality:
            # Filter to matching municipalities (case-insensitive)
            munis_lower = {m.lower() for m in municipality}
            geo = geo[geo["municipality"].str.lower().isin(munis_lower)]
        for val in geo["street"].dropna():
            clean = str(val).strip()
            clean = re.sub(r'\s+\d+[a-zA-Z]*\s*$', '', clean).strip()
            if len(clean) < 3: continue
            if clean.lower() in exclude: continue
            if clean.lower() in all_munis: continue
            locations.add(clean)

    # Build street → coords lookup for geocoded streets (validated against municipality listings)
    street_coords = {}
    if not _GEOCODED_STREETS.empty:
        geo_for_coords = _GEOCODED_STREETS.copy()
        if municipality:
            munis_lower = {m.lower() for m in municipality}
            geo_for_coords = geo_for_coords[geo_for_coords["municipality"].str.lower().isin(munis_lower)]

        # Compute municipality listing centroid for distance validation
        _opt_centroid = None
        if municipality:
            _opt_lids = list(base["listing_id"].unique())[:100]
            _opt_mc = [(LISTING_COORDS[lid]["lat"], LISTING_COORDS[lid]["lng"])
                       for lid in _opt_lids if lid in LISTING_COORDS]
            if _opt_mc:
                _opt_centroid = (sum(c[0] for c in _opt_mc) / len(_opt_mc),
                                 sum(c[1] for c in _opt_mc) / len(_opt_mc))

        for _, row in geo_for_coords.iterrows():
            if pd.notna(row.get("latitude")) and pd.notna(row.get("longitude")):
                g_lat, g_lng = float(row["latitude"]), float(row["longitude"])
                # Skip if geocoded location is >20 km from municipality listings
                if _opt_centroid and _haversine_km(_opt_centroid[0], _opt_centroid[1], g_lat, g_lng) > 20:
                    continue
                street_coords[str(row["street"]).strip()] = {"lat": g_lat, "lng": g_lng}

    return safe_json({
        "municipalities": munis,
        "locations": sorted(locations),
        "street_coords": street_coords,
    })


@app.get("/search/listings")
def search_listings(
    municipality: Optional[List[str]] = Query(None),
    street:       Optional[List[str]] = Query(None),
    radius_km:    Optional[float]     = Query(None),
    lat:          Optional[float]     = Query(None),
    lng:          Optional[float]     = Query(None),
    unit_type:    Optional[List[str]] = Query(None),
    min_price:    Optional[float]     = Query(None),
    max_price:    Optional[float]     = Query(None),
    min_m2:       Optional[float]     = Query(None),
    max_m2:       Optional[float]     = Query(None),
    esg:          Optional[List[str]] = Query(None),
    house_type:   Optional[List[str]] = Query(None),
):
    def _parse_esg_grade(val):
        """Extract best (lowest) grade letter from 'Consumption: A, Emissions: B' etc."""
        if not val or str(val).lower() in ("nan", "unknown", ""):
            return None
        m = re.findall(r':\s*([A-G])', str(val), re.IGNORECASE)
        if m:
            return sorted(m, key=lambda g: "ABCDEFG".index(g.upper()))[0].upper()
        m2 = re.match(r'^([A-G])$', str(val).strip(), re.IGNORECASE)
        return m2.group(1).upper() if m2 else None

    d = _raw.copy()

    # ESG filter — compare parsed grade
    if esg:
        d = d[d["esg_certificate"].apply(lambda v: _parse_esg_grade(v) in esg)]

    if municipality: d = d[d["municipality"].isin(municipality)]
    if unit_type:    d = d[d["unit_type"].isin(unit_type)]
    if house_type:   d = d[d["house_type"].isin(house_type)]

    # Area/locality/street filter — match against city_area OR use geocoded coords
    geo_centers = []  # populated inside street block, needed later for area_center
    if street:
        # Check if any selected street is in the geocoded CSV — if so, use its coords as center
        if not _GEOCODED_STREETS.empty:
            # Build municipality listing centroid for distance validation
            _muni_centroid = None
            if municipality:
                _lids = d["listing_id"].unique()[:100]  # sample for speed
                _mc = [(LISTING_COORDS[lid]["lat"], LISTING_COORDS[lid]["lng"])
                       for lid in _lids if lid in LISTING_COORDS]
                if _mc:
                    _muni_centroid = (sum(c[0] for c in _mc) / len(_mc),
                                      sum(c[1] for c in _mc) / len(_mc))

            for s in street:
                s_lower = s.strip().lower()
                match = _GEOCODED_STREETS[_GEOCODED_STREETS["street"].str.lower().str.strip() == s_lower]
                if not match.empty:
                    row = match.iloc[0]
                    if pd.notna(row.get("latitude")) and pd.notna(row.get("longitude")):
                        g_lat, g_lng = float(row["latitude"]), float(row["longitude"])
                        # Validate: skip if geocoded location is >20 km from municipality listings
                        if _muni_centroid is None or _haversine_km(_muni_centroid[0], _muni_centroid[1], g_lat, g_lng) <= 20:
                            geo_centers.append((g_lat, g_lng))

        def _street_match(city_area_val):
            v = str(city_area_val).lower()
            for s in street:
                s_clean = re.sub(r'\s+nn\s*$', '', s, flags=re.IGNORECASE).strip().lower()
                if s_clean in v:
                    return True
            return False
        mask = d["city_area"].apply(_street_match)
        d_matched = d[mask]

        # If geocoded streets found and radius_km provided, also include nearby listings
        if geo_centers and radius_km:
            geo_listing_ids = set()
            for c_lat, c_lng in geo_centers:
                for lid, c in LISTING_COORDS.items():
                    if _haversine_km(c_lat, c_lng, c["lat"], c["lng"]) <= radius_km:
                        geo_listing_ids.add(lid)
            d_geo = d[d["listing_id"].isin(geo_listing_ids)]
            d = pd.concat([d_matched, d_geo]).drop_duplicates(subset=["listing_id"]) if not d_geo.empty else d_matched
        elif geo_centers and not radius_km:
            # Geocoded street selected but no radius — use 2 km default to find nearby listings
            _default_r = 2.0
            geo_listing_ids = set()
            for c_lat, c_lng in geo_centers:
                for lid, c in LISTING_COORDS.items():
                    if _haversine_km(c_lat, c_lng, c["lat"], c["lng"]) <= _default_r:
                        geo_listing_ids.add(lid)
            d_geo = d[d["listing_id"].isin(geo_listing_ids)]
            d = pd.concat([d_matched, d_geo]).drop_duplicates(subset=["listing_id"]) if not d_geo.empty else d_matched
        else:
            d = d_matched

    # Price filters
    if min_price: d = d[d["price"] >= min_price]
    if max_price: d = d[d["price"] <= max_price]
    if min_m2:    d = d[d["price_per_m2"] >= min_m2]
    if max_m2:    d = d[d["price_per_m2"] <= max_m2]

    if d.empty:
        _early_center = None
        if geo_centers:
            _early_center = {"lat": sum(c[0] for c in geo_centers) / len(geo_centers),
                             "lng": sum(c[1] for c in geo_centers) / len(geo_centers)}
        return safe_json({"listings": [], "total": 0, "area_center": _early_center})

    # Group to listing level
    # Build safe agg dict — only include columns that exist
    _agg = dict(
        property_name  =("property_name","first"),
        municipality   =("municipality","first"),
        province       =("province","first"),
        developer      =("developer","first"),
        units          =("sub_listing_id","nunique"),
        avg_price      =("price","mean"),
        avg_price_m2   =("price_per_m2","mean"),
        min_price      =("price","min"),
        max_price      =("price","max"),
        unit_types     =("unit_type",  lambda x: ", ".join(sorted(x.dropna().unique()))),
        house_types    =("house_type", lambda x: ", ".join(sorted(t for t in x.dropna().unique() if t))),
        city_area      =("city_area","first"),
    )
    if "esg_certificate" in d.columns:
        _agg["esg_certificate"] = ("esg_certificate","first")
    _desc_col_s = next((c for c in _DESC_COLS if c in d.columns), None)
    if _desc_col_s:
        _agg["_desc_raw"] = (_desc_col_s, "first")
    grp = d.groupby("listing_id").agg(**_agg).reset_index()
    grp["avg_price"]    = grp["avg_price"].round(0)
    grp["avg_price_m2"] = grp["avg_price_m2"].round(1)
    grp["min_price"]    = grp["min_price"].round(0)
    grp["max_price"]    = grp["max_price"].round(0)

    # Attach coords and parsed ESG grade
    rows = []
    for _, r in grp.iterrows():
        lid = int(r["listing_id"])
        lat_v, lng_v, _ = _listing_coords(lid, r["municipality"])
        row = {**{k: _clean(v) for k, v in r.items() if k != "_desc_raw"}, "lat": lat_v, "lng": lng_v,
               "lat_exact": lid in LISTING_COORDS}
        row["esg_grade"] = _parse_esg_grade(r.get("esg_certificate", ""))
        row["stated_total_units"] = _extract_stated_units(str(r["_desc_raw"])) if "_desc_raw" in r.index and pd.notna(r.get("_desc_raw")) else None
        row["nearest_beach_km"], row["nearest_beach_name"] = _nearest_beach(lid, lat_v, lng_v)
        rows.append(row)

    # Radius filter — compute centroid from all rows that have real coords
    if radius_km:
        # Use provided lat/lng or compute centroid of current results
        if lat is not None and lng is not None:
            c_lat, c_lng = lat, lng
        else:
            coords = [(r["lat"], r["lng"]) for r in rows if r["lat"] and r["lng"]]
            if coords:
                c_lat = sum(c[0] for c in coords) / len(coords)
                c_lng = sum(c[1] for c in coords) / len(coords)
            else:
                c_lat, c_lng = None, None

        if c_lat and c_lng:
            rows = [r for r in rows if r["lat"] and r["lng"]
                    and _haversine_km(c_lat, c_lng, r["lat"], r["lng"]) <= radius_km]

    # Per-unit-type stats from the apartment-level data (accurate counts + prices per type)
    # Use province-aware _is_latest flag (same as municipality endpoint) so listings that
    # disappeared before the province's true latest period are correctly treated as sold.
    # Use deduplicated df (not _raw) to avoid duplicate rows with conflicting house_type values.
    _radius_lids = set(r["listing_id"] for r in rows)
    _d_for_ut_all = df[df["listing_id"].isin(_radius_lids)] if _radius_lids else df.iloc[0:0]
    # Active = province-latest rows (not per-listing max, which misses listings absent from true latest)
    _d_for_ut = _d_for_ut_all[_d_for_ut_all["_is_latest"] & _d_for_ut_all["listing_id"].isin(_radius_lids)]

    # Per-listing, per-unit-type counts for accurate frontend breakdown
    # When house_type filter is active, only count unit types from those house type sub-listings
    _d_for_ut_ht = (
        _d_for_ut[_d_for_ut["house_type"].isin(house_type)]
        if house_type and "house_type" in _d_for_ut.columns
        else _d_for_ut
    )
    _lid_ut_counts = (
        _d_for_ut_ht.groupby(["listing_id","unit_type"])["sub_listing_id"]
        .nunique().reset_index(name="cnt")
    )
    _lid_ut_map = {}
    for _, _r in _lid_ut_counts.iterrows():
        _lid_ut_map.setdefault(int(_r["listing_id"]), {})[_r["unit_type"]] = int(_r["cnt"])
    # Per-listing, per-house-type counts for accurate frontend breakdown
    _lid_ht_counts = (
        _d_for_ut[_d_for_ut["house_type"].notna() & (_d_for_ut["house_type"] != "")]
        .groupby(["listing_id","house_type"])["sub_listing_id"]
        .nunique().reset_index(name="cnt")
    )
    _lid_ht_map = {}
    for _, _r in _lid_ht_counts.iterrows():
        _lid_ht_map.setdefault(int(_r["listing_id"]), {})[_r["house_type"]] = int(_r["cnt"])

    # Per-listing, per-house-type × per-unit-type cross-reference (active period, always unfiltered)
    # Built from _d_for_ut (not _d_for_ut_ht) so frontend always gets the complete cross-reference
    _lid_ht_ut_map = {}
    if "unit_type" in _d_for_ut.columns and "house_type" in _d_for_ut.columns:
        _ht_ut_grp = (
            _d_for_ut[_d_for_ut["unit_type"].notna() & _d_for_ut["house_type"].notna() & (_d_for_ut["house_type"] != "")]
            .groupby(["listing_id","house_type","unit_type"])["sub_listing_id"]
            .nunique().reset_index(name="cnt")
        )
        for _, _r in _ht_ut_grp.iterrows():
            _lid_ht_ut_map.setdefault(int(_r["listing_id"]), {}).setdefault(_r["house_type"], {})[_r["unit_type"]] = int(_r["cnt"])

    # Previous period data for ALL result listings
    # sold = unique sub_listing_ids in non-latest periods that are NOT in latest period (truly removed)
    _all_result_lids = {int(r["listing_id"]) for r in rows}
    _prev_lid_ut_map      = {}
    _prev_lid_ut_stats    = {}
    _prev_lid_ht_map      = {}
    _sold_per_lid: dict   = {}  # listing_id → count of truly-removed (sold) sub-listings
    _d_prev_removed_all   = pd.DataFrame()  # initialized here; populated below if historical data exists
    _prev_all_non_latest  = pd.DataFrame()  # same
    # Non-latest = all rows NOT in province latest period
    _d_prev_all_lids = _d_for_ut_all[~_d_for_ut_all["_is_latest"]]
    _active_sub_cnt = _d_for_ut.groupby("listing_id")["sub_listing_id"].nunique().to_dict()
    # Per-listing per-unit-type price stats for active (latest) period
    _lid_ut_stats: dict = {}
    if "unit_type" in _d_for_ut.columns:
        _active_price_agg = _d_for_ut_ht[_d_for_ut_ht["unit_type"].notna()].groupby(["listing_id","unit_type"]).agg(
            avg_price=("price","mean"), min_price=("price","min"),
            max_price=("price","max"), avg_pm2=("price_per_m2","mean"), avg_size=("size","mean"),
        ).reset_index()
        for _, _r in _active_price_agg.iterrows():
            _lid_ut_stats.setdefault(int(_r["listing_id"]), {})[_r["unit_type"]] = {
                "avg_price": round(float(_r["avg_price"])) if pd.notna(_r["avg_price"]) else None,
                "min_price": round(float(_r["min_price"])) if pd.notna(_r["min_price"]) else None,
                "max_price": round(float(_r["max_price"])) if pd.notna(_r["max_price"]) else None,
                "avg_pm2":   round(float(_r["avg_pm2"]))   if pd.notna(_r["avg_pm2"])   else None,
                "avg_size":  round(float(_r["avg_size"]),1) if pd.notna(_r["avg_size"]) else None,
            }
    _prev_all_non_latest = _d_prev_all_lids  # already non-latest by _is_latest flag
    if not _prev_all_non_latest.empty:
        # Truly removed = sub_listing_id not present anywhere in the global active set
        _prev_all_non_latest = _prev_all_non_latest[_prev_all_non_latest["sub_listing_id"].notna()].copy()
        _prev_all_non_latest["_sid_i"] = _prev_all_non_latest["sub_listing_id"].astype(int)
        _d_prev_removed_all = _prev_all_non_latest[~_prev_all_non_latest["_sid_i"].isin(_global_active_sub_ids)]
        # Build per-listing sold count from truly-removed unique sub-listings
        _sold_per_lid = {
            int(k): int(v)
            for k, v in _d_prev_removed_all.groupby("listing_id")["sub_listing_id"].nunique().items()
        }
        _lids_with_sold = set(_sold_per_lid.keys())
        if _lids_with_sold:
            _d_prev_removed = _d_prev_removed_all[_d_prev_removed_all["listing_id"].isin(_lids_with_sold)]
            _d_prev_sold = _prev_all_non_latest[_prev_all_non_latest["listing_id"].isin(_lids_with_sold)]
            # When house_type filter is active, restrict historical unit type counts to those house types
            _d_prev_removed_ht = (
                _d_prev_removed[_d_prev_removed["house_type"].isin(house_type)]
                if house_type and "house_type" in _d_prev_removed.columns
                else _d_prev_removed
            )
            _prev_lid_ut_counts = (
                _d_prev_removed_ht[_d_prev_removed_ht["unit_type"].notna()]
                .groupby(["listing_id","unit_type"])["sub_listing_id"]
                .nunique().reset_index(name="cnt")
            )
            for _, _r in _prev_lid_ut_counts.iterrows():
                _prev_lid_ut_map.setdefault(int(_r["listing_id"]), {})[_r["unit_type"]] = int(_r["cnt"])
            # Price stats: most recent previous period per listing
            _prev_max_ord = _d_prev_sold.groupby("listing_id")["period_ord"].max().reset_index().rename(columns={"period_ord":"_pm"})
            _d_prev_latest = _d_prev_sold.merge(_prev_max_ord, on="listing_id")
            _d_prev_latest = _d_prev_latest[_d_prev_latest["period_ord"] == _d_prev_latest["_pm"]]
            _prev_price_agg = (
                _d_prev_latest.groupby(["listing_id","unit_type"]).agg(
                    avg_price=("price","mean"), min_price=("price","min"),
                    max_price=("price","max"), avg_pm2=("price_per_m2","mean"),
                    avg_size=("size","mean"),
                ).reset_index()
            )
            for _, _r in _prev_price_agg.iterrows():
                _prev_lid_ut_stats.setdefault(int(_r["listing_id"]), {})[_r["unit_type"]] = {
                    "avg_price": round(float(_r["avg_price"])) if pd.notna(_r["avg_price"]) else None,
                    "min_price": round(float(_r["min_price"])) if pd.notna(_r["min_price"]) else None,
                    "max_price": round(float(_r["max_price"])) if pd.notna(_r["max_price"]) else None,
                    "avg_pm2":   round(float(_r["avg_pm2"]))   if pd.notna(_r["avg_pm2"])   else None,
                    "avg_size":  round(float(_r["avg_size"]),1) if pd.notna(_r["avg_size"]) else None,
                }
            # Historical house_type counts — only truly removed sub-listings (not still active)
            _d_prev_ht = _d_prev_removed[_d_prev_removed["house_type"].notna() & (_d_prev_removed["house_type"] != "")]
            for _, _r in _d_prev_ht.groupby(["listing_id","house_type"])["sub_listing_id"].nunique().reset_index(name="cnt").iterrows():
                _prev_lid_ht_map.setdefault(int(_r["listing_id"]), {})[_r["house_type"]] = int(_r["cnt"])
            # Fallback: listings with sold unit-type counts but no sold house-type data
            # (sub-listings had unit_type but null house_type). Distribute sold total
            # proportionally across the listing's active house types.
            for _lid_s in _lids_with_sold:
                if _lid_s in _prev_lid_ut_map and _lid_s not in _prev_lid_ht_map:
                    _active_ht = _lid_ht_map.get(_lid_s, {})
                    if not _active_ht:
                        continue
                    _sold_total = sum(_prev_lid_ut_map[_lid_s].values())
                    _ht_total   = sum(_active_ht.values()) or 1
                    for _ht, _ht_cnt in _active_ht.items():
                        _prev_lid_ht_map.setdefault(_lid_s, {})[_ht] = round(_sold_total * _ht_cnt / _ht_total)

    # Historical cross-reference: per-listing per-house-type × per-unit-type (sold period)
    _prev_lid_ht_ut_map = {}
    if not _d_prev_removed_all.empty and "unit_type" in _d_prev_removed_all.columns and "house_type" in _d_prev_removed_all.columns:
        _prev_ht_ut_grp = (
            _d_prev_removed_all[
                _d_prev_removed_all["unit_type"].notna() &
                _d_prev_removed_all["house_type"].notna() &
                (_d_prev_removed_all["house_type"] != "")
            ].groupby(["listing_id","house_type","unit_type"])["sub_listing_id"]
            .nunique().reset_index(name="cnt")
        )
        for _, _r in _prev_ht_ut_grp.iterrows():
            _prev_lid_ht_ut_map.setdefault(int(_r["listing_id"]), {}).setdefault(_r["house_type"], {})[_r["unit_type"]] = int(_r["cnt"])

    # Fallback: sold sub-listings often lack house_type; distribute prev_unit_type_counts
    # proportionally using the active cross-reference to fill prev_house_type_unit_counts
    for _lid in _all_result_lids:
        if _lid in _prev_lid_ut_map and _lid not in _prev_lid_ht_ut_map:
            _active_xref = _lid_ht_ut_map.get(_lid, {})
            if not _active_xref:
                continue
            _estimated: dict = {}
            for _ut, _sold_cnt in _prev_lid_ut_map[_lid].items():
                _ht_shares = {_ht: _active_xref[_ht].get(_ut, 0) for _ht in _active_xref if _ut in _active_xref[_ht]}
                _total = sum(_ht_shares.values()) or 1
                for _ht, _cnt in _ht_shares.items():
                    _estimated.setdefault(_ht, {})[_ut] = round(_sold_cnt * _cnt / _total)
            if _estimated:
                _prev_lid_ht_ut_map[_lid] = _estimated

    # Attach unit_type_counts, house_type_counts and partial-delisted flag to each row
    for _row in rows:
        _lid_int = int(_row["listing_id"])
        _row["unit_type_counts"]            = _lid_ut_map.get(_lid_int, {})
        _row["unit_type_stats"]             = _lid_ut_stats.get(_lid_int, {})
        _row["house_type_counts"]           = _lid_ht_map.get(_lid_int, {})
        _row["house_type_unit_counts"]      = _lid_ht_ut_map.get(_lid_int, {})
        _row["prev_unit_type_counts"]       = _prev_lid_ut_map.get(_lid_int, {})
        _row["prev_unit_type_stats"]        = _prev_lid_ut_stats.get(_lid_int, {})
        _row["prev_house_type_counts"]      = _prev_lid_ht_map.get(_lid_int, {})
        _row["prev_house_type_unit_counts"] = _prev_lid_ht_ut_map.get(_lid_int, {})
        # sold = truly removed sub-listings (in non-latest periods, not present in latest at all)
        _active_cnt_s = _active_sub_cnt.get(_lid_int, 0)
        _sold_s = _sold_per_lid.get(_lid_int, 0)
        _row["prev_total_units"] = _sold_s  # sold count only
        # Partial delisted: recent sub-listing loss OR sold units with no active type data
        _active_has_types_s = bool(_lid_ut_map.get(_lid_int))
        _row["is_partial_delisted"] = (_lid_int in PARTIAL_DELISTED_IDS) or (
            _active_cnt_s > 0 and _sold_s > 0 and not _active_has_types_s
        )
        # Exclude phantom active sub-listings (no unit_type data) from the display count
        _active_display_s = _active_cnt_s if _active_has_types_s else 0
        _row["units"] = _active_display_s + _sold_s

    # Remove fully sold-out listings — no active typed units, only sold history
    rows = [r for r in rows if r.get("unit_type_counts")]

    _ut_order = ["Studio","1BR","2BR","3BR","4BR","5BR","Penthouse"]
    _ut_agg = _d_for_ut_ht.groupby("unit_type").agg(
        count     =("sub_listing_id","nunique"),
        min_price =("price","min"),
        avg_price =("price","mean"),
        max_price =("price","max"),
        avg_size  =("size","mean"),
        avg_pm2   =("price_per_m2","mean"),
    ).reset_index()
    for _c in ["min_price","avg_price","max_price"]: _ut_agg[_c] = _ut_agg[_c].round(0)
    _ut_agg["avg_size"] = _ut_agg["avg_size"].round(1)
    _ut_agg["avg_pm2"]  = _ut_agg["avg_pm2"].round(0)
    _ut_agg["_s"] = _ut_agg["unit_type"].apply(lambda x: _ut_order.index(x) if x in _ut_order else 99)
    _ut_agg = _ut_agg.sort_values("_s").drop("_s", axis=1)

    # Per-house-type stats (unit level, active period)
    _ht_agg = _d_for_ut[_d_for_ut["house_type"].notna() & (_d_for_ut["house_type"] != "")].groupby("house_type").agg(
        count     =("sub_listing_id","nunique"),
        min_price =("price","min"),
        avg_price =("price","mean"),
        max_price =("price","max"),
        avg_size  =("size","mean"),
        avg_pm2   =("price_per_m2","mean"),
    ).reset_index()
    for _c in ["min_price","avg_price","max_price"]: _ht_agg[_c] = _ht_agg[_c].round(0)
    _ht_agg["avg_size"] = _ht_agg["avg_size"].round(1)
    _ht_agg["avg_pm2"]  = _ht_agg["avg_pm2"].round(0)
    _ht_agg = _ht_agg.sort_values("count", ascending=False)
    # Per-house-type stats for historical/sold sub-listings (fallback for types not in active period)
    _prev_ht_agg = pd.DataFrame()
    if not _prev_all_non_latest.empty:
        _prev_removed_ht = _d_prev_removed_all[
            _d_prev_removed_all["house_type"].notna() & (_d_prev_removed_all["house_type"] != "")
        ] if not _d_prev_removed_all.empty else pd.DataFrame()
        if not _prev_removed_ht.empty:
            _prev_ht_agg = _prev_removed_ht.groupby("house_type").agg(
                count     =("sub_listing_id","nunique"),
                min_price =("price","min"),
                avg_price =("price","mean"),
                max_price =("price","max"),
                avg_size  =("size","mean"),
                avg_pm2   =("price_per_m2","mean"),
            ).reset_index()
            for _c in ["min_price","avg_price","max_price"]: _prev_ht_agg[_c] = _prev_ht_agg[_c].round(0)
            _prev_ht_agg["avg_size"] = _prev_ht_agg["avg_size"].round(1)
            _prev_ht_agg["avg_pm2"]  = _prev_ht_agg["avg_pm2"].round(0)

    # Determine area center for the frontend map
    area_center = None
    if street and geo_centers:
        # Geocoded street: use its exact coordinates
        area_center = {
            "lat": sum(c[0] for c in geo_centers) / len(geo_centers),
            "lng": sum(c[1] for c in geo_centers) / len(geo_centers),
        }
    elif street and rows:
        # Neighbourhood/area name: use centroid of matched result listings (much tighter than full municipality)
        matched_coords = [(r["lat"], r["lng"]) for r in rows if r.get("lat") and r.get("lng")]
        if matched_coords:
            area_center = {
                "lat": sum(c[0] for c in matched_coords) / len(matched_coords),
                "lng": sum(c[1] for c in matched_coords) / len(matched_coords),
            }
    elif municipality:
        muni_lids = df[df["municipality"].isin(municipality)]["listing_id"].unique()
        muni_coords = [
            (LISTING_COORDS[lid]["lat"], LISTING_COORDS[lid]["lng"])
            for lid in muni_lids
            if lid in LISTING_COORDS
        ]
        if muni_coords:
            area_center = {
                "lat": sum(c[0] for c in muni_coords) / len(muni_coords),
                "lng": sum(c[1] for c in muni_coords) / len(muni_coords),
            }

    # Per-unit-type stats for historical/sold sub-listings
    _prev_ut_agg = pd.DataFrame()
    if not _prev_all_non_latest.empty and not _d_prev_removed_all.empty:
        _prev_removed_ut_src = (
            _d_prev_removed_all[_d_prev_removed_all["house_type"].isin(house_type)]
            if house_type and "house_type" in _d_prev_removed_all.columns
            else _d_prev_removed_all
        )
        _prev_removed_ut = _prev_removed_ut_src[_prev_removed_ut_src["unit_type"].notna()]
        if not _prev_removed_ut.empty:
            _prev_ut_agg = _prev_removed_ut.groupby("unit_type").agg(
                count     =("sub_listing_id","nunique"),
                min_price =("price","min"),
                avg_price =("price","mean"),
                max_price =("price","max"),
                avg_size  =("size","mean"),
                avg_pm2   =("price_per_m2","mean"),
            ).reset_index()
            for _c in ["min_price","avg_price","max_price"]: _prev_ut_agg[_c] = _prev_ut_agg[_c].round(0)
            _prev_ut_agg["avg_size"] = _prev_ut_agg["avg_size"].round(1)
            _prev_ut_agg["avg_pm2"]  = _prev_ut_agg["avg_pm2"].round(0)

    return safe_json({"listings": rows, "total": len(rows),
                      "unit_type_stats":      _clean(_ut_agg.to_dict(orient="records")),
                      "prev_unit_type_stats": _clean(_prev_ut_agg.to_dict(orient="records")) if not _prev_ut_agg.empty else [],
                      "house_type_stats":      _clean(_ht_agg.to_dict(orient="records")),
                      "prev_house_type_stats": _clean(_prev_ht_agg.to_dict(orient="records")) if not _prev_ht_agg.empty else [],
                      "area_center": area_center})


# ══════════════════════════════════════════════════════════════════════════
#  DESCRIPTION SEARCH  — keyword / fuzzy search across property descriptions
# ══════════════════════════════════════════════════════════════════════════
_DESC_SEARCH_CACHE: dict = {}
_DESC_SEARCH_CACHE_MAX = 500

@app.get("/description/search")
def description_search_endpoint(
    q:            str                    = Query(default=""),
    municipality: Optional[List[str]]   = Query(None),
    province:     Optional[List[str]]   = Query(None),
    unit_type:    Optional[List[str]]   = Query(None),
    esg:          Optional[List[str]]   = Query(None),
    house_type:   Optional[List[str]]   = Query(None),
    min_price:    Optional[float]       = Query(None),
    max_price:    Optional[float]       = Query(None),
    min_m2:       Optional[float]       = Query(None),
    max_m2:       Optional[float]       = Query(None),
):
    from difflib import SequenceMatcher as _SM

    q = q.strip()
    if not q:
        return safe_json({"listings": [], "total": 0, "unit_type_stats": [], "house_type_stats": [], "delisted": []})

    _cache_key = (
        q.lower(),
        tuple(sorted(municipality or [])),
        tuple(sorted(province     or [])),
        tuple(sorted(unit_type    or [])),
        tuple(sorted(esg          or [])),
        tuple(sorted(house_type   or [])),
        min_price, max_price, min_m2, max_m2,
    )
    if _cache_key in _DESC_SEARCH_CACHE:
        return _DESC_SEARCH_CACHE[_cache_key]

    q_tokens = [w.lower() for w in re.split(r'[\s\W]+', q) if len(w) >= 2]
    if not q_tokens:
        return safe_json({"listings": [], "total": 0, "unit_type_stats": [], "house_type_stats": [], "delisted": []})

    _MARK = lambda s: f'<mark style="background:#FEF3C7;color:#92400E;border-radius:2px;padding:0 2px">{s}</mark>'

    def _highlight(text, tokens):
        for tok in sorted(tokens, key=len, reverse=True):
            text = re.sub(re.escape(tok), lambda m: _MARK(m.group()), text, flags=re.IGNORECASE)
        return text

    def _score_entry(entry):
        wset = entry["word_set"]
        score, matched = 0, set()
        for tok in q_tokens:
            if tok in wset:
                score += 3; matched.add(tok)
            elif any(w.startswith(tok) for w in wset if len(w) >= len(tok)):
                score += 2; matched.add(tok)
            else:
                best = max((_SM(None, tok, w).ratio() for w in wset if abs(len(w) - len(tok)) <= 3), default=0)
                if best >= 0.82:
                    score += 1; matched.add(tok)
        return score, matched

    def _make_snippet_hl(entry, matched):
        text = entry["desc"]
        tl   = text.lower()
        pos   = min((tl.find(t) for t in matched if tl.find(t) >= 0), default=0)
        start = max(0, pos - 60)
        end   = min(len(text), start + 280)
        snip  = ("…" if start else "") + text[start:end] + ("…" if end < len(text) else "")
        return _highlight(snip, matched), _highlight(text, matched)

    def _filter_entry(e):
        if municipality and e["municipality"] not in municipality: return False
        if province     and e["province"]     not in province:     return False
        if unit_type    and not e["unit_type_set"].intersection(unit_type):    return False
        if house_type   and not e["house_type_set"].intersection(house_type):  return False
        if esg          and e["esg_grade"] not in esg:             return False
        if min_price and e["max_price"] is not None and e["max_price"] < min_price: return False
        if max_price and e["min_price"] is not None and e["min_price"] > max_price: return False
        if min_m2    and e["max_pm2"]   is not None and e["max_pm2"]   < min_m2:   return False
        if max_m2    and e["min_pm2"]   is not None and e["min_pm2"]   > max_m2:   return False
        return True

    def _search_index(dsi, dsi_words, dsi_word_idx):
        candidate_lids: set = set()
        for tok in q_tokens:
            candidate_lids.update(dsi_word_idx.get(tok, frozenset()))
            i = _bisect.bisect_left(dsi_words, tok)
            while i < len(dsi_words) and dsi_words[i].startswith(tok):
                candidate_lids.update(dsi_word_idx.get(dsi_words[i], frozenset()))
                i += 1
        hits = []
        for lid in candidate_lids:
            e = dsi.get(lid)
            if e is None or not _filter_entry(e): continue
            score, matched = _score_entry(e)
            if score == 0: continue
            snip, full_hl = _make_snippet_hl(e, matched)
            rec = dict(e["record"])
            rec["match_score"]           = score
            rec["description_snippet"]   = snip
            rec["full_description_html"] = full_hl
            hits.append(rec)
        return hits

    results = _search_index(_DSI, _DSI_words, _DSI_word_idx)
    results.sort(key=lambda x: (-x["match_score"], x.get("property_name", "")))

    _delisted_results = _search_index(_DSI_DEL, _DSI_DEL_words, _DSI_DEL_word_idx)
    _delisted_results.sort(key=lambda x: (-x["match_score"], x.get("property_name", "")))

    _matched_ids = {r["listing_id"] for r in results}
    _UT_ORD = ["Studio", "1BR", "2BR", "3BR", "4BR", "5BR", "Penthouse"]
    _srv_ut, _srv_ht = [], []
    if _matched_ids:
        _dm = df[df["_is_latest"] & df["listing_id"].isin(_matched_ids)]
        if not _dm.empty:
            if "unit_type" in _dm.columns:
                _uta = _dm[_dm["unit_type"].notna()].copy()
                if not _uta.empty:
                    _uag = {"count": ("sub_listing_id", "nunique"), "avg_price": ("price", "mean"),
                            "min_price": ("price", "min"), "max_price": ("price", "max")}
                    if "price_per_m2" in _uta.columns: _uag["avg_pm2"] = ("price_per_m2", "mean")
                    if "size" in _uta.columns:         _uag["avg_size"] = ("size", "mean")
                    _uta_df = _uta.groupby("unit_type").agg(**_uag).reset_index()
                    _uta_df["_o"] = _uta_df["unit_type"].map(lambda x: _UT_ORD.index(x) if x in _UT_ORD else 99)
                    for _, _r in _uta_df.sort_values("_o").iterrows():
                        _srv_ut.append({
                            "unit_type": _r["unit_type"], "count": int(_r["count"]),
                            "avg_price": round(float(_r["avg_price"])) if pd.notna(_r.get("avg_price")) else None,
                            "min_price": round(float(_r["min_price"])) if pd.notna(_r.get("min_price")) else None,
                            "max_price": round(float(_r["max_price"])) if pd.notna(_r.get("max_price")) else None,
                            "avg_pm2":   round(float(_r["avg_pm2"]))   if "avg_pm2"  in _r.index and pd.notna(_r.get("avg_pm2"))  else None,
                            "avg_size":  round(float(_r["avg_size"]), 1) if "avg_size" in _r.index and pd.notna(_r.get("avg_size")) else None,
                        })
            if "house_type" in _dm.columns:
                _hta = _dm[_dm["house_type"].notna() & (_dm["house_type"] != "")].copy()
                if not _hta.empty:
                    _hag = {"count": ("sub_listing_id", "nunique"), "avg_price": ("price", "mean"),
                            "min_price": ("price", "min"), "max_price": ("price", "max")}
                    if "price_per_m2" in _hta.columns: _hag["avg_pm2"] = ("price_per_m2", "mean")
                    if "size" in _hta.columns:         _hag["avg_size"] = ("size", "mean")
                    for _, _r in _hta.groupby("house_type").agg(**_hag).reset_index().sort_values("count", ascending=False).iterrows():
                        _srv_ht.append({
                            "house_type": _r["house_type"], "count": int(_r["count"]),
                            "avg_price":  round(float(_r["avg_price"])) if pd.notna(_r.get("avg_price")) else None,
                            "min_price":  round(float(_r["min_price"])) if pd.notna(_r.get("min_price")) else None,
                            "max_price":  round(float(_r["max_price"])) if pd.notna(_r.get("max_price")) else None,
                            "avg_pm2":    round(float(_r["avg_pm2"]))   if "avg_pm2"  in _r.index and pd.notna(_r.get("avg_pm2"))  else None,
                            "avg_size":   round(float(_r["avg_size"]), 1) if "avg_size" in _r.index and pd.notna(_r.get("avg_size")) else None,
                        })

    _resp = safe_json({"listings": results, "total": len(results),
                       "unit_type_stats": _srv_ut, "house_type_stats": _srv_ht,
                       "delisted": _delisted_results})
    if len(_DESC_SEARCH_CACHE) >= _DESC_SEARCH_CACHE_MAX:
        _DESC_SEARCH_CACHE.clear()
    _DESC_SEARCH_CACHE[_cache_key] = _resp
    return _resp


# ══════════════════════════════════════════════════════════════════════════
#  EXCEL EXPORT  — selected listings in wide per-unit-type format
# ══════════════════════════════════════════════════════════════════════════
@app.get("/search/export")
def export_listings_excel(ids: str = Query(...)):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError:
        return JSONResponse({"error": "openpyxl not installed"}, status_code=500)

    listing_ids = [int(x) for x in ids.split(",") if x.strip().isdigit()]
    if not listing_ids:
        return JSONResponse({"error": "No listing IDs"}, status_code=400)

    AME_COLS = [("has_lift","Lift"),("has_parking","Parking"),("has_pool","Pool"),
                ("has_garden","Garden"),("has_ac","AC"),("has_storage","Storage"),
                ("has_terrace","Terrace"),("has_wardrobes","Wardrobes")]

    # Collect per-listing data (meta + sorted apartment rows)
    listings_data = []
    for lid in listing_ids:
        d = df[df["listing_id"] == lid]
        if d.empty: continue
        latest = d[d["_is_latest"]]
        if latest.empty: latest = d
        meta = latest.iloc[0]

        esg = str(meta.get("esg_certificate","")) if pd.notna(meta.get("esg_certificate","")) else ""
        ht  = ", ".join(sorted(t for t in latest["house_type"].dropna().unique() if t)) if "house_type" in latest.columns else ""
        amenities = [lbl for c_,lbl in AME_COLS if c_ in latest.columns and latest[c_].any()]

        # Sort apartments: by unit_type order then price
        UT_ORDER = {"Studio":0,"1BR":1,"2BR":2,"3BR":3,"4BR":4,"5BR":5,"Penthouse":6}
        apt_cols = ["sub_listing_id","unit_type","price","size","price_per_m2",
                    "floor","floor_num","bedrooms","bathrooms","unit_url"]
        for ac in ["house_type","has_terrace","has_parking","has_pool","has_garden",
                   "has_lift","has_ac","has_storage","has_wardrobes"]:
            if ac in latest.columns: apt_cols.append(ac)
        apts_df = latest[apt_cols].drop_duplicates("sub_listing_id").copy()
        apts_df["_s"] = apts_df["unit_type"].map(lambda x: UT_ORDER.get(x, 99))
        apts_df = apts_df.sort_values(["_s","price"]).drop("_s", axis=1)

        apts = []
        for _, a in apts_df.iterrows():
            apt_ame = [lbl for c_,lbl in AME_COLS if c_ in a.index and bool(a.get(c_))]
            apts.append({
                "unit_type":  str(a.get("unit_type","")) if pd.notna(a.get("unit_type")) else "",
                "house_type": str(a.get("house_type","")) if "house_type" in a.index and pd.notna(a.get("house_type")) else "",
                "floor":      str(a.get("floor",""))     if pd.notna(a.get("floor"))     else "",
                "bedrooms":   int(a["bedrooms"])  if pd.notna(a.get("bedrooms"))  else None,
                "bathrooms":  int(a["bathrooms"]) if pd.notna(a.get("bathrooms")) else None,
                "size":       round(float(a["size"]),1) if pd.notna(a.get("size")) else None,
                "price":      int(round(a["price"])) if pd.notna(a.get("price")) else None,
                "pm2":        int(round(a["price_per_m2"])) if pd.notna(a.get("price_per_m2")) else None,
                "amenities":  "; ".join(apt_ame),
                "url":        str(a["unit_url"]) if "unit_url" in a.index and pd.notna(a.get("unit_url")) else "",
            })

        desc_raw = next((str(meta[c]) for c in ["description","property_description","descripcion","desc","comments"]
                         if c in d.columns and pd.notna(meta.get(c))), "")
        _strip_lines = [
            "This comment was automatically translated and may not be 100% accurate.",
            "See description in the original language",
        ]
        for _sl in _strip_lines:
            desc_raw = desc_raw.replace(_sl, "")
        desc_raw = desc_raw.strip()
        _lat, _lng, _ = _listing_coords(lid, str(meta.get("municipality","")))
        _lat_exact = _lat is not None
        if _lat is None: _lat = 39.47
        if _lng is None: _lng = -0.38
        listings_data.append({
            "property_name": str(meta.get("property_name","")),
            "developer":     str(meta.get("developer","")),
            "city_area":     str(meta.get("city_area","")) if pd.notna(meta.get("city_area")) else "",
            "municipality":  str(meta.get("municipality","")),
            "province":      str(meta.get("province","")),
            "house_type":    ht,
            "delivery_date": str(meta.get("delivery_date","")).replace("Delivery : ",""),
            "esg":           esg,
            "amenities":     "; ".join(amenities),
            "description":   desc_raw,
            "total_units":   len(apts),
            "apartments":    apts,
            "lat":           round(_lat, 6),
            "lng":           round(_lng, 6),
            "lat_exact":     _lat_exact,
        })

    # ── Build Excel ────────────────────────────────────────────────────────
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "New Construction Projects"

    from openpyxl.styles import Border, Side

    NAVY  = "0B1239"
    DGREY = "3A4A6B"
    LGREY = "F2F4F6"

    FONT_NAME = "Aptos Narrow"
    def fill(hex_): return PatternFill("solid", fgColor=hex_)
    def font(bold=False, color="000000", size=10, italic=False, underline=None):
        return Font(name=FONT_NAME, bold=bold, color=color, size=size, italic=italic, underline=underline)
    def aln(h="left", v="center", wrap=False):
        return Alignment(horizontal=h, vertical=v, wrap_text=wrap)
    thin = Side(style="thin", color="D0D4DE")
    def border(): return Border(bottom=thin)
    def col(n): return get_column_letter(n)

    # Column layout:
    # A-E  (1-5):  Property Name, Developer, City Area, Municipality, Province       [listing info]
    # F-M  (6-13): Unit Type, House Type, Price (€), €/m², Size (m²), Floor, Beds, Bath [per-unit]
    # N-R  (14-18): Delivery, ESG, Total Units, Latitude, Longitude                  [listing tail]
    # S    (19):   Link                                                               [per-unit]
    # T    (20):   Description                                                        [merged rows]
    NL   = 5   # listing prefix cols (A-E)
    NAPT = 8   # per-unit cols (F-M): Unit Type, House Type, Price, €/m², Size (m²), Floor, Beds, Bath
    NTAIL= 5   # listing tail cols (N-R): Delivery, ESG, Total Units, Lat, Lng
    NCOLS = NL + NAPT + NTAIL + 1 + 1  # = 20  (+Link +Description)
    COL_LINK = NL + NAPT + NTAIL + 1   # col 19 = S
    COL_DESC = NL + NAPT + NTAIL + 2   # col 20 = T

    APT_HEADERS  = ["Unit Type","House Type","Price (€)","€/m²","Size (m²)","Floor","Beds","Bath"]
    TAIL_HEADERS = ["Delivery","ESG","Total Units","Latitude","Longitude"]

    import datetime as _dt
    export_date = _dt.datetime.now().strftime("%d %B %Y")

    # ── Row 1: Title ──────────────────────────────────────────────────────
    ws.merge_cells(f"A1:{col(NCOLS)}1")
    c = ws["A1"]
    c.value = "New Construction Projects - Spain Housing Intelligence"
    c.font = Font(name=FONT_NAME, bold=True, size=15, color=NAVY)
    c.alignment = aln(); ws.row_dimensions[1].height = 28

    # ── Row 2: Subheading ─────────────────────────────────────────────────
    ws.merge_cells(f"A2:{col(NCOLS)}2")
    c = ws["A2"]
    c.value = f"Residential new-build developments | Unit-level pricing & specifications | Exported {export_date}"
    c.font = Font(name=FONT_NAME, size=10, italic=True, color="6B7A9F")
    c.fill = fill("F5F7FC")
    c.alignment = aln()
    ws.row_dimensions[2].height = 18

    # ── Row 3: Column headers ─────────────────────────────────────────────
    ws.row_dimensions[3].height = 30
    listing_prefix = ["Property Name","Developer","City Area","Municipality","Province"]
    all_headers = listing_prefix + APT_HEADERS + TAIL_HEADERS + ["Link","Description"]
    for i, h in enumerate(all_headers):
        c = ws[f"{col(i+1)}3"]
        c.value = h
        c.font = font(bold=True, color="FFFFFF", size=9)
        if i < NL:
            c.fill = fill(NAVY)
        elif i < NL + NAPT:
            c.fill = fill(DGREY)
        elif i < NL + NAPT + NTAIL + 1:   # tail + link
            c.fill = fill("1A3060")
        else:                               # description
            c.fill = fill("2A3F6F")
        c.alignment = aln("center", wrap=True)
        c.border = border()
    ws.freeze_panes = "A4"

    # ── Data rows ─────────────────────────────────────────────────────────
    row = 4
    for li, ld in enumerate(listings_data):
        apts  = ld["apartments"]
        n_apt = len(apts)
        bg_l  = fill("2A3F6F")
        bg_a0 = fill(LGREY)
        bg_a1 = fill("FFFFFF")

        lat_label = str(ld["lat"]) if ld["lat_exact"] else f"~{ld['lat']} (approx)"
        lng_label = str(ld["lng"]) if ld["lat_exact"] else f"~{ld['lng']} (approx)"

        prefix_vals = [
            ld["property_name"], ld["developer"], ld["city_area"],
            ld["municipality"],  ld["province"],
        ]
        tail_vals = [
            ld["delivery_date"], ld["esg"], f'{n_apt} units',
            lat_label, lng_label,
        ]

        # ── Listing header row (full width navy) ──────────────────────────
        ws.row_dimensions[row].height = 18
        for ci, v in enumerate(prefix_vals):
            c = ws[f"{col(ci+1)}{row}"]
            c.value = v
            c.font  = font(bold=True, color="FFFFFF", size=10)
            c.fill  = bg_l; c.alignment = aln("left")
        for ci in range(NL, NL + NAPT):
            ws[f"{col(ci+1)}{row}"].fill = bg_l
        for ci, v in enumerate(tail_vals):
            c = ws[f"{col(NL + NAPT + ci + 1)}{row}"]
            c.value = v
            c.font  = font(bold=True, color="FFFFFF", size=10)
            c.fill  = bg_l; c.alignment = aln("left")
        # Link col (header) — amenities note
        c = ws[f"{col(COL_LINK)}{row}"]
        c.value = f'Amenities: {ld["amenities"]}' if ld["amenities"] else ""
        c.font  = font(bold=False, color="C5CBE9", size=9, italic=True)
        c.fill  = bg_l
        # Description col (header) — blank, will be merged across unit rows below
        ws[f"{col(COL_DESC)}{row}"].fill = bg_l
        header_row = row
        row += 1

        # ── Apartment rows ────────────────────────────────────────────────
        apt_start_row = row
        for ai, apt in enumerate(apts):
            ws.row_dimensions[row].height = 14
            bg = bg_a0 if ai % 2 == 0 else bg_a1

            # Prefix cols A-E
            for ci, v in enumerate(prefix_vals):
                c = ws[f"{col(ci+1)}{row}"]
                c.value = v if ci == 0 else ""
                c.font  = font(color="9AA0B4", size=9)
                c.fill  = bg; c.border = border()

            # Per-unit cols F-M: Unit Type, House Type, Price, €/m², Size (m²), Floor, Beds, Bath
            apt_vals = [
                apt["unit_type"], apt["house_type"],
                apt["price"],     apt["pm2"],
                apt["size"],
                apt["floor"],     apt["bedrooms"], apt["bathrooms"],
            ]
            for ci, v in enumerate(apt_vals):
                c = ws[f"{col(NL + ci + 1)}{row}"]
                c.fill = bg; c.border = border()
                is_price = ci == 2; is_pm2 = ci == 3; is_size = ci == 4
                c.alignment = aln("right" if ci in (2,3,4,6,7) else "left")
                if is_price and v:
                    c.value = v; c.number_format = "#,##0"
                    c.font  = font(bold=True, color=NAVY, size=10)
                elif is_pm2 and v:
                    c.value = v; c.number_format = "#,##0"
                    c.font  = font(color=DGREY, size=10)
                elif is_size and v:
                    c.value = v; c.number_format = "#,##0.0"
                    c.font  = font(color=DGREY, size=10)
                elif ci == 0:
                    c.value = v; c.font = font(bold=True, color=NAVY, size=10)
                else:
                    c.value = v; c.font = font(color=DGREY, size=9)

            # Tail cols M-Q (show only on first unit row)
            for ci, v in enumerate(tail_vals):
                c = ws[f"{col(NL + NAPT + ci + 1)}{row}"]
                c.value = v if ai == 0 else ""
                c.font  = font(color="9AA0B4", size=9)
                c.fill  = bg; c.border = border(); c.alignment = aln("left")

            # Link col R
            c = ws[f"{col(COL_LINK)}{row}"]
            c.fill = bg; c.border = border()
            if apt["url"]:
                c.value = apt["url"]; c.hyperlink = apt["url"]
                c.font  = Font(name=FONT_NAME, color="0563C1", underline="single", size=9)

            # Description col S — filled but content added via merge below
            ws[f"{col(COL_DESC)}{row}"].fill = bg_a0; ws[f"{col(COL_DESC)}{row}"].border = border()
            row += 1

        apt_end_row = row - 1

        # ── Description: merged rows in last column (S), spanning all unit rows + header ──
        if ld.get("description") and n_apt > 0:
            merge_start = header_row
            merge_end   = apt_end_row
            if merge_end > merge_start:
                ws.merge_cells(f"{col(COL_DESC)}{merge_start}:{col(COL_DESC)}{merge_end}")
            c = ws[f"{col(COL_DESC)}{merge_start}"]
            c.value = ld["description"]
            c.font  = Font(name=FONT_NAME, size=9, italic=True, color="6B7A9F")
            c.fill  = fill("F5F7FC")
            c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

        # Blank separator row
        for ci in range(NCOLS):
            ws[f"{col(ci+1)}{row}"].fill = fill("FFFFFF")
        row += 1

    # ── Column widths ─────────────────────────────────────────────────────
    # A-E: listing prefix, F-M: per-unit, N-R: tail, S: link, T: description
    widths = [28, 20, 18, 16, 12,   # A-E
              11, 16, 11,  9,  9, 8, 6, 6,  # F-M
              14,  8, 11, 13, 13,        # N-R
              35, 50]                    # S (link), T (description)
    for i, w in enumerate(widths):
        ws.column_dimensions[col(i+1)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=new_construction_projects.xlsx"})


_FALLBACK_SOLD_DATE = "4 Apr 2026"  # fallback when no expired_listing entry found

def _listing_sold_date(listing_id: int) -> str:
    """Return the sold date for a listing: latest removed_date among its sub-listings, or fallback."""
    sub_ids = df[df["listing_id"] == listing_id]["sub_listing_id"].dropna().astype(int).unique()
    dates = [_sub_to_sold_date[s] for s in sub_ids if s in _sub_to_sold_date]
    if not dates:
        return _FALLBACK_SOLD_DATE
    # Return the most recent date
    import datetime as _dt
    def _parse(d):
        try: return _dt.datetime.strptime(d, "%d %b %Y")
        except: return _dt.datetime.min
    return max(dates, key=_parse)

@app.get("/delisted/export")
def export_delisted_excel():
    """Export all sold-out properties: Sheet 1 = listings, Sheet 2 = apartments."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        return JSONResponse({"error": "openpyxl not installed"}, status_code=500)

    import datetime as _dt
    export_date = _dt.datetime.now().strftime("%d %B %Y")

    d = df.copy()
    has_latest     = set(d[d["_is_latest"] & d["sub_listing_id"].notna()]["listing_id"].unique())
    has_non_latest = set(d[~d["_is_latest"] & d["sub_listing_id"].notna()]["listing_id"].unique())
    delisted_ids   = has_non_latest - has_latest
    if not delisted_ids:
        return JSONResponse({"error": "No sold-out properties found"}, status_code=404)

    d_non_latest = d[(d["listing_id"].isin(delisted_ids)) & (~d["_is_latest"])]
    max_ords = d_non_latest.groupby("listing_id")["period_ord"].max().reset_index().rename(columns={"period_ord":"_max_ord"})
    d_snap   = d_non_latest.merge(max_ords, on="listing_id")
    d_snap   = d_snap[d_snap["period_ord"] == d_snap["_max_ord"]]

    NAVY   = "0B1239"
    RED    = "7F1D1D"
    LRED   = "FEF2F2"
    DGREY  = "3A4A6B"
    LGREY  = "F2F4F6"
    STRIPE = "EBF0F7"
    FONT   = "Aptos Narrow"

    def fill(hex_): return PatternFill("solid", fgColor=hex_)
    def font(bold=False, color="000000", size=10, italic=False, underline=None):
        return Font(name=FONT, bold=bold, color=color, size=size, italic=italic, underline=underline)
    def aln(h="left", v="center", wrap=False):
        return Alignment(horizontal=h, vertical=v, wrap_text=wrap)
    thin = Side(style="thin", color="D0D4DE")
    def bdr(): return Border(bottom=thin)
    def col(n): return get_column_letter(n)

    wb = openpyxl.Workbook()

    # ══════════════════════════════════════════════════════════════
    # SHEET 1 — Sold Out Listings (one row per development)
    # ══════════════════════════════════════════════════════════════
    ws1 = wb.active
    ws1.title = "Sold Out Properties"

    LISTING_COLS = ["Property Name","Developer","Municipality","Province","City Area",
                    "House Type","ESG","Last Seen Period","Sold Date","Total Units",
                    "Avg Price (€)","Avg €/m²","Min Price (€)","Max Price (€)","Avg Size (m²)"]
    NC1 = len(LISTING_COLS)

    # Title
    ws1.merge_cells(f"A1:{col(NC1)}1")
    c = ws1["A1"]
    c.value = "Sold Out Properties — Spain Housing Intelligence"
    c.font = Font(name=FONT, bold=True, size=14, color=RED)
    c.alignment = aln(); ws1.row_dimensions[1].height = 26

    ws1.merge_cells(f"A2:{col(NC1)}2")
    c = ws1["A2"]
    c.value = f"Developments no longer listed as of {LATEST_PERIOD} | Exported {export_date}"
    c.font = Font(name=FONT, size=10, italic=True, color=DGREY)
    c.fill = fill("FFF1F1"); c.alignment = aln(); ws1.row_dimensions[2].height = 16

    # Headers
    ws1.row_dimensions[3].height = 26
    for i, h in enumerate(LISTING_COLS):
        c = ws1[f"{col(i+1)}3"]
        c.value = h; c.font = font(bold=True, color="FFFFFF", size=9)
        c.fill = fill(RED); c.alignment = aln("center", wrap=True); c.border = bdr()
    ws1.freeze_panes = "A4"

    # Data
    listings_meta = d_snap.groupby("listing_id").first().reset_index()
    listings_agg  = d_snap.groupby("listing_id").agg(
        units=("sub_listing_id","nunique"),
        avg_price=("price","mean"),
        avg_pm2=("price_per_m2","mean"),
        min_price=("price","min"),
        max_price=("price","max"),
        avg_size=("size","mean"),
        last_period=("period","max"),
    ).reset_index()
    listings_meta = listings_meta.merge(listings_agg, on="listing_id", suffixes=("","_agg"))

    row = 4
    for i, r_ in listings_meta.iterrows():
        bg = fill(LGREY) if i % 2 == 0 else fill("FFFFFF")
        ht = str(r_.get("house_type","")) if pd.notna(r_.get("house_type")) else ""
        vals = [
            str(r_.get("property_name","")),
            str(r_.get("developer","")) if pd.notna(r_.get("developer")) else "",
            str(r_.get("municipality","")),
            str(r_.get("province","")) if pd.notna(r_.get("province")) else "",
            str(r_.get("city_area","")) if pd.notna(r_.get("city_area")) else "",
            ht,
            str(r_.get("esg_grade","")) if pd.notna(r_.get("esg_grade")) else "",
            str(r_.get("last_period","")),
            _listing_sold_date(int(r_["listing_id"])),
            int(r_.get("units",0)) if pd.notna(r_.get("units")) else 0,
            int(round(r_.get("avg_price",0))) if pd.notna(r_.get("avg_price")) else None,
            int(round(r_.get("avg_pm2",0))) if pd.notna(r_.get("avg_pm2")) else None,
            int(round(r_.get("min_price",0))) if pd.notna(r_.get("min_price")) else None,
            int(round(r_.get("max_price",0))) if pd.notna(r_.get("max_price")) else None,
            round(float(r_.get("avg_size",0)),1) if pd.notna(r_.get("avg_size")) else None,
        ]
        ws1.row_dimensions[row].height = 16
        for ci, v in enumerate(vals):
            c = ws1[f"{col(ci+1)}{row}"]
            c.value = v; c.fill = bg; c.border = bdr()
            is_price = ci in (10,11,12,13)
            c.alignment = aln("right" if ci >= 9 else "left")
            if is_price and v:
                c.number_format = "#,##0"
                c.font = font(color=RED if ci in (10,11) else DGREY, size=9)
            elif ci == 8:  # Sold Date
                c.font = font(bold=True, color=RED, size=9)
            else:
                c.font = font(color=DGREY, size=9)
        row += 1

    widths1 = [28,18,16,12,18,14,6,12,12,10,13,10,13,13,12]
    for i, w in enumerate(widths1):
        ws1.column_dimensions[col(i+1)].width = w

    # ══════════════════════════════════════════════════════════════
    # SHEET 2 — Sold Out Sub-listings (one row per apartment)
    # ══════════════════════════════════════════════════════════════
    ws2 = wb.create_sheet("Sold Out Units")

    APT_COLS = ["Property Name","Municipality","Sold Date","Unit ID","Unit Type",
                "Floor","Size (m²)","Bedrooms","Bathrooms","Price (€)","€/m²",
                "Terrace","Parking","Pool","Lift","AC","Link"]
    NC2 = len(APT_COLS)

    ws2.merge_cells(f"A1:{col(NC2)}1")
    c = ws2["A1"]
    c.value = "Sold Out Units — Spain Housing Intelligence"
    c.font = Font(name=FONT, bold=True, size=14, color=RED)
    c.alignment = aln(); ws2.row_dimensions[1].height = 26

    ws2.merge_cells(f"A2:{col(NC2)}2")
    c = ws2["A2"]
    c.value = f"Individual apartment records for all sold-out developments | Sold Date: {LATEST_PERIOD} | Exported {export_date}"
    c.font = Font(name=FONT, size=10, italic=True, color=DGREY)
    c.fill = fill("FFF1F1"); c.alignment = aln(); ws2.row_dimensions[2].height = 16

    ws2.row_dimensions[3].height = 26
    for i, h in enumerate(APT_COLS):
        c = ws2[f"{col(i+1)}3"]
        c.value = h; c.font = font(bold=True, color="FFFFFF", size=9)
        c.fill = fill(RED); c.alignment = aln("center", wrap=True); c.border = bdr()
    ws2.freeze_panes = "A4"

    apt_cols_need = ["listing_id","sub_listing_id","property_name","municipality",
                     "unit_type","floor","size","bedrooms","bathrooms","price","price_per_m2",
                     "has_terrace","has_parking","has_pool","has_lift","has_ac","unit_url"]
    apts_df = d_snap[[c_ for c_ in apt_cols_need if c_ in d_snap.columns]].drop_duplicates("sub_listing_id").copy()
    apts_df = apts_df.sort_values(["listing_id","unit_type","price"])

    row = 4
    for i, r_ in apts_df.iterrows():
        bg = fill(LGREY) if i % 2 == 0 else fill("FFFFFF")
        def yn(col_): return "Yes" if pd.notna(r_.get(col_)) and bool(r_.get(col_)) else "No"
        url = str(r_.get("unit_url","")) if "unit_url" in r_.index and pd.notna(r_.get("unit_url")) else ""
        sub_id = int(r_["sub_listing_id"]) if "sub_listing_id" in r_.index and pd.notna(r_.get("sub_listing_id")) else None
        apt_sold = _sub_to_sold_date.get(sub_id, _FALLBACK_SOLD_DATE) if sub_id else _FALLBACK_SOLD_DATE
        vals = [
            str(r_.get("property_name","")),
            str(r_.get("municipality","")),
            apt_sold,
            str(r_.get("sub_listing_id","")) if pd.notna(r_.get("sub_listing_id")) else "",
            str(r_.get("unit_type","")) if pd.notna(r_.get("unit_type")) else "",
            str(r_.get("floor","")) if pd.notna(r_.get("floor")) else "",
            round(float(r_.get("size",0)),1) if pd.notna(r_.get("size")) else None,
            int(r_["bedrooms"]) if "bedrooms" in r_.index and pd.notna(r_.get("bedrooms")) else None,
            int(r_["bathrooms"]) if "bathrooms" in r_.index and pd.notna(r_.get("bathrooms")) else None,
            int(round(r_["price"])) if pd.notna(r_.get("price")) else None,
            int(round(r_["price_per_m2"])) if "price_per_m2" in r_.index and pd.notna(r_.get("price_per_m2")) else None,
            yn("has_terrace"), yn("has_parking"), yn("has_pool"), yn("has_lift"), yn("has_ac"),
            url,
        ]
        ws2.row_dimensions[row].height = 14
        for ci, v in enumerate(vals):
            c = ws2[f"{col(ci+1)}{row}"]
            c.value = v; c.fill = bg; c.border = bdr()
            is_num = ci in (6,7,8,9,10)
            c.alignment = aln("right" if is_num else "left")
            if ci in (9,10) and v:
                c.number_format = "#,##0"
                c.font = font(bold=(ci==9), color=RED if ci==9 else DGREY, size=9)
            elif ci == 2:  # Sold Date
                c.font = font(bold=True, color=RED, size=9)
            elif ci == 16 and url:  # Link
                c.hyperlink = url
                c.font = Font(name=FONT, color="0563C1", underline="single", size=9)
            else:
                c.font = font(color=DGREY, size=9)
        row += 1

    widths2 = [28,14,12,12,10,8,9,7,7,13,10,8,8,6,6,6,40]
    for i, w in enumerate(widths2):
        ws2.column_dimensions[col(i+1)].width = w

    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    return StreamingResponse(buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=sold_out_properties.xlsx"})


@app.get("/export/by-filter")
def export_by_filter(
    provinces:     Optional[List[str]] = Query(None),
    municipalities: Optional[List[str]] = Query(None),
):
    """Export all listings matching given provinces and/or municipalities."""
    d = df[df["_is_latest"]]
    if provinces:
        d = d[d["province"].isin(provinces)]
    if municipalities:
        d = d[d["municipality"].isin(municipalities)]

    listing_ids = sorted(d["listing_id"].dropna().unique().astype(int).tolist())
    if not listing_ids:
        return JSONResponse({"error": "No listings found"}, status_code=404)

    ids_str = ",".join(str(i) for i in listing_ids)
    # Reuse the existing export function
    from fastapi import Request
    class _FakeQuery:
        pass
    return export_listings_excel(ids=ids_str)


@app.get("/summary/export")
def export_summary_excel(
    municipality: Optional[List[str]] = Query(None),
    province:     Optional[List[str]] = Query(None),
    unit_type:    Optional[List[str]] = Query(None),
    year:         Optional[List[str]] = Query(None),
    esg:          Optional[List[str]] = Query(None),
    house_type:   Optional[List[str]] = Query(None),
):
    """Sheet 1 = active listings in analysis format; Sheet 2 = sold-out listings."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        return JSONResponse({"error": "openpyxl not installed"}, status_code=500)

    import datetime as _dt
    export_date = _dt.datetime.now().strftime("%d %B %Y")

    NAVY  = "0B1239"; DGREY = "3A4A6B"; LGREY = "F2F4F6"
    RED   = "7F1D1D"; FONT  = "Aptos Narrow"
    AME_COLS = [("has_lift","Lift"),("has_parking","Parking"),("has_pool","Pool"),
                ("has_garden","Garden"),("has_ac","AC"),("has_storage","Storage"),
                ("has_terrace","Terrace"),("has_wardrobes","Wardrobes")]
    UT_ORDER = {"Studio":0,"1BR":1,"2BR":2,"3BR":3,"4BR":4,"5BR":5,"Penthouse":6}

    def fill(hex_): return PatternFill("solid", fgColor=hex_)
    def font(bold=False, color="000000", size=10, italic=False, underline=None):
        return Font(name=FONT, bold=bold, color=color, size=size, italic=italic, underline=underline)
    def aln(h="left", v="center", wrap=False):
        return Alignment(horizontal=h, vertical=v, wrap_text=wrap)
    thin = Side(style="thin", color="D0D4DE")
    def border(): return Border(bottom=thin)
    def col(n): return get_column_letter(n)

    # ── Collect active listing IDs from filters ───────────────────────────
    d_active = _filter(municipality, unit_type, year, esg, None, province, house_type=house_type)
    active_ids = sorted(d_active["listing_id"].dropna().unique().astype(int).tolist())

    # ── Build per-listing data (same logic as /search/export) ─────────────
    listings_data = []
    for lid in active_ids:
        d = df[df["listing_id"] == lid]
        if d.empty: continue
        latest = d[d["_is_latest"]]
        if latest.empty: latest = d
        meta = latest.iloc[0]
        esg_val = str(meta.get("esg_certificate","")) if pd.notna(meta.get("esg_certificate","")) else ""
        ht = ", ".join(sorted(t for t in latest["house_type"].dropna().unique() if t)) if "house_type" in latest.columns else ""
        amenities = [lbl for c_,lbl in AME_COLS if c_ in latest.columns and latest[c_].any()]
        apt_cols = ["sub_listing_id","unit_type","price","size","price_per_m2","floor","floor_num","bedrooms","bathrooms","unit_url"]
        for ac in ["house_type","has_terrace","has_parking","has_pool","has_garden","has_lift","has_ac","has_storage","has_wardrobes"]:
            if ac in latest.columns: apt_cols.append(ac)
        apts_df = latest[[c_ for c_ in apt_cols if c_ in latest.columns]].drop_duplicates("sub_listing_id").copy()
        apts_df["_s"] = apts_df["unit_type"].map(lambda x: UT_ORDER.get(x, 99))
        apts_df = apts_df.sort_values(["_s","price"]).drop("_s", axis=1)
        apts = []
        for _, a in apts_df.iterrows():
            if pd.isna(a.get("sub_listing_id")): continue
            apt_ame = [lbl for c_,lbl in AME_COLS if c_ in a.index and bool(a.get(c_))]
            apts.append({
                "unit_type":  str(a.get("unit_type","")) if pd.notna(a.get("unit_type")) else "",
                "house_type": str(a.get("house_type","")) if "house_type" in a.index and pd.notna(a.get("house_type")) else "",
                "floor":      str(a.get("floor","")) if pd.notna(a.get("floor")) else "",
                "bedrooms":   int(a["bedrooms"])  if pd.notna(a.get("bedrooms"))  else None,
                "bathrooms":  int(a["bathrooms"]) if pd.notna(a.get("bathrooms")) else None,
                "size":       round(float(a["size"]),1) if pd.notna(a.get("size")) else None,
                "price":      int(round(a["price"])) if pd.notna(a.get("price")) else None,
                "pm2":        int(round(a["price_per_m2"])) if pd.notna(a.get("price_per_m2")) else None,
                "amenities":  "; ".join(apt_ame),
                "url":        str(a["unit_url"]) if "unit_url" in a.index and pd.notna(a.get("unit_url")) else "",
            })
        if not apts: continue
        _lat, _lng, _ = _listing_coords(lid, str(meta.get("municipality","")))
        _lat_exact = _lat is not None
        if _lat is None: _lat = 39.47
        if _lng is None: _lng = -0.38
        desc_raw = next((str(meta[c]) for c in ["description","property_description","descripcion","desc","comments"]
                         if c in d.columns and pd.notna(meta.get(c))), "")
        for _sl in ["This comment was automatically translated and may not be 100% accurate.",
                    "See description in the original language"]:
            desc_raw = desc_raw.replace(_sl, "")
        listings_data.append({
            "property_name": str(meta.get("property_name","")),
            "developer":     str(meta.get("developer","")) if pd.notna(meta.get("developer")) else "",
            "city_area":     str(meta.get("city_area","")) if pd.notna(meta.get("city_area")) else "",
            "municipality":  str(meta.get("municipality","")),
            "province":      str(meta.get("province","")),
            "house_type":    ht,
            "delivery_date": str(meta.get("delivery_date","")).replace("Delivery : ",""),
            "esg":           esg_val,
            "amenities":     "; ".join(amenities),
            "description":   desc_raw.strip(),
            "total_units":   len(apts),
            "apartments":    apts,
            "lat":           round(_lat, 6), "lng": round(_lng, 6),
            "lat_exact":     _lat_exact,
        })

    # ── Sheet 1: Active listings (analysis format) ────────────────────────
    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "Active Listings"

    NL=5; NAPT=8; NTAIL=5
    NCOLS = NL + NAPT + NTAIL + 2
    COL_LINK = NL + NAPT + NTAIL + 1
    COL_DESC = NL + NAPT + NTAIL + 2

    ws1.merge_cells(f"A1:{col(NCOLS)}1")
    c = ws1["A1"]
    c.value = "New Construction Projects — Spain Housing Intelligence"
    c.font = Font(name=FONT, bold=True, size=15, color=NAVY)
    c.alignment = aln(); ws1.row_dimensions[1].height = 28

    ws1.merge_cells(f"A2:{col(NCOLS)}2")
    c = ws1["A2"]
    c.value = f"Residential new-build developments | Unit-level pricing & specifications | Snapshot: {LATEST_PERIOD} | Exported {export_date}"
    c.font = Font(name=FONT, size=10, italic=True, color="6B7A9F")
    c.fill = fill("F5F7FC"); c.alignment = aln(); ws1.row_dimensions[2].height = 18

    ws1.row_dimensions[3].height = 30
    all_headers = ["Property Name","Developer","City Area","Municipality","Province",
                   "Unit Type","House Type","Price (€)","€/m²","Size (m²)","Floor","Beds","Bath",
                   "Delivery","ESG","Total Units","Latitude","Longitude","Link","Description"]
    for i, h in enumerate(all_headers):
        c = ws1[f"{col(i+1)}3"]
        c.value = h; c.font = font(bold=True, color="FFFFFF", size=9)
        if i < NL:              c.fill = fill(NAVY)
        elif i < NL+NAPT:       c.fill = fill(DGREY)
        elif i < NL+NAPT+NTAIL+1: c.fill = fill("1A3060")
        else:                   c.fill = fill("2A3F6F")
        c.alignment = aln("center", wrap=True); c.border = border()
    ws1.freeze_panes = "A4"

    row = 4
    for ld in listings_data:
        apts = ld["apartments"]; n_apt = len(apts)
        bg_l = fill("2A3F6F"); bg_a0 = fill(LGREY); bg_a1 = fill("FFFFFF")
        lat_lbl = str(ld["lat"]) if ld["lat_exact"] else f"~{ld['lat']} (approx)"
        lng_lbl = str(ld["lng"]) if ld["lat_exact"] else f"~{ld['lng']} (approx)"
        prefix_vals = [ld["property_name"],ld["developer"],ld["city_area"],ld["municipality"],ld["province"]]
        tail_vals   = [ld["delivery_date"],ld["esg"],f"{n_apt} units",lat_lbl,lng_lbl]
        ws1.row_dimensions[row].height = 18
        for ci, v in enumerate(prefix_vals):
            c = ws1[f"{col(ci+1)}{row}"]
            c.value = v; c.font = font(bold=True, color="FFFFFF", size=10); c.fill = bg_l; c.alignment = aln("left")
        for ci in range(NL, NL+NAPT):
            ws1[f"{col(ci+1)}{row}"].fill = bg_l
        for ci, v in enumerate(tail_vals):
            c = ws1[f"{col(NL+NAPT+ci+1)}{row}"]
            c.value = v; c.font = font(bold=True, color="FFFFFF", size=10); c.fill = bg_l; c.alignment = aln("left")
        c = ws1[f"{col(COL_LINK)}{row}"]
        c.value = f'Amenities: {ld["amenities"]}' if ld["amenities"] else ""
        c.font = font(bold=False, color="C5CBE9", size=9, italic=True); c.fill = bg_l
        ws1[f"{col(COL_DESC)}{row}"].fill = bg_l
        header_row = row; row += 1
        apt_start = row
        for ai, apt in enumerate(apts):
            ws1.row_dimensions[row].height = 14
            bg = bg_a0 if ai % 2 == 0 else bg_a1
            for ci, v in enumerate(prefix_vals):
                c = ws1[f"{col(ci+1)}{row}"]
                c.value = v if ci == 0 else ""; c.font = font(color="9AA0B4", size=9); c.fill = bg; c.border = border()
            apt_vals = [apt["unit_type"],apt["house_type"],apt["price"],apt["pm2"],apt["size"],apt["floor"],apt["bedrooms"],apt["bathrooms"]]
            for ci, v in enumerate(apt_vals):
                c = ws1[f"{col(NL+ci+1)}{row}"]
                c.fill = bg; c.border = border()
                c.alignment = aln("right" if ci in (2,3,4,6,7) else "left")
                if ci == 2 and v:   c.value = v; c.number_format = "#,##0"; c.font = font(bold=True, color=NAVY, size=10)
                elif ci == 3 and v: c.value = v; c.number_format = "#,##0"; c.font = font(color=DGREY, size=10)
                elif ci == 4 and v: c.value = v; c.number_format = "#,##0.0"; c.font = font(color=DGREY, size=10)
                elif ci == 0:       c.value = v; c.font = font(bold=True, color=NAVY, size=10)
                else:               c.value = v; c.font = font(color=DGREY, size=9)
            for ci, v in enumerate(tail_vals):
                c = ws1[f"{col(NL+NAPT+ci+1)}{row}"]
                c.value = v if ai == 0 else ""; c.font = font(color="9AA0B4", size=9)
                c.fill = bg; c.border = border(); c.alignment = aln("left")
            c = ws1[f"{col(COL_LINK)}{row}"]
            c.fill = bg; c.border = border()
            if apt["url"]:
                c.value = apt["url"]; c.hyperlink = apt["url"]
                c.font = Font(name=FONT, color="0563C1", underline="single", size=9)
            ws1[f"{col(COL_DESC)}{row}"].fill = bg_a0; ws1[f"{col(COL_DESC)}{row}"].border = border()
            row += 1
        apt_end = row - 1
        if ld.get("description") and n_apt > 0:
            if apt_end > header_row:
                ws1.merge_cells(f"{col(COL_DESC)}{header_row}:{col(COL_DESC)}{apt_end}")
            c = ws1[f"{col(COL_DESC)}{header_row}"]
            c.value = ld["description"]; c.font = Font(name=FONT, size=9, italic=True, color="6B7A9F")
            c.fill = fill("F5F7FC"); c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        for ci in range(NCOLS):
            ws1[f"{col(ci+1)}{row}"].fill = fill("FFFFFF")
        row += 1

    for i, w in enumerate([28,20,18,16,12, 11,16,11,9,9,8,6,6, 14,8,11,13,13, 35,50]):
        ws1.column_dimensions[col(i+1)].width = w

    # ── Sheet 2: Sold-out listings ────────────────────────────────────────
    ws2 = wb.create_sheet("Sold Out")

    d_all = df.copy()
    if province:     d_all = d_all[d_all["province"].isin(province)]
    if municipality: d_all = d_all[d_all["municipality"].isin(municipality)]
    if unit_type:    d_all = d_all[d_all["unit_type"].isin(unit_type)]
    if year:         d_all = d_all[d_all["delivery_year"].isin([int(y) for y in year])]
    if esg:          d_all = d_all[d_all["esg_grade"].isin(esg)]
    if house_type and "house_type" in d_all.columns:
        d_all = d_all[d_all["house_type"].isin(house_type)]

    has_latest     = set(d_all[d_all["_is_latest"] & d_all["sub_listing_id"].notna()]["listing_id"].unique())
    has_non_latest = set(d_all[~d_all["_is_latest"] & d_all["sub_listing_id"].notna()]["listing_id"].unique())
    sold_out_ids   = has_non_latest - has_latest

    S_COLS = ["Property Name","Developer","Municipality","Province","City Area",
              "House Type","ESG","Last Seen Period","Sold Date","Total Units",
              "Avg Price (€)","Avg €/m²","Min Price (€)","Max Price (€)","Avg Size (m²)"]
    NS = len(S_COLS)

    ws2.merge_cells(f"A1:{col(NS)}1")
    c = ws2["A1"]
    c.value = "Sold Out Properties — Spain Housing Intelligence"
    c.font = Font(name=FONT, bold=True, size=14, color=RED)
    c.alignment = aln(); ws2.row_dimensions[1].height = 26

    ws2.merge_cells(f"A2:{col(NS)}2")
    c = ws2["A2"]
    c.value = f"Developments no longer listed as of {LATEST_PERIOD} | Exported {export_date}"
    c.font = Font(name=FONT, size=10, italic=True, color=DGREY)
    c.fill = fill("FFF1F1"); c.alignment = aln(); ws2.row_dimensions[2].height = 16

    ws2.row_dimensions[3].height = 26
    for i, h in enumerate(S_COLS):
        c = ws2[f"{col(i+1)}3"]
        c.value = h; c.font = font(bold=True, color="FFFFFF", size=9)
        c.fill = fill(RED); c.alignment = aln("center", wrap=True); c.border = border()
    ws2.freeze_panes = "A4"

    d_snap = pd.DataFrame()
    if sold_out_ids:
        d_nl = d_all[~d_all["_is_latest"] & d_all["listing_id"].isin(sold_out_ids)]
        max_ords = d_nl.groupby("listing_id")["period_ord"].max().reset_index().rename(columns={"period_ord":"_mo"})
        d_snap = d_nl.merge(max_ords, on="listing_id")
        d_snap = d_snap[d_snap["period_ord"] == d_snap["_mo"]]
        s_meta = d_snap.groupby("listing_id").first().reset_index()
        s_agg  = d_snap.groupby("listing_id").agg(
            units=("sub_listing_id","nunique"), avg_price=("price","mean"),
            avg_pm2=("price_per_m2","mean"), min_price=("price","min"),
            max_price=("price","max"), avg_size=("size","mean"), last_period=("period","max"),
        ).reset_index()
        s_meta = s_meta.merge(s_agg, on="listing_id", suffixes=("","_agg"))
        srow = 4
        for i, r_ in s_meta.iterrows():
            bg = fill(LGREY) if i % 2 == 0 else fill("FFFFFF")
            vals = [
                str(r_.get("property_name","")),
                str(r_.get("developer","")) if pd.notna(r_.get("developer")) else "",
                str(r_.get("municipality","")),
                str(r_.get("province","")) if pd.notna(r_.get("province")) else "",
                str(r_.get("city_area","")) if pd.notna(r_.get("city_area")) else "",
                str(r_.get("house_type","")) if pd.notna(r_.get("house_type")) else "",
                str(r_.get("esg_grade","")) if pd.notna(r_.get("esg_grade")) else "",
                str(r_.get("last_period","")),
                _listing_sold_date(int(r_["listing_id"])),
                int(r_.get("units",0)) if pd.notna(r_.get("units")) else 0,
                int(round(r_.get("avg_price",0))) if pd.notna(r_.get("avg_price")) else None,
                int(round(r_.get("avg_pm2",0)))   if pd.notna(r_.get("avg_pm2"))   else None,
                int(round(r_.get("min_price",0))) if pd.notna(r_.get("min_price")) else None,
                int(round(r_.get("max_price",0))) if pd.notna(r_.get("max_price")) else None,
                round(float(r_.get("avg_size",0)),1) if pd.notna(r_.get("avg_size")) else None,
            ]
            ws2.row_dimensions[srow].height = 16
            for ci, v in enumerate(vals):
                c = ws2[f"{col(ci+1)}{srow}"]
                c.value = v; c.fill = bg; c.border = border()
                is_price = ci in (10,11,12,13)
                c.alignment = aln("right" if ci >= 9 else "left")
                if is_price and v:
                    c.number_format = "#,##0"; c.font = font(color=RED if ci in (10,11) else DGREY, size=9)
                elif ci == 8:
                    c.font = font(bold=True, color=RED, size=9)
                else:
                    c.font = font(color=DGREY, size=9)
            srow += 1
    else:
        ws2.merge_cells(f"A4:{col(NS)}4")
        c = ws2["A4"]
        c.value = "No sold-out properties found for the selected filters."
        c.font = font(italic=True, color=DGREY, size=10); c.alignment = aln()

    for i, w in enumerate([28,18,16,12,18,14,6,12,12,10,13,10,13,13,12]):
        ws2.column_dimensions[col(i+1)].width = w

    # ── Sheet 3: Sold-out units (one row per apartment) ───────────────────
    ws3 = wb.create_sheet("Sold Out Units")

    APT_COLS = ["Property Name","Municipality","Sold Date","Unit ID","Unit Type",
                "Floor","Size (m²)","Bedrooms","Bathrooms","Price (€)","€/m²",
                "Terrace","Parking","Pool","Lift","AC","Link"]
    NC3 = len(APT_COLS)

    ws3.merge_cells(f"A1:{col(NC3)}1")
    c = ws3["A1"]
    c.value = "Sold Out Units — Spain Housing Intelligence"
    c.font = Font(name=FONT, bold=True, size=14, color=RED)
    c.alignment = aln(); ws3.row_dimensions[1].height = 26

    ws3.merge_cells(f"A2:{col(NC3)}2")
    c = ws3["A2"]
    c.value = f"Individual apartment records for all sold-out developments | Exported {export_date}"
    c.font = Font(name=FONT, size=10, italic=True, color=DGREY)
    c.fill = fill("FFF1F1"); c.alignment = aln(); ws3.row_dimensions[2].height = 16

    ws3.row_dimensions[3].height = 26
    for i, h in enumerate(APT_COLS):
        c = ws3[f"{col(i+1)}3"]
        c.value = h; c.font = font(bold=True, color="FFFFFF", size=9)
        c.fill = fill(RED); c.alignment = aln("center", wrap=True); c.border = border()
    ws3.freeze_panes = "A4"

    if not d_snap.empty:
        apt_cols_need = ["listing_id","sub_listing_id","property_name","municipality",
                         "unit_type","floor","size","bedrooms","bathrooms","price","price_per_m2",
                         "has_terrace","has_parking","has_pool","has_lift","has_ac","unit_url"]
        apts_df = d_snap[[c_ for c_ in apt_cols_need if c_ in d_snap.columns]].drop_duplicates("sub_listing_id").copy()
        apts_df = apts_df.sort_values(["listing_id","unit_type","price"])
        row3 = 4
        for i, r_ in apts_df.iterrows():
            bg = fill(LGREY) if i % 2 == 0 else fill("FFFFFF")
            def yn(col_): return "Yes" if pd.notna(r_.get(col_)) and bool(r_.get(col_)) else "No"
            url = str(r_.get("unit_url","")) if "unit_url" in r_.index and pd.notna(r_.get("unit_url")) else ""
            sub_id = int(r_["sub_listing_id"]) if "sub_listing_id" in r_.index and pd.notna(r_.get("sub_listing_id")) else None
            apt_sold = _sub_to_sold_date.get(sub_id, _FALLBACK_SOLD_DATE) if sub_id else _FALLBACK_SOLD_DATE
            vals3 = [
                str(r_.get("property_name","")),
                str(r_.get("municipality","")),
                apt_sold,
                str(r_.get("sub_listing_id","")) if pd.notna(r_.get("sub_listing_id")) else "",
                str(r_.get("unit_type","")) if pd.notna(r_.get("unit_type")) else "",
                str(r_.get("floor","")) if pd.notna(r_.get("floor")) else "",
                round(float(r_.get("size",0)),1) if pd.notna(r_.get("size")) else None,
                int(r_["bedrooms"]) if "bedrooms" in r_.index and pd.notna(r_.get("bedrooms")) else None,
                int(r_["bathrooms"]) if "bathrooms" in r_.index and pd.notna(r_.get("bathrooms")) else None,
                int(round(r_["price"])) if pd.notna(r_.get("price")) else None,
                int(round(r_["price_per_m2"])) if "price_per_m2" in r_.index and pd.notna(r_.get("price_per_m2")) else None,
                yn("has_terrace"), yn("has_parking"), yn("has_pool"), yn("has_lift"), yn("has_ac"),
                url,
            ]
            ws3.row_dimensions[row3].height = 14
            for ci, v in enumerate(vals3):
                c = ws3[f"{col(ci+1)}{row3}"]
                c.value = v; c.fill = bg; c.border = border()
                is_num = ci in (6,7,8,9,10)
                c.alignment = aln("right" if is_num else "left")
                if ci in (9,10) and v:
                    c.number_format = "#,##0"
                    c.font = font(bold=(ci==9), color=RED if ci==9 else DGREY, size=9)
                elif ci == 2:
                    c.font = font(bold=True, color=RED, size=9)
                elif ci == 16 and url:
                    c.hyperlink = url
                    c.font = Font(name=FONT, color="0563C1", underline="single", size=9)
                else:
                    c.font = font(color=DGREY, size=9)
            row3 += 1
    else:
        ws3.merge_cells(f"A4:{col(NC3)}4")
        c = ws3["A4"]
        c.value = "No sold-out units found for the selected filters."
        c.font = font(italic=True, color=DGREY, size=10); c.alignment = aln()

    for i, w in enumerate([28,14,12,12,10,8,9,7,7,13,10,8,8,6,6,6,40]):
        ws3.column_dimensions[col(i+1)].width = w

    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    return StreamingResponse(buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=market_summary.xlsx"})


@app.get("/resolve-url")
def resolve_url(url: str):
    """Follow redirects on a short URL (e.g. goo.gl/maps/...) and return the final URL."""
    try:
        req = urllib.request.Request(url, headers={"User-Agent": "Mozilla/5.0"})
        opener = urllib.request.build_opener(urllib.request.HTTPRedirectHandler())
        with opener.open(req, timeout=8) as resp:
            final_url = resp.geturl()
        return {"url": final_url}
    except Exception as e:
        return JSONResponse({"error": str(e)}, status_code=400)

if __name__ == "__main__":
    # SHAREPOINT_FILE_URL = "https://molnirre.sharepoint.com/..."
    # response = requests.get(SHAREPOINT_FILE_URL)
    # if response.status_code==200:
    print("Starting FastAPI server...")
    threading.Timer(3, webbrowser.open("http://localhost:8000/spain_new_frontend/")).start()
    uvicorn.run(app, host="127.0.0.1", port=8000)
    # else:
    #     print(response.status_code)
